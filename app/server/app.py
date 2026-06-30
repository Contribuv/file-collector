#!/usr/bin/env python3
"""
文件收集器 - fnOS Native 应用
基于 Flask + SQLite 的多链接文件收集系统

功能：
- 后台生成多个文件收集链接，每个链接有独立通行证
- 支持收集文件描述，告知用户上传内容
- 上传记录管理（查看、删除，删除链接保留文件）
- 响应式 UI，微信浏览器检测提示
"""
import os
import re
import json
import time
import uuid
import sqlite3
import shutil
import secrets
import math
import logging
import warnings
import unicodedata
import traceback
import tempfile
import threading
import urllib.request
import urllib.error
from datetime import datetime, timedelta, timezone
from functools import wraps
from pathlib import Path

from flask import (
    Flask, request, render_template, redirect, url_for,
    session, jsonify, flash, send_from_directory, send_file, abort, make_response, after_this_request
)
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
import bleach
import mimetypes
import zipfile
import hmac
import hashlib
# smtplib / email 延迟导入（仅邮件功能需要，启动时跳过 ~50ms）

# 确保 CSS / JS 等静态资源的 MIME 类型注册正确
# 某些精简环境（如 fnOS）的 mimetypes 数据库可能缺失这些映射
_ESSENTIAL_MIMETYPES = {
    '.css':  'text/css',
    '.js':   'application/javascript',
    '.mjs':  'application/javascript',
    '.json': 'application/json',
    '.svg':  'image/svg+xml',
    '.ico':  'image/vnd.microsoft.icon',
    '.png':  'image/png',
    '.jpg':  'image/jpeg',
    '.jpeg': 'image/jpeg',
    '.gif':  'image/gif',
    '.webp': 'image/webp',
    '.woff': 'font/woff',
    '.woff2':'font/woff2',
    '.ttf':  'font/ttf',
    '.html': 'text/html',
    '.htm':  'text/html',
    '.xml':  'application/xml',
    '.txt':  'text/plain; charset=utf-8',
    '.pdf':  'application/pdf',
}
for _ext, _mime in _ESSENTIAL_MIMETYPES.items():
    mimetypes.add_type(_mime, _ext)

# HEIC 支持（可选依赖，打包时可能不包含）
try:
    from pillow_heif import register_heif_opener
    register_heif_opener()
    HEIC_SUPPORT = True
except ImportError:
    HEIC_SUPPORT = False


# ============================================================
# HTML 压缩 — 去除注释、压缩标签间空白，减小 ~15-30% 响应体积
# ============================================================
import re as _re

_HTML_COMMENT_RE = _re.compile(r'<!--(?!\[if\b).*?-->', _re.DOTALL)
_HTML_TAG_WS_RE  = _re.compile(r'>\s{2,}<')
_HTML_SPACE_RE   = _re.compile(r' {2,}')
_MAX_COMPRESS_BYTES = 512 * 1024  # 超过 512KB 的页面跳过压缩，避免阻塞

def _minify_html(html: str) -> str:
    """压缩 HTML：去掉注释、压缩标签间空白、合并多余空格。
    保护 <pre>/<textarea>/<script>/<style> 内部内容不被篡改。"""
    if len(html) > _MAX_COMPRESS_BYTES:
        return html

    # 1. 保护 <pre>/<textarea>/<script>/<style> 块 — 内容原样保留
    preserved: list[str] = []
    def _save(m: _re.Match) -> str:
        preserved.append(m.group(0))
        return f'\x00P{preserved.__len__()-1}\x00'
    for tag in ('script', 'style', 'pre', 'textarea'):
        html = _re.sub(
            rf'<{tag}[\s>][\s\S]*?</{tag}>',
            _save, html, flags=_re.IGNORECASE
        )

    # 2. 去掉 HTML 注释（保留 IE 条件注释）
    html = _HTML_COMMENT_RE.sub('', html)

    # 3. 压缩标签间连续空白
    html = _HTML_TAG_WS_RE.sub('><', html)

    # 4. 合并文本中多余空格
    html = _HTML_SPACE_RE.sub(' ', html)

    # 5. 去掉首尾空白
    html = html.strip()

    # 6. 还原受保护块
    for i, chunk in enumerate(preserved):
        html = html.replace(f'\x00P{i}\x00', chunk)

    return html


# ============================================================
# 配置 - 适配 fnOS 环境
# ============================================================
VERSION = "2.3.10"

# 模板目录指向 app/server/templates
_TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates')
_STATIC_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')

app = Flask(__name__, template_folder=_TEMPLATE_DIR, static_folder=_STATIC_DIR)
app.config['TEMPLATES_AUTO_RELOAD'] = True  # 开发环境开启，修改模板后自动刷新
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 3600 * 24  # 静态资源缓存 1 天
app.config['MAX_CONTENT_LENGTH'] = 64 * 1024 * 1024 * 1024  # 64GB 硬限制，防止超大文件耗尽磁盘
app.config['MAX_FORM_MEMORY_SIZE'] = 1 * 1024 * 1024  # 超过 1MB 的文件流式写入磁盘，避免内存溢出

# Office 预览模块（基于 flyfish-dev/file-viewer）
from office import office_bp
app.register_blueprint(office_bp)

# 内置反向代理引擎（Go 版本）
import rproxy_manager
from cert_manager import CertManager

RPROXY_PM = rproxy_manager.GoRProxyManager()

# 反向代理支持：修正 request.remote_addr / request.scheme
# x_for=1 信任 1 层反向代理（最常见的 Nginx/Unix Socket 单层反代场景）
# ProxyFix 从 X-Forwarded--For 最右侧取第 1 个值作为 remote_addr
# x_proto=1 信任 X-Forwarded-Proto，确保 request.scheme 正确识别 HTTPS
# x_host=1 信任 X-Forwarded-Host，x_port=1 信任 X-Forwarded-Port
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1, x_prefix=1)

# 会话安全配置
# 注意：SESSION_COOKIE_SECURE 不在 before_request 中动态修改（会导致多线程竞态条件）
# 反向代理场景：ProxyFix 修正 request.scheme 后，Flask 内部会正确处理
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('SESSION_COOKIE_SECURE', '0') == '1'  # HTTPS 部署时设为 1
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)

@app.before_request
def ensure_csrf_token():
    """确保 CSRF token 在 session 中初始化，并保持已登录 session 永久有效"""
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)
    # 已登录用户每请求标记 permanent=True，防止后续 session 修改（如 flash）导致
    # cookie 降级为浏览器会话级而丢失（表现为"等一会儿又要登陆"）
    if session.get('user_id'):
        session.permanent = True

# 上传通行证浏览器缓存有效期默认值（秒）
DEFAULT_PASSCODE_TTL = 7200  # 2小时

def get_passcode_ttl():
    """获取通行证缓存有效期（秒），从数据库设置读取"""
    minutes = get_setting('passcode_ttl_minutes', '120')
    try:
        return int(minutes) * 60
    except (ValueError, TypeError):
        return DEFAULT_PASSCODE_TTL

# 从环境变量读取配置
# 数据库存储：TRIM_PKGVAR（应用私有目录）→ /tmp
# 上传文件存储：UPLOAD_BASE 环境变量（cmd/main 设置为 data-share 子目录）
_PKGVAR = os.environ.get('TRIM_PKGVAR', '/tmp/file-collector')
DATA_DIR = os.environ.get('DATA_DIR', os.path.join(_PKGVAR, 'data'))

PORT = int(os.environ.get('PORT', 5557))

# 上传目录：优先 UPLOAD_BASE 环境变量（cmd/main 设置为 data-share 的 uploads 子目录）
# → 回退到 TRIM_PKGVAR 下的 uploads
_DEFAULT_UPLOAD_BASE = os.environ.get('UPLOAD_BASE',
    os.path.join(_PKGVAR, 'uploads'))
UPLOAD_BASE = _DEFAULT_UPLOAD_BASE  # 启动时默认，后续可能被自定义覆盖

def get_upload_base():
    """获取当前上传目录：自定义路径优先 → 默认路径"""
    custom = get_setting('custom_upload_path', '').strip()
    if custom and os.path.isdir(custom):
        return custom
    return _DEFAULT_UPLOAD_BASE

def refresh_upload_base():
    """刷新全局 UPLOAD_BASE（数据库设置变更后调用）"""
    global UPLOAD_BASE
    UPLOAD_BASE = get_upload_base()

def _migrate_stored_paths(old_base, new_base):
    """迁移数据库中 upload_records.stored_path 从旧基础路径到新基础路径。
    同时处理「新旧路径相同但数据库中仍有旧前缀记录」的边缘情况（如用户已手动迁移文件后再次保存）。"""
    conn = get_db()
    try:
        records = conn.execute("SELECT COUNT(*) as cnt FROM upload_records").fetchone()
        if not records or records['cnt'] == 0:
            return
        new_base_norm = os.path.normpath(new_base) + os.sep
        # 查找第一条不以新 base 开头的记录，推断旧 base
        sample = conn.execute(
            "SELECT stored_path FROM upload_records WHERE stored_path NOT LIKE ? LIMIT 1",
            (new_base_norm + '%',)
        ).fetchone()
        if not sample:
            return  # 所有记录都已匹配，无需迁移
        # 从样本路径提取旧基础路径（upload_records 路径结构：old_base/username/link_id/filename）
        sample_path = sample['stored_path']
        # 取 uploads 目录的父级作为旧 base（兼容 /path/uploads 和 /path 两种格式）
        # 向上找两层：去掉 /username/link_id/filename
        parts = os.path.normpath(sample_path).split(os.sep)
        if len(parts) >= 3:
            old_base_norm = os.sep.join(parts[:-3])
        else:
            old_base_norm = os.path.normpath(old_base)
        if not old_base_norm.endswith(os.sep):
            old_base_norm += os.sep
        if os.path.normpath(old_base_norm.rstrip(os.sep)) == os.path.normpath(new_base_norm.rstrip(os.sep)):
            return
        count_before = conn.execute(
            "SELECT COUNT(*) as cnt FROM upload_records WHERE stored_path LIKE ?",
            (old_base_norm + '%',)
        ).fetchone()['cnt']
        if count_before == 0:
            return
        conn.execute(
            "UPDATE upload_records SET stored_path = REPLACE(stored_path, ?, ?) WHERE stored_path LIKE ?",
            (old_base_norm, new_base_norm, old_base_norm + '%')
        )
        conn.commit()
        actual_changed = conn.total_changes
        flash(f'已自动修复 {actual_changed} 条文件记录路径（{old_base_norm} → {new_base_norm}）。')
    finally:
        conn.close()

DEBUG_MODE = os.environ.get('FLASK_DEBUG', '0') == '1'

# 确保目录存在
os.makedirs(DATA_DIR, exist_ok=True)
try:
    os.makedirs(UPLOAD_BASE, mode=0o755, exist_ok=True)
except PermissionError:
    pass  # 目录已存在但无权限修改，后续会检查可写性

# 数据库路径
DB_PATH = os.path.join(DATA_DIR, 'file_collector.db')

# 日志配置
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger('file_collector')

# 默认设置
DEFAULT_MAX_FILE_SIZE_GB = 1
DEFAULT_MAX_FILES = 10
DEFAULT_ADMIN_USER = 'admin'
DEFAULT_ADMIN_PASS = 'admin123'

# 批量下载限制：总文件大小上限（字节），超过则提示分批下载
MAX_BATCH_DOWNLOAD_SIZE = 500 * 1024 * 1024  # 500MB

# 已压缩格式扩展名，批量打包时跳过重新压缩（ZIP_STORED）
SKIP_COMPRESS_EXTENSIONS = {
    # 图片
    '.jpg', '.jpeg', '.png', '.gif', '.webp', '.heic', '.heif', '.bmp',
    # 视频
    '.mp4', '.mov', '.avi', '.mkv', '.webm', '.flv', '.wmv', '.m4v', '.3gp',
    # 音频
    '.mp3', '.aac', '.ogg', '.wma', '.m4a', '.opus', '.flac', '.wav',
    # 压缩包
    '.zip', '.7z', '.rar', '.gz', '.bz2', '.xz', '.tar',
    # 文档（Office 新版格式本身是 zip 包）
    '.pdf', '.docx', '.xlsx', '.pptx', '.odt', '.ods', '.odp',
}

# ============================================================
# 数据库初始化
# ============================================================
def get_db():
    """获取数据库连接"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn

def init_db():
    """初始化数据库表"""
    conn = get_db()
    
    # 数据库迁移 - 重要：必须先创建users表（其他表依赖它）
    # 1. 检查并创建 users 表（如果不存在）
    try:
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='users'")
        if not cursor.fetchone():
            conn.execute('''
                CREATE TABLE users (
                    id TEXT PRIMARY KEY,
                    username TEXT NOT NULL UNIQUE,
                    password_hash TEXT NOT NULL,
                    email TEXT DEFAULT '',
                    nickname TEXT DEFAULT '',
                    is_admin INTEGER DEFAULT 0,
                    status TEXT DEFAULT 'active',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    last_login_at TIMESTAMP
                )
            ''')
            conn.commit()
            logger.info("已创建 users 表")
    except Exception as e:
        logger.error(f"数据库迁移错误(users): {e}")
    
    # 2. 检查并创建 invite_codes 表（如果不存在）
    try:
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='invite_codes'")
        if not cursor.fetchone():
            conn.execute('''
                CREATE TABLE invite_codes (
                    id TEXT PRIMARY KEY,
                    code TEXT NOT NULL UNIQUE,
                    expires_at TIMESTAMP NOT NULL,
                    used_by TEXT,
                    used_at TIMESTAMP,
                    created_by TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            conn.commit()
            logger.info("已创建 invite_codes 表")
        # 索引始终尝试创建（升级场景下表可能已由 upgrade_callback 创建）
        conn.execute("CREATE INDEX IF NOT EXISTS idx_invite_codes_code ON invite_codes(code)")
        conn.commit()
    except Exception as e:
        logger.error(f"数据库迁移错误(invite_codes): {e}")
    
    # 3. 创建其他基础表（使用单独的CREATE TABLE语句，避免外键约束问题）
    try:
        # 3.1 创建 settings 表
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='settings'")
        if not cursor.fetchone():
            conn.execute('''
                CREATE TABLE settings (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL
                )
            ''')
            conn.commit()
            logger.info("已创建 settings 表")
    except Exception as e:
        logger.error(f"数据库迁移错误(settings): {e}")

    try:
        # 3.2 创建 links 表（不使用外键约束，应用层处理关联）
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='links'")
        if not cursor.fetchone():
            conn.execute('''
                CREATE TABLE links (
                    id TEXT PRIMARY KEY,
                    user_id TEXT NOT NULL DEFAULT '',
                    title TEXT NOT NULL,
                    description TEXT DEFAULT '',
                    passcode TEXT NOT NULL,
                    max_file_size_gb REAL DEFAULT 1,
                    max_files INTEGER DEFAULT 10,
                    target_folder TEXT DEFAULT '',
                    folder_name TEXT DEFAULT '',
                    status TEXT DEFAULT 'active',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            conn.commit()
            logger.info("已创建 links 表")
        # 索引始终尝试创建
        conn.execute("CREATE INDEX IF NOT EXISTS idx_links_user_id ON links(user_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_links_created_at ON links(created_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_links_user_created ON links(user_id, created_at)")
        conn.commit()
    except Exception as e:
        logger.error(f"数据库迁移错误(links): {e}")

    try:
        # 3.3 创建 upload_records 表（不使用外键约束，应用层处理关联）
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='upload_records'")
        if not cursor.fetchone():
            conn.execute('''
                CREATE TABLE upload_records (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id TEXT NOT NULL DEFAULT '',
                    link_id TEXT NOT NULL,
                    original_name TEXT NOT NULL,
                    stored_name TEXT NOT NULL,
                    stored_path TEXT NOT NULL,
                    file_size INTEGER NOT NULL,
                    file_size_display TEXT DEFAULT '',
                    uploader_ip TEXT DEFAULT '',
                    uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    download_count INTEGER DEFAULT 0
                )
            ''')
            conn.commit()
            logger.info("已创建 upload_records 表")
        # 索引始终尝试创建
        conn.execute("CREATE INDEX IF NOT EXISTS idx_records_link_id ON upload_records(link_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_records_uploaded_at ON upload_records(uploaded_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_records_user_id ON upload_records(user_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_records_link_uploaded ON upload_records(link_id, uploaded_at)")
        conn.commit()
    except Exception as e:
        logger.error(f"数据库迁移错误(upload_records): {e}")

    try:
        # 3.3.1 创建 chunk_uploads 表（分片上传断点续传支持）
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='chunk_uploads'")
        if not cursor.fetchone():
            conn.execute('''
                CREATE TABLE chunk_uploads (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    upload_id TEXT NOT NULL UNIQUE,
                    link_id TEXT NOT NULL,
                    user_id TEXT DEFAULT '',
                    original_name TEXT NOT NULL,
                    stored_name TEXT NOT NULL,
                    total_size INTEGER NOT NULL,
                    chunk_size INTEGER NOT NULL,
                    total_chunks INTEGER NOT NULL,
                    uploaded_chunks TEXT DEFAULT '',
                    status TEXT DEFAULT 'uploading',
                    stored_path TEXT DEFAULT '',
                    uploader_ip TEXT DEFAULT '',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            conn.commit()
            logger.info("已创建 chunk_uploads 表")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_chunk_uploads_upload_id ON chunk_uploads(upload_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_chunk_uploads_link_id ON chunk_uploads(link_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_chunk_uploads_status ON chunk_uploads(status)")
        conn.commit()
    except Exception as e:
        logger.error(f"数据库迁移错误(chunk_uploads): {e}")

    try:
        # 3.3.2 创建 tus_uploads 表（Tus 协议断点续传）
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='tus_uploads'")
        if not cursor.fetchone():
            conn.execute('''
                CREATE TABLE tus_uploads (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    upload_id TEXT NOT NULL UNIQUE,
                    link_id TEXT NOT NULL,
                    user_id TEXT DEFAULT '',
                    original_name TEXT NOT NULL,
                    stored_name TEXT NOT NULL,
                    total_size INTEGER NOT NULL,
                    uploaded_offset INTEGER DEFAULT 0,
                    status TEXT DEFAULT 'uploading',
                    stored_path TEXT DEFAULT '',
                    temp_path TEXT DEFAULT '',
                    uploader_ip TEXT DEFAULT '',
                    uploader_name TEXT DEFAULT '',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            conn.commit()
            logger.info("已创建 tus_uploads 表")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_tus_uploads_upload_id ON tus_uploads(upload_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_tus_uploads_link_id ON tus_uploads(link_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_tus_uploads_status ON tus_uploads(status)")
        conn.commit()
    except Exception as e:
        logger.error(f"数据库迁移错误(tus_uploads): {e}")

    # 数据库迁移 - 字段升级
    # 0. 检查并创建 verification_codes 表（如果不存在）
    try:
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='verification_codes'")
        if not cursor.fetchone():
            conn.execute('''
                CREATE TABLE verification_codes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    email TEXT NOT NULL,
                    code TEXT NOT NULL,
                    purpose TEXT DEFAULT 'reset_password',
                    expires_at TIMESTAMP NOT NULL,
                    used INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            conn.commit()
            logger.info("已创建 verification_codes 表")
        # 索引始终尝试创建
        conn.execute("CREATE INDEX IF NOT EXISTS idx_vcodes_email ON verification_codes(email)")
        conn.commit()
    except Exception as e:
        logger.error(f"数据库迁移错误(verification_codes): {e}")

    # 3.4 检查并创建 rate_limits 表（基于 SQLite 的频率限制，支持多 worker）
    try:
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='rate_limits'")
        if not cursor.fetchone():
            conn.execute('''
                CREATE TABLE rate_limits (
                    key TEXT PRIMARY KEY,
                    attempts INTEGER DEFAULT 1,
                    first_attempt REAL NOT NULL
                )
            ''')
            conn.commit()
            logger.info("已创建 rate_limits 表")
    except Exception as e:
        logger.error(f"数据库迁移错误(rate_limits): {e}")

    # 4. 检查并创建 user_settings 表（用户级设置）
    try:
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='user_settings'")
        if not cursor.fetchone():
            conn.execute('''
                CREATE TABLE user_settings (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id TEXT NOT NULL,
                    key TEXT NOT NULL,
                    value TEXT DEFAULT '',
                    UNIQUE(user_id, key)
                )
            ''')
            conn.commit()
            logger.info("已创建 user_settings 表")
        # 索引始终尝试创建
        conn.execute("CREATE INDEX IF NOT EXISTS idx_user_settings_user_id ON user_settings(user_id)")
        conn.commit()
    except Exception as e:
        logger.error(f"数据库迁移错误(user_settings): {e}")

    # 数据库迁移 - 字段升级（links/upload_records 等）
    # 1. 检查并迁移 links 表添加 user_id 字段
    try:
        cursor = conn.execute("PRAGMA table_info(links)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'user_id' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN user_id TEXT DEFAULT ''")
            conn.commit()
    except Exception as e:
        logger.error(f"数据库迁移错误(user_id links): {e}")

    # 2. 检查并迁移 upload_records 表添加 user_id 字段
    try:
        cursor = conn.execute("PRAGMA table_info(upload_records)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'user_id' not in columns:
            conn.execute("ALTER TABLE upload_records ADD COLUMN user_id TEXT DEFAULT ''")
            conn.commit()
    except Exception as e:
        logger.error(f"数据库迁移错误(user_id upload_records): {e}")

    # 3. 原有迁移保持不变
    try:
        cursor = conn.execute("PRAGMA table_info(links)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'max_file_size_mb' in columns and 'max_file_size_gb' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN max_file_size_gb REAL DEFAULT 1")
            conn.execute("UPDATE links SET max_file_size_gb = max_file_size_mb / 1024.0 WHERE max_file_size_mb IS NOT NULL AND max_file_size_gb = 1")
            conn.commit()
    except Exception as e:
        logger.error(f"数据库迁移错误: {e}")

    try:
        cursor = conn.execute("PRAGMA table_info(upload_records)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'download_count' not in columns:
            conn.execute("ALTER TABLE upload_records ADD COLUMN download_count INTEGER DEFAULT 0")
            conn.commit()
    except Exception as e:
        logger.error(f"数据库迁移错误(download_count): {e}")

    # 添加 source 字段（标识文件来源）
    try:
        cursor = conn.execute("PRAGMA table_info(upload_records)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'source' not in columns:
            conn.execute("ALTER TABLE upload_records ADD COLUMN source TEXT DEFAULT 'upload'")
            conn.commit()
            logger.info("已为 upload_records 表添加 source 字段")
    except Exception as e:
        logger.error(f"数据库迁移错误(source): {e}")

    # 添加 passcode_empty 字段（标记空通行证，避免每次请求 check_password_hash）
    try:
        cursor = conn.execute("PRAGMA table_info(links)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'passcode_empty' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN passcode_empty INTEGER DEFAULT 0")
            conn.commit()
            # 迁移已有数据：对 passcode 为空或空哈希的标记为 1
            conn.execute(
                "UPDATE links SET passcode_empty = 1 WHERE passcode = '' OR passcode IS NULL"
            )
            conn.commit()
            logger.info("已为 links 表添加 passcode_empty 字段")
    except Exception as e:
        logger.error(f"数据库迁移错误(passcode_empty): {e}")

    # 规范化 uploaded_at 时间格式（去除微秒）
    try:
        conn.execute(
            """UPDATE upload_records SET uploaded_at = 
               substr(uploaded_at, 1, 19) 
               WHERE uploaded_at LIKE '%.%'"""
        )
        conn.commit()
    except Exception as e:
        logger.error(f"数据库迁移错误(uploaded_at normalize): {e}")

    # 清理孤儿记录（实际文件已被删除）
    try:
        orphan_records = conn.execute(
            "SELECT id, stored_path FROM upload_records"
        ).fetchall()
        deleted_count = 0
        for r in orphan_records:
            real_path = os.path.realpath(r['stored_path'])
            real_base = os.path.realpath(get_upload_base())
            if (real_path.startswith(real_base + os.sep) or real_path == real_base):
                if not os.path.exists(real_path) or not os.path.isfile(real_path):
                    conn.execute("DELETE FROM upload_records WHERE id = ?", (r['id'],))
                    deleted_count += 1
        if deleted_count > 0:
            conn.commit()
            logger.info(f"已清理 {deleted_count} 条孤儿记录（文件已不存在）")
    except Exception as e:
        logger.error(f"数据库迁移错误(orphan cleanup): {e}")

    try:
        cursor = conn.execute("PRAGMA table_info(links)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'expires_at' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN expires_at TIMESTAMP")
            conn.commit()
    except Exception as e:
        logger.error(f"数据库迁移错误(expires_at): {e}")

    try:
        cursor = conn.execute("PRAGMA table_info(links)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'allow_delete' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN allow_delete INTEGER DEFAULT 0")
            conn.commit()
        if 'allow_preview_download' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN allow_preview_download INTEGER DEFAULT 0")
            conn.commit()
    except Exception as e:
        logger.error(f"数据库迁移错误(allow_delete): {e}")

    try:
        cursor = conn.execute("PRAGMA table_info(links)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'passcode_plain' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN passcode_plain TEXT DEFAULT ''")
            conn.commit()
    except Exception as e:
        logger.error(f"数据库迁移错误(passcode_plain): {e}")

    # 新增分享页开关和独立通行证字段
    try:
        cursor = conn.execute("PRAGMA table_info(links)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'share_enabled' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN share_enabled INTEGER DEFAULT 0")
            conn.commit()
            logger.info("已为 links 表添加 share_enabled 字段")
    except Exception as e:
        logger.error(f"数据库迁移错误(share_enabled): {e}")
    try:
        cursor = conn.execute("PRAGMA table_info(links)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'share_passcode' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN share_passcode TEXT DEFAULT ''")
            conn.commit()
            logger.info("已为 links 表添加 share_passcode 字段")
    except Exception as e:
        logger.error(f"数据库迁移错误(share_passcode): {e}")
    try:
        cursor = conn.execute("PRAGMA table_info(links)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'share_passcode_empty' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN share_passcode_empty INTEGER DEFAULT 0")
            conn.commit()
            logger.info("已为 links 表添加 share_passcode_empty 字段")
    except Exception as e:
        logger.error(f"数据库迁移错误(share_passcode_empty): {e}")

    try:
        cursor = conn.execute("PRAGMA table_info(links)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'share_passcode_plain' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN share_passcode_plain TEXT DEFAULT ''")
            conn.commit()
            logger.info("已为 links 表添加 share_passcode_plain 字段")
    except Exception as e:
        logger.error(f"数据库迁移错误(share_passcode_plain): {e}")

    # 分享页独立描述、独立有效期、收集开关
    try:
        cursor = conn.execute("PRAGMA table_info(links)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'share_description' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN share_description TEXT DEFAULT ''")
            conn.commit()
            logger.info("已为 links 表添加 share_description 字段")
    except Exception as e:
        logger.error(f"数据库迁移错误(share_description): {e}")
    try:
        cursor = conn.execute("PRAGMA table_info(links)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'share_expires_at' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN share_expires_at TIMESTAMP")
            conn.commit()
            logger.info("已为 links 表添加 share_expires_at 字段")
    except Exception as e:
        logger.error(f"数据库迁移错误(share_expires_at): {e}")
    try:
        cursor = conn.execute("PRAGMA table_info(links)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'collect_enabled' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN collect_enabled INTEGER DEFAULT 1")
            conn.commit()
            logger.info("已为 links 表添加 collect_enabled 字段")
    except Exception as e:
        logger.error(f"数据库迁移错误(collect_enabled): {e}")

    # 为 links 表添加 folder_name 列（用于文件系统文件夹，基于收集名称生成）
    try:
        cursor = conn.execute("PRAGMA table_info(links)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'folder_name' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN folder_name TEXT DEFAULT ''")
            conn.commit()
            logger.info("已为 links 表添加 folder_name 字段")
            # 迁移旧数据：folder_name 默认为 link_id
            conn.execute("UPDATE links SET folder_name = id WHERE folder_name = '' OR folder_name IS NULL")
            conn.commit()
            logger.info("已迁移旧链接的 folder_name 为 link_id")
    except Exception as e:
        logger.error(f"数据库迁移错误(folder_name): {e}")

    # 为 links 表添加 require_uploader 列（空通行证时是否要求上传者填写身份）
    try:
        cursor = conn.execute("PRAGMA table_info(links)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'require_uploader' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN require_uploader INTEGER DEFAULT 0")
            conn.commit()
            logger.info("已为 links 表添加 require_uploader 字段")
    except Exception as e:
        logger.error(f"数据库迁移错误(require_uploader): {e}")

    # 为 links 表添加 collect_slug 和 share_slug 列（自定义链接 ID）
    try:
        cursor = conn.execute("PRAGMA table_info(links)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'collect_slug' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN collect_slug TEXT DEFAULT ''")
            conn.commit()
            logger.info("已为 links 表添加 collect_slug 字段")
        if 'share_slug' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN share_slug TEXT DEFAULT ''")
            conn.commit()
            logger.info("已为 links 表添加 share_slug 字段")
    except Exception as e:
        logger.error(f"数据库迁移错误(collect_slug/share_slug): {e}")

    # 为 upload_records 表添加 uploader_name 列（记录上传者身份）
    try:
        cursor = conn.execute("PRAGMA table_info(upload_records)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'uploader_name' not in columns:
            conn.execute("ALTER TABLE upload_records ADD COLUMN uploader_name TEXT DEFAULT ''")
            conn.commit()
            logger.info("已为 upload_records 表添加 uploader_name 字段")
    except Exception as e:
        logger.error(f"数据库迁移错误(upload_records.uploader_name): {e}")

    # 为 users 表添加 email 列（旧数据库可能缺失）
    try:
        cursor = conn.execute("PRAGMA table_info(users)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'email' not in columns:
            conn.execute("ALTER TABLE users ADD COLUMN email TEXT DEFAULT ''")
            conn.commit()
            logger.info("已为 users 表添加 email 字段")
    except Exception as e:
        logger.error(f"数据库迁移错误(users email): {e}")

    # 为 users 表添加 nickname 列
    try:
        cursor = conn.execute("PRAGMA table_info(users)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'nickname' not in columns:
            conn.execute("ALTER TABLE users ADD COLUMN nickname TEXT DEFAULT ''")
            conn.commit()
            logger.info("已为 users 表添加 nickname 字段")
    except Exception as e:
        logger.error(f"数据库迁移错误(users nickname): {e}")



    # 下载日志表
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS download_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                record_id INTEGER NOT NULL,
                downloader_ip TEXT DEFAULT '',
                downloaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                source TEXT DEFAULT 'admin',
                user_agent TEXT DEFAULT ''
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_download_logs_record_id ON download_logs(record_id)")
        conn.commit()
        logger.info("download_logs 表检查/创建完成")
    except Exception as e:
        logger.error(f"数据库迁移错误(download_logs): {e}")

    # 上传日志表
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS upload_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                record_id INTEGER,
                link_id TEXT NOT NULL DEFAULT '',
                uploader_ip TEXT DEFAULT '',
                event TEXT NOT NULL DEFAULT '',
                event_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                details TEXT DEFAULT '',
                success INTEGER DEFAULT 1,
                uploader_name TEXT DEFAULT ''
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_upload_logs_record_id ON upload_logs(record_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_upload_logs_link_id ON upload_logs(link_id)")
        conn.commit()
        logger.info("upload_logs 表检查/创建完成")
    except Exception as e:
        logger.error(f"数据库迁移错误(upload_logs): {e}")

    # 为 upload_logs 表添加 uploader_name 列（仅对旧表做迁移 + 回填）
    try:
        cursor = conn.execute("PRAGMA table_info(upload_logs)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'uploader_name' not in columns:
            conn.execute("ALTER TABLE upload_logs ADD COLUMN uploader_name TEXT DEFAULT ''")
            conn.commit()
            logger.info("已为 upload_logs 表添加 uploader_name 字段")
            # 回填已有日志的 uploader_name
            conn.execute("""
                UPDATE upload_logs SET uploader_name = (
                    SELECT ur.uploader_name FROM upload_records ur WHERE ur.id = upload_logs.record_id
                ) WHERE record_id IS NOT NULL AND uploader_name = ''
            """)
            conn.commit()
            logger.info("已回填 upload_logs 的 uploader_name 字段")
    except Exception as e:
        logger.error(f"数据库迁移错误(upload_logs.uploader_name): {e}")

    # 为 links 表添加附件字段（老师上传附件供学生下载）
    try:
        cursor = conn.execute("PRAGMA table_info(links)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'attachment_name' not in columns:
            conn.execute("ALTER TABLE links ADD COLUMN attachment_name TEXT DEFAULT ''")
            conn.execute("ALTER TABLE links ADD COLUMN attachment_path TEXT DEFAULT ''")
            conn.execute("ALTER TABLE links ADD COLUMN attachment_size INTEGER DEFAULT 0")
            conn.commit()
            logger.info("已为 links 表添加附件字段(attachment_name/path/size)")
    except Exception as e:
        logger.error(f"数据库迁移错误(links attachment): {e}")

    # 检测是否已有数据库（升级场景）
    existing_admin = conn.execute(
        "SELECT value FROM settings WHERE key = 'admin_username'"
    ).fetchone()

    # 检查是否已有用户表数据（先检查表是否存在）
    existing_users = 0
    try:
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='users'")
        if cursor.fetchone():
            existing_users = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    except Exception as e:
        logger.error(f"检查用户表失败: {e}")

    if existing_admin:
        # 升级安装：数据库已存在，保留所有已有设置，忽略 wizard 环境变量
        logger.info("检测到已有数据库，保留所有数据（升级安装）")
        
        # 创建默认管理员用户（如果不存在）
        if existing_users == 0:
            admin_user = existing_admin['value']
            # 兼容老版本：admin_password_hash 可能不存在
            admin_pass_row = conn.execute(
                "SELECT value FROM settings WHERE key = 'admin_password_hash'"
            ).fetchone()
            if admin_pass_row:
                admin_pass_hash = admin_pass_row['value']
            else:
                # 老版本无该字段，用默认密码生成哈希
                admin_pass_hash = generate_password_hash(DEFAULT_ADMIN_PASS)
                conn.execute(
                    "INSERT OR REPLACE INTO settings (key, value) VALUES ('admin_password_hash', ?)",
                    (admin_pass_hash,)
                )
                logger.info("老版本数据库：自动生成 admin_password_hash（默认密码）")
            
            admin_id = str(uuid.uuid4())
            conn.execute('''
                INSERT INTO users (id, username, password_hash, is_admin, status)
                VALUES (?, ?, ?, 1, 'active')
            ''', (admin_id, admin_user, admin_pass_hash))
            
            # 更新旧数据的 user_id
            conn.execute("UPDATE links SET user_id = ? WHERE user_id = ''", (admin_id,))
            conn.execute("UPDATE upload_records SET user_id = ? WHERE user_id = ''", (admin_id,))
            
            logger.info(f"升级安装：创建管理员用户 {admin_user}，迁移旧数据")
        
        # 仅补充可能缺失的新增设置项（不影响已有数据）
        new_defaults = {
            'passcode_ttl_minutes': '120',
            'landing_page_enabled': '1',
            'collect_footer_text': '',
            'share_page_title': '',
            'share_footer_text': '',
            'blocked_extensions': '',
            'allow_registration': '0',
            'default_invite_expire_days': '7',
            'default_link_expire_days': '30',
            'smtp_host': '',
            'smtp_port': '587',
            'smtp_username': '',
            'smtp_password': '',
            'smtp_use_tls': '1',
            'smtp_from_email': '',
            'smtp_from_name': '文件收集器',
        }
        for key, val in new_defaults.items():
            conn.execute(
                "INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)",
                (key, val)
            )
    else:
        # 首次安装：使用 wizard 环境变量或默认值初始化
        wizard_admin_user = os.environ.get('wizard_admin_user', '').strip()
        wizard_admin_pass = os.environ.get('wizard_admin_pass', '').strip()
        init_admin_user = wizard_admin_user if wizard_admin_user else DEFAULT_ADMIN_USER
        init_admin_pass = wizard_admin_pass if wizard_admin_pass else DEFAULT_ADMIN_PASS
        init_login_tip = '默认账户 admin / admin123，请及时修改' if not wizard_admin_pass else '账户已由安装向导设置'

        # 创建管理员用户
        admin_id = str(uuid.uuid4())
        conn.execute('''
            INSERT INTO users (id, username, password_hash, is_admin, status)
            VALUES (?, ?, ?, 1, 'active')
        ''', (admin_id, init_admin_user, generate_password_hash(init_admin_pass)))

        defaults = {
            'admin_username': init_admin_user,
            'admin_password_hash': generate_password_hash(init_admin_pass),
            'max_file_size_gb': str(DEFAULT_MAX_FILE_SIZE_GB),
            'max_files': str(DEFAULT_MAX_FILES),
            'site_title': '文件收集器',
            'secret_key': secrets.token_hex(32),
            'login_tip': init_login_tip,
            'collect_footer_text': '',
            'share_page_title': '',
            'share_footer_text': '',
            'passcode_ttl_minutes': '120',
            'landing_page_enabled': '1',
            'allow_registration': '0',
            'default_invite_expire_days': '7',
            'default_link_expire_days': '30',
            'links_per_page': '50',
            'smtp_host': '',
            'smtp_port': '587',
            'smtp_username': '',
            'smtp_password': '',
            'smtp_use_tls': '1',
            'smtp_from_email': '',
            'smtp_from_name': '文件收集器',
        }
        for key, val in defaults.items():
            conn.execute(
                "INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)",
                (key, val)
            )
    conn.commit()
    conn.close()

def get_setting(key, default=None):
    """获取单个设置值"""
    conn = get_db()
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    conn.close()
    return row['value'] if row else default

def set_setting(key, value):
    """设置单个值"""
    conn = get_db()
    conn.execute(
        "INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)",
        (key, value)
    )
    conn.commit()
    conn.close()

def get_user_setting(user_id, key, default=None):
    """获取用户级设置，优先用户设置，回退到全局设置"""
    conn = get_db()
    row = conn.execute(
        "SELECT value FROM user_settings WHERE user_id = ? AND key = ?",
        (str(user_id), key)
    ).fetchone()
    conn.close()
    if row:
        return row['value']
    return get_setting(key, default)

def set_user_setting(user_id, key, value):
    """设置用户级设置"""
    conn = get_db()
    conn.execute(
        "INSERT OR REPLACE INTO user_settings (user_id, key, value) VALUES (?, ?, ?)",
        (str(user_id), key, str(value))
    )
    conn.commit()
    conn.close()

def get_upload_batch_limit(user_id=None):
    """获取单次上传个数限制，优先用户级设置，回退全局设置，默认 30"""
    batch = get_user_setting(user_id, 'upload_batch_limit', '30') if user_id else get_setting('upload_batch_limit', '30')
    try:
        batch = int(batch)
    except (ValueError, TypeError):
        batch = 30
    return max(1, min(batch, 100))

# 初始化
init_db()
refresh_upload_base()  # 从数据库恢复自定义上传路径（如果有）
_key = get_setting('secret_key', None)
if _key is None:
    _key = secrets.token_hex(32)
    set_setting('secret_key', _key)
app.secret_key = _key

# 预编译所有模板，消除首次请求时的编译延迟（TTFB 从 ~3s 降至 ~50ms）
_TEMPLATE_NAMES = [
    'base.html', '_admin_sidebar.html',
    'landing.html', 'collect.html', 'share.html', 'error.html',
    'admin_login.html', 'admin_dashboard.html', 'admin_links.html',
    'admin_records.html', 'admin_settings.html', 'admin_users.html',
    'admin_invite_codes.html', 'admin_user_settings.html',
    'register.html', 'forgot_password.html', 'reset_password.html',
]
for _tn in _TEMPLATE_NAMES:
    try:
        app.jinja_env.get_or_select_template(_tn)
    except Exception:
        pass  # 模板可能不存在或依赖上下文变量，忽略编译错误
del _tn, _TEMPLATE_NAMES

# ============================================================
# 频率限制 & CSRF 保护
# ============================================================
_rate_limit_calls = 0

def rate_limit(key, max_attempts=5, window_seconds=60):
    """基于 SQLite 的频率限制（支持多 worker 部署）
    
    将限流计数存储在数据库中，确保 gunicorn 多 worker 场景下计数共享。
    """
    global _rate_limit_calls
    now = time.time()
    _rate_limit_calls += 1
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT attempts, first_attempt FROM rate_limits WHERE key = ?", (key,)
        ).fetchone()
        if row:
            attempts = row['attempts']
            first = row['first_attempt']
            if now - first > window_seconds:
                # 窗口过期，重置计数
                conn.execute(
                    "UPDATE rate_limits SET attempts = 1, first_attempt = ? WHERE key = ?",
                    (now, key)
                )
                conn.commit()
            elif attempts >= max_attempts:
                return False
            else:
                conn.execute(
                    "UPDATE rate_limits SET attempts = attempts + 1 WHERE key = ?",
                    (key,)
                )
                conn.commit()
        else:
            conn.execute(
                "INSERT INTO rate_limits (key, attempts, first_attempt) VALUES (?, 1, ?)",
                (key, now)
            )
            conn.commit()
        # 每 100 次调用清理一次过期条目（防止表膨胀）
        if _rate_limit_calls % 100 == 0:
            conn.execute(
                "DELETE FROM rate_limits WHERE ? - first_attempt > ?",
                (now, max(window_seconds, 3600))
            )
            conn.commit()
        return True
    except Exception as e:
        logger.warning(f"rate_limit 异常，放行请求: {e}")
        return True
    finally:
        conn.close()

def generate_csrf_token():
    """生成 CSRF token"""
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)
    return session['csrf_token']

def validate_csrf():
    """验证 CSRF token"""
    token = request.form.get('csrf_token', '') or request.headers.get('X-CSRFToken', '')
    expected = session.get('csrf_token', '')
    if not token:
        flash('安全验证失败：缺少验证令牌，请刷新页面重试')
        return False
    if not expected or not secrets.compare_digest(token, expected):
        flash('安全验证失败，请刷新页面重试')
        return False
    return True

# ============================================================
# 辅助函数
# ============================================================
def generate_link_id(title=None):
    """生成 8 位纯字母数字短链接ID"""
    return uuid.uuid4().hex[:8]

def _resolve_collect_id(raw_id, conn=None):
    """解析收集页链接 ID：优先 collect_slug，回退 id"""
    close_conn = conn is None
    if close_conn:
        conn = get_db()
    try:
        link = conn.execute(
            "SELECT * FROM links WHERE (collect_slug = ? OR id = ?) AND status = 'active'",
            (raw_id, raw_id)
        ).fetchone()
        return link
    finally:
        if close_conn:
            conn.close()

def _resolve_share_id(raw_id, conn=None):
    """解析分享页链接 ID：优先 share_slug，回退 id"""
    close_conn = conn is None
    if close_conn:
        conn = get_db()
    try:
        link = conn.execute(
            "SELECT * FROM links WHERE (share_slug = ? OR id = ?) AND status = 'active'",
            (raw_id, raw_id)
        ).fetchone()
        return link
    finally:
        if close_conn:
            conn.close()

def _check_slug_unique(slug, exclude_id=None):
    """检查自定义 slug 是否全局唯一（与所有 id、collect_slug、share_slug 不冲突）"""
    if not slug:
        return True
    conn = get_db()
    try:
        # 检查是否与主键 id 冲突
        row = conn.execute("SELECT id FROM links WHERE id = ?", (slug,)).fetchone()
        if row and row['id'] != exclude_id:
            return False
        # 检查是否与 collect_slug 冲突
        row = conn.execute(
            "SELECT id FROM links WHERE collect_slug = ?", (slug,)
        ).fetchone()
        if row and row['id'] != exclude_id:
            return False
        # 检查是否与 share_slug 冲突
        row = conn.execute(
            "SELECT id FROM links WHERE share_slug = ?", (slug,)
        ).fetchone()
        if row and row['id'] != exclude_id:
            return False
        return True
    finally:
        conn.close()

def format_file_size(size_bytes):
    """格式化文件大小"""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024:
            return f"{size_bytes:.1f} {unit}"
        size_bytes /= 1024
    return f"{size_bytes:.1f} TB"


def _get_zip_compress_type(filename, skip_set=SKIP_COMPRESS_EXTENSIONS):
    """根据扩展名决定 ZIP 压缩方式，已压缩格式跳过重复压缩"""
    ext = os.path.splitext(filename)[1].lower()
    return zipfile.ZIP_STORED if ext in skip_set else zipfile.ZIP_DEFLATED

def _get_client_ip():
    """获取客户端真实IP
    
    支持多种反向代理场景：
    - Nginx HTTP 反向代理（ProxyFix 修正 remote_addr）
    - Unix Socket 反向代理（remote_addr 为 127.0.0.1 或特殊值时回退到 header）
    
    ProxyFix(x_for=1) 已修正 remote_addr 为 X-Forwarded-For 最右侧值。
    但在 Unix Socket 或未配置 ProxyFix 信任链时，remote_addr 可能是 127.0.0.1，
    此时回退到 X-Real-IP / X-Forwarded-For 等 header。
    """
    ip = (request.remote_addr or '').strip()
    # remote_addr 有效且非本地回环 → 直接用
    if ip and ip != 'unknown' and not ip.startswith('unix') and ip not in ('127.0.0.1', '::1'):
        return ip
    
    # 回退：remote_addr 是本地回环或 Unix Socket → 从 header 取真实 IP
    # 优先 X-Real-IP（Nginx 常用，代理服务器设置，客户端无法伪造）
    x_real_ip = request.headers.get('X-Real-IP', '').strip()
    if x_real_ip and x_real_ip != 'unknown' and x_real_ip not in ('127.0.0.1', '::1'):
        return x_real_ip
    
    # X-Forwarded-For（取第一个值，即最原始的客户端 IP）
    x_forwarded_for = request.headers.get('X-Forwarded-For', '')
    if x_forwarded_for:
        first_ip = x_forwarded_for.split(',')[0].strip()
        if first_ip and first_ip != 'unknown' and first_ip not in ('127.0.0.1', '::1'):
            return first_ip
    
    # 其他可能的 header
    for header in ['X-Client-IP', 'X-Forwarded', 'Forwarded-For', 'Forwarded']:
        val = request.headers.get(header, '').strip()
        if val and val != 'unknown':
            if header.lower() == 'forwarded':
                for part in val.split(';'):
                    part = part.strip()
                    if part.lower().startswith('for='):
                        ip_val = part[4:].strip()
                        if ip_val.startswith(('"', "'")):
                            ip_val = ip_val[1:-1]
                        return ip_val
            else:
                return val
    
    # 所有方法都失败，返回 remote_addr 原始值（可能是 127.0.0.1）
    return ip if ip else 'unknown'
    # 如果都没有，返回一个标识性值
    return 'unknown'

def _log_download(record_id, source='admin'):
    """记录下载日志"""
    try:
        conn = get_db()
        ua = (request.headers.get('User-Agent', '') or '')[:512]
        conn.execute(
            """INSERT INTO download_logs (record_id, downloader_ip, downloaded_at, source, user_agent)
               VALUES (?, ?, ?, ?, ?)""",
            (record_id, _get_client_ip(), datetime.now().strftime('%Y-%m-%d %H:%M:%S'), source, ua)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        logger.error(f"下载日志记录失败: {e}")

def _log_upload(link_id, event, record_id=None, details=None, success=1, uploader_name=''):
    """记录上传日志"""
    try:
        conn = get_db()
        detail_str = json.dumps(details, ensure_ascii=False) if isinstance(details, dict) else (details or '')
        conn.execute(
            """INSERT INTO upload_logs (record_id, link_id, uploader_ip, event, event_time, details, success, uploader_name)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (record_id, link_id, _get_client_ip(), event,
             datetime.now().strftime('%Y-%m-%d %H:%M:%S'), detail_str, success, uploader_name)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        logger.error(f"上传日志记录失败: {e}")

def is_wechat_browser():
    """检测是否来自微信浏览器"""
    ua = request.headers.get('User-Agent', '').lower()
    return 'micromessenger' in ua

def is_verified(link_id, link=None):
    """检查当前 session 中通行证是否已验证且未过期
    如果传入 link 对象且为空通行证，直接返回 True
    """
    if link is not None:
        pp = link['passcode_plain']
        if not pp or not pp.strip():
            return True
    ts = session.get(f'verified_{link_id}', 0)
    return isinstance(ts, (int, float)) and (time.time() - ts) < get_passcode_ttl()

def is_share_verified(link_id, link=None):
    """检查分享页独立通行证是否已验证（share_passcode 优先，fallback 到收集页通行证）"""
    if link is not None:
        link = dict(link)  # 兼容 sqlite3.Row 传入
        _sp = link.get('share_passcode', '')
        _spe = link.get('share_passcode_empty', 0)
        if _spe:
            return True
        if _sp and _sp.strip():
            has_pass = True
        else:
            # fallback 到收集页通行证
            has_pass = bool(link.get('passcode_plain', '') and link['passcode_plain'].strip())
        if not has_pass:
            return True
    ts = session.get(f'share_verified_{link_id}', 0)
    return isinstance(ts, (int, float)) and (time.time() - ts) < get_passcode_ttl()

def link_has_passcode(link_id):
    """检查链接是否设置了通行证（passcode_plain 非空）"""
    conn = get_db()
    row = conn.execute("SELECT passcode_plain FROM links WHERE id = ?", (link_id,)).fetchone()
    conn.close()
    if not row:
        return False
    return bool(row['passcode_plain'] and row['passcode_plain'].strip())

# ============================================================
# 下载令牌（防盗链防护）— 仅对公开链接（无通行证）生效
# ============================================================

_DL_TOKEN_TTL = 900  # 令牌有效期 15 分钟


def _parse_expire_input(days_str, unit_str, max_days, current_time=None):
    """将用户输入的天数+单位转换为 expires_at 字符串。
    返回 (expires_at_str_or_None, error_msg_or_None)"""
    if not days_str or not days_str.strip():
        return None, None
    try:
        value = int(days_str.strip())
    except (ValueError, TypeError):
        return None, '请输入有效的整数'
    if value <= 0:
        return None, None
    unit = (unit_str or '天').strip()
    if unit == '分钟':
        total_minutes = value
    elif unit == '小时':
        total_minutes = value * 60
    else:
        total_minutes = value * 1440
    max_minutes = max_days * 1440
    if total_minutes > max_minutes:
        return None, f'有效期不能超过 {max_days} 天'
    now = current_time or datetime.now()
    expires_at = (now + timedelta(minutes=total_minutes)).strftime('%Y-%m-%dT%H:%M')
    return expires_at, None


def _format_expire_for_edit(expires_at_str):
    """根据 expires_at 反算编辑表单的值+单位+展示字符串。
    返回 (value_str, unit_str, display_str)"""
    if not expires_at_str:
        return '', '天', ''
    try:
        et = datetime.strptime(expires_at_str, '%Y-%m-%dT%H:%M')
        remaining_minutes = (et - datetime.now()).total_seconds() / 60
        display = et.strftime('%Y-%m-%d %H:%M')
        if remaining_minutes <= 0:
            return '', '天', display
        if remaining_minutes >= 1440 * 2:
            days = math.ceil(remaining_minutes / 1440)
            return str(days), '天', display
        elif remaining_minutes >= 120:
            hours = math.ceil(remaining_minutes / 60)
            return str(hours), '小时', display
        else:
            mins = max(1, math.ceil(remaining_minutes))
            return str(mins), '分钟', display
    except (ValueError, TypeError):
        return '', '天', ''

def _is_link_public(link, for_share=False):
    """判断链接是否为公开链接（无通行证保护）"""
    link = dict(link)  # 兼容 sqlite3.Row 传入
    if for_share:
        _sp = link.get('share_passcode', '')
        _spe = link.get('share_passcode_empty', 0)
        if _spe:
            return True
        if _sp and _sp.strip():
            return False
        return not (link.get('passcode_plain', '') and link['passcode_plain'].strip())
    else:
        return not (link.get('passcode_plain', '') and link['passcode_plain'].strip())

def _generate_download_token(link_id):
    """为公开链接生成 HMAC-SHA256 下载令牌"""
    expires = int(time.time()) + _DL_TOKEN_TTL
    msg = f"{link_id}|{expires}"
    signature = hmac.new(
        app.secret_key.encode('utf-8'),
        msg.encode('utf-8'),
        hashlib.sha256
    ).hexdigest()
    return signature, expires

def _verify_download_token(link_id, token, expires):
    """验证下载令牌的签名和过期时间"""
    try:
        expires = int(expires)
    except (ValueError, TypeError):
        return False
    if time.time() > expires:
        return False
    expected = hmac.new(
        app.secret_key.encode('utf-8'),
        f"{link_id}|{expires}".encode('utf-8'),
        hashlib.sha256
    ).hexdigest()
    return hmac.compare_digest(expected, token)

def _check_download_rate_limit():
    """IP 级别下载速率限制（对所有链接生效）"""
    client_ip = _get_client_ip()
    if not rate_limit(f'dl_{client_ip}', max_attempts=10, window_seconds=60):
        return False, '下载过于频繁，请稍后再试'
    return True, ''

def _check_preview_rate_limit():
    """IP 级别预览速率限制（放宽阈值，视频 Range 请求较多）"""
    client_ip = _get_client_ip()
    if not rate_limit(f'pv_{client_ip}', max_attempts=100, window_seconds=60):
        return False, '预览请求过于频繁，请稍后再试'
    return True, ''

def _check_public_link_token(link_id, link, for_share=False):
    """
    公开链接的下载令牌检查。仅对无通行证链接生效。
    返回 (通过: bool, 错误信息: str)
    """
    if not _is_link_public(link, for_share):
        return True, ''  # 有通行证保护，跳过令牌检查

    token = request.args.get('token', '')
    expires = request.args.get('expires', '')

    if not token or not expires:
        return False, '缺少访问令牌，请刷新页面后重试'

    if not _verify_download_token(link_id, token, expires):
        return False, '访问令牌无效或已过期，请刷新页面后重试'

    return True, ''

# ============================================================
# 用户认证相关函数
# ============================================================

def get_current_user():
    """获取当前登录用户信息"""
    user_id = session.get('user_id')
    if not user_id:
        return None
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id = ? AND status = 'active'", (user_id,)).fetchone()
    conn.close()
    return dict(user) if user else None

def validate_user_login(username, password):
    """验证用户登录"""
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE username = ? AND status = 'active'", (username,)).fetchone()
    conn.close()
    if user and check_password_hash(user['password_hash'], password):
        return dict(user)
    return None

def validate_invite_code(code):
    """验证邀请码是否有效"""
    conn = get_db()
    invite = None
    try:
        # 检查表是否存在
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='invite_codes'")
        if cursor.fetchone():
            invite = conn.execute(
                "SELECT * FROM invite_codes WHERE code = ? AND expires_at > CURRENT_TIMESTAMP AND used_by IS NULL",
                (code,)
            ).fetchone()
    except Exception as e:
        logger.error(f"验证邀请码失败: {e}")
    conn.close()
    return dict(invite) if invite else None

def mark_invite_code_used(code, user_id):
    """标记邀请码已使用"""
    conn = get_db()
    try:
        conn.execute(
            "UPDATE invite_codes SET used_by = ?, used_at = CURRENT_TIMESTAMP WHERE code = ?",
            (user_id, code)
        )
        conn.commit()
    except Exception as e:
        logger.error(f"标记邀请码失败: {e}")
    conn.close()

def generate_invite_code(created_by, expire_days=7):
    """生成邀请码"""
    code = secrets.token_urlsafe(16)
    expires_at = (datetime.now() + timedelta(days=expire_days)).strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO invite_codes (id, code, expires_at, created_by) VALUES (?, ?, ?, ?)",
            (str(uuid.uuid4()), code, expires_at, created_by)
        )
        conn.commit()
    except Exception as e:
        logger.error(f"生成邀请码失败: {e}")
        code = None
    conn.close()
    return code

def create_user(username, password, email=''):
    """创建新用户"""
    user_id = str(uuid.uuid4())
    password_hash = generate_password_hash(password)
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO users (id, username, password_hash, email, is_admin, status) VALUES (?, ?, ?, ?, 0, 'active')",
            (user_id, username, password_hash, email)
        )
        conn.commit()
        return user_id
    except sqlite3.IntegrityError:
        logger.error(f"创建用户失败（用户名已存在）: {username}")
        return None
    except Exception as e:
        logger.error(f"创建用户失败: {e}")
        return None
    finally:
        conn.close()

def get_user_folder(user_id):
    """获取用户的文件存储目录"""
    user = get_user_by_id(user_id)
    if not user:
        return None
    # 使用用户名作为文件夹名（去除特殊字符）
    safe_username = re.sub(r'[^\w\-]', '_', user['username'])
    return os.path.join(get_upload_base(), safe_username)

def get_link_by_id(link_id):
    """根据ID或slug获取链接信息"""
    conn = get_db()
    try:
        link = conn.execute(
            "SELECT * FROM links WHERE (collect_slug = ? OR share_slug = ? OR id = ?)",
            (link_id, link_id, link_id)
        ).fetchone()
        return dict(link) if link else None
    finally:
        conn.close()

def get_link_folder(link_id, user_id=None):
    """获取链接的文件存储目录（兼容新旧数据）"""
    link = get_link_by_id(link_id)
    
    if link and link.get('user_id'):
        # 有用户归属：UPLOAD_BASE/<username>/<folder_name>/
        user = get_user_by_id(link['user_id'])
        if user:
            safe_username = re.sub(r'[^\w\-]', '_', user['username'])
            folder_name = link.get('folder_name', '') or link_id
            return os.path.join(get_upload_base(), safe_username, folder_name)
    
    # 旧数据：保持原位置
    return os.path.join(get_upload_base(), link_id)

def scan_link_folder(link_id, conn=None):
    """扫描链接文件夹，识别手动放入的文件（高性能版：递归扫描子目录、批量 INSERT）"""
    should_close = False
    if conn is None:
        conn = get_db()
        should_close = True
    
    try:
        # 从已有连接获取 link 信息（避免额外 DB 连接）
        link = conn.execute(
            "SELECT id, user_id, folder_name FROM links WHERE (collect_slug = ? OR share_slug = ? OR id = ?)",
            (link_id, link_id, link_id)
        ).fetchone()
        if not link:
            return []
        
        user_id = link['user_id']
        # 直接构建文件夹路径（内联 get_link_folder 逻辑）
        if user_id:
            user = get_user_by_id(user_id)
            if user:
                safe_username = re.sub(r'[^\w\-]', '_', user['username'])
                folder_name = link['folder_name'] or link_id
                folder_path = os.path.join(get_upload_base(), safe_username, folder_name)
            else:
                folder_path = os.path.join(get_upload_base(), link_id)
        else:
            folder_path = os.path.join(get_upload_base(), link_id)
        
        if not os.path.exists(folder_path):
            return []
        
        # 获取已记录文件（用 stored_path 做去重，支持同名文件在不同子目录）
        records = conn.execute(
            "SELECT stored_path FROM upload_records WHERE link_id = ?",
            (link_id,)
        ).fetchall()
        existing_paths = {r['stored_path'] for r in records}
        
        # 递归扫描并批量 INSERT 新文件
        new_files = []
        batch = []
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        folder_real = os.path.realpath(folder_path)
        
        for root, dirs, files in os.walk(folder_path):
            # 跳过 .chunks 分片临时目录和 _attachment 附件目录
            dirs[:] = [d for d in dirs if d not in ('.chunks', '_attachment')]
            
            # 计算相对于链接根目录的子路径，用于推断 uploader_name
            rel_dir = os.path.relpath(os.path.realpath(root), folder_real)
            # 取第一级子目录名作为 uploader_name
            if rel_dir == '.':
                uploader_name = ''
            else:
                uploader_name = rel_dir.split(os.sep)[0]
            
            for filename in files:
                filepath = os.path.join(root, filename)
                if not os.path.isfile(filepath):
                    continue
                if filepath in existing_paths:
                    continue
                
                file_size = os.path.getsize(filepath)
                batch.append((link_id, user_id or '', filename, filename, filepath,
                             file_size, uploader_name, 'manual', now_str, 'manual'))
                new_files.append(filepath)
                existing_paths.add(filepath)
        
        if batch:
            conn.executemany(
                """INSERT INTO upload_records 
                   (link_id, user_id, original_name, stored_name, stored_path, file_size, 
                    uploader_name, uploader_ip, uploaded_at, source)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                batch
            )
            conn.commit()
            logger.info(f"扫描链接 {link_id} 发现 {len(new_files)} 个新文件（含递归）")
        
        return new_files
        
    except Exception as e:
        logger.error(f"扫描链接文件夹错误: {e}")
        return []
    finally:
        if should_close and conn:
            try:
                conn.close()
            except Exception:
                pass

def get_user_by_id(user_id):
    """根据ID获取用户信息"""
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    conn.close()
    return dict(user) if user else None

def get_user_by_email(email):
    """根据邮箱获取用户信息"""
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE email = ? AND status = 'active'", (email,)).fetchone()
    conn.close()
    return dict(user) if user else None

def sanitize_html(html_text):
    """使用 bleach 清洗 HTML，只允许安全标签和属性
    
    白名单策略：
    - 允许基本格式标签: p, br, strong, b, em, i, u, s, del
    - 允许标题: h1, h2, h3, h4
    - 允许列表: ul, ol, li
    - 允许引用和代码: blockquote, pre, code, sub, sup
    - 允许链接: a (仅 href 属性，强制 nofollow)
    - 允许行内容: span (仅 style 中的颜色相关属性)
    - 禁止: img, video, audio, iframe, script, object, embed 等
    """
    if not html_text or not html_text.strip():
        return ''
    
    allowed_tags = [
        'p', 'br', 'strong', 'b', 'em', 'i', 'u', 's', 'del',
        'h1', 'h2', 'h3', 'h4',
        'ul', 'ol', 'li',
        'blockquote', 'pre', 'code', 'sub', 'sup',
        'a', 'span', 'div', 'hr',
    ]
    
    allowed_attrs = {
        'a': ['href', 'title'],
        'span': ['style'],
    }
    
    allowed_styles = [
        'color', 'background-color', 'background',
        'font-size', 'font-weight', 'font-style',
        'text-decoration',
    ]
    
    # 清洗 HTML（CSSSanitizer 需要 tinycss2，若未安装则降级处理）
    css_sanitizer = None
    try:
        from bleach.css_sanitizer import CSSSanitizer
        css_sanitizer = CSSSanitizer(allowed_css_properties=allowed_styles)
    except ModuleNotFoundError:
        logger.warning("tinycss2 未安装，CSS 样式过滤已降级（建议: pip install tinycss2）")
        # 抑制 bleach 每次处理 style 属性时发出的 NoCssSanitizerWarning
        warnings.filterwarnings('ignore', message='.*css_sanitizer.*', module='bleach')
    
    clean_kwargs = dict(
        tags=allowed_tags,
        attributes=allowed_attrs,
        strip=True,
    )
    if css_sanitizer:
        clean_kwargs['css_sanitizer'] = css_sanitizer
    
    cleaned = bleach.clean(html_text, **clean_kwargs)
    
    # 对所有链接强制添加 rel="nofollow noopener noreferrer" 和 target="_blank"
    # bleach 的 linkify 可以做到这点
    cleaned = bleach.linkify(
        cleaned,
        callbacks=[
            lambda attrs, new: dict(attrs, **{
                'rel': 'nofollow noopener noreferrer',
                'target': '_blank',
            }) if 'href' in attrs else attrs
        ]
    )
    
    return cleaned


def get_smtp_config():
    """获取 SMTP 配置字典"""
    try:
        port = int(get_setting('smtp_port', '587'))
    except (ValueError, TypeError):
        port = 587
    return {
        'host': get_setting('smtp_host', ''),
        'port': port,
        'username': get_setting('smtp_username', ''),
        'password': get_setting('smtp_password', ''),
        'use_tls': get_setting('smtp_use_tls', '1') == '1',
        'from_email': get_setting('smtp_from_email', ''),
        'from_name': get_setting('smtp_from_name', '文件收集器'),
    }

def send_email(to_email, subject, body_html):
    """发送邮件，返回 (success, message)"""
    # 延迟导入重型邮件模块（启动时不加载）
    import smtplib as _smtplib
    import email.utils as _email_utils
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    config = get_smtp_config()
    if not config['host'] or not config['from_email']:
        return False, 'SMTP 未配置'
    server = None
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = _email_utils.formataddr((config['from_name'], config['from_email']))
        msg['To'] = to_email
        msg['Message-ID'] = _email_utils.make_msgid(domain=config['from_email'].split('@')[-1] if '@' in config['from_email'] else 'localhost')
        msg['Date'] = _email_utils.formatdate(localtime=True)
        msg.attach(MIMEText(body_html, 'html', 'utf-8'))

        if config['use_tls']:
            # STARTTLS: 先建立明文连接，再升级到 TLS
            server = _smtplib.SMTP(config['host'], config['port'], timeout=30)
            server.ehlo()
            server.starttls()
            server.ehlo()
        else:
            # SSL: 直接建立加密连接
            server = _smtplib.SMTP_SSL(config['host'], config['port'], timeout=30)
            server.ehlo()

        if config['username'] and config['password']:
            server.login(config['username'], config['password'])
        server.sendmail(config['from_email'], [to_email], msg.as_string())
        logger.info(f"邮件发送成功: {to_email} - {subject}")
        return True, '发送成功'
    except _smtplib.SMTPAuthenticationError:
        return False, 'SMTP 认证失败，请检查用户名和密码'
    except _smtplib.SMTPException as e:
        return False, f'SMTP 错误: {str(e)[:100]}'
    except Exception as e:
        return False, f'发送失败: {str(e)[:100]}'
    finally:
        if server:
            try:
                server.quit()
            except Exception:
                pass

def cleanup_orphan_records(conn=None):
    """清理孤儿记录（批量版：内存比对 + 批量 DELETE，避免逐行磁盘 I/O）"""
    should_close = False
    if conn is None:
        conn = get_db()
        should_close = True
    try:
        records = conn.execute(
            "SELECT id, stored_path, stored_name, link_id FROM upload_records"
        ).fetchall()
        upload_base = get_upload_base()
        real_base = os.path.realpath(upload_base)
        
        # 构建：(link_id, stored_name) -> record 的映射
        # 同时预检查路径是否存在
        orphan_ids = []
        for r in records:
            real_path = os.path.realpath(r['stored_path'])
            if (real_path.startswith(real_base + os.sep) or real_path == real_base):
                if not os.path.exists(real_path) or not os.path.isfile(real_path):
                    orphan_ids.append(r['id'])
        
        if orphan_ids:
            # 批量删除（SQLite 限制一次最多 999 个参数）
            chunk_size = 500
            for i in range(0, len(orphan_ids), chunk_size):
                chunk = orphan_ids[i:i + chunk_size]
                placeholders = ','.join(['?' for _ in chunk])
                conn.execute(f"DELETE FROM upload_records WHERE id IN ({placeholders})", chunk)
            conn.commit()
            logger.info(f"清理孤儿记录: {len(orphan_ids)} 条")
        return len(orphan_ids)
    except Exception as e:
        logger.error(f"清理孤儿记录失败: {e}")
        return 0
    finally:
        if should_close and conn:
            try:
                conn.close()
            except Exception:
                pass

# 孤儿记录清理缓存：避免每次请求都扫文件系统（NAS/HDD 下尤其耗时）
_orphan_cleanup_cache = {}  # {link_id: timestamp}

def cleanup_orphan_records_for_link(conn, link_id):
    """清理指定链接的孤儿记录（批量版），60秒内不重复扫描"""
    now = time.time()
    last = _orphan_cleanup_cache.get(link_id, 0)
    if now - last < 60:
        return 0  # 60 秒内已清理过，跳过文件系统扫描
    
    try:
        records = conn.execute(
            "SELECT id, stored_path FROM upload_records WHERE link_id = ?",
            (link_id,)
        ).fetchall()
        if not records:
            _orphan_cleanup_cache[link_id] = now
            return 0
        
        upload_base = get_upload_base()
        real_base = os.path.realpath(upload_base)
        
        orphan_ids = []
        for r in records:
            real_path = os.path.realpath(r['stored_path'])
            if (real_path.startswith(real_base + os.sep) or real_path == real_base):
                if not os.path.exists(real_path) or not os.path.isfile(real_path):
                    orphan_ids.append(r['id'])
        
        if orphan_ids:
            placeholders = ','.join(['?' for _ in orphan_ids])
            conn.execute(f"DELETE FROM upload_records WHERE id IN ({placeholders})", orphan_ids)
            conn.commit()
        
        _orphan_cleanup_cache[link_id] = now
        # 定期清理过期缓存，防止内存泄漏
        cutoff = now - 3600
        stale = [k for k, v in _orphan_cleanup_cache.items() if v < cutoff]
        for k in stale:
            _orphan_cleanup_cache.pop(k, None)
        
        return len(orphan_ids)
    except Exception as e:
        _orphan_cleanup_cache[link_id] = now
        logger.error(f"清理链接 {link_id} 孤儿记录失败: {e}")
        return 0

# .part 文件清理缓存（60秒内不重复扫描）
_part_cleanup_cache = {}

def cleanup_stale_part_files(conn, link_id):
    """静默清理超过24小时的 .part 断点文件"""
    now = time.time()
    last = _part_cleanup_cache.get(link_id, 0)
    if now - last < 60:
        return 0  # 60 秒内已清理过，跳过
    
    try:
        # 查询超过24小时的 .part 文件记录
        stale_records = conn.execute(
            "SELECT id, stored_path, original_name FROM upload_records WHERE link_id = ? AND original_name LIKE '%.part'",
            (link_id,)
        ).fetchall()
        
        if not stale_records:
            _part_cleanup_cache[link_id] = now
            return 0
        
        cutoff_time = datetime.fromtimestamp(now - 86400)  # 24小时前
        stale_ids = []
        
        for r in stale_records:
            uploaded_at = datetime.fromisoformat(r['uploaded_at']) if r['uploaded_at'] else None
            if uploaded_at and uploaded_at < cutoff_time:
                stale_ids.append(r['id'])
                # 静默删除 .part 文件
                try:
                    if r['stored_path'] and os.path.exists(r['stored_path']):
                        os.remove(r['stored_path'])
                    part_file = r['stored_path'] + '.part'
                    if os.path.exists(part_file):
                        os.remove(part_file)
                except Exception:
                    pass  # 静默处理，不影响流程
        
        if stale_ids:
            placeholders = ','.join(['?' for _ in stale_ids])
            conn.execute(f"DELETE FROM upload_records WHERE id IN ({placeholders})", stale_ids)
            conn.commit()
        
        _part_cleanup_cache[link_id] = now
        
        # 定期清理过期缓存
        cutoff = now - 3600
        stale = [k for k, v in _part_cleanup_cache.items() if v < cutoff]
        for k in stale:
            _part_cleanup_cache.pop(k, None)
        
        return len(stale_ids)
    except Exception as e:
        _part_cleanup_cache[link_id] = now
        logger.error(f"清理链接 {link_id} 断点文件失败: {e}")
        return 0

def get_user_files(user_id):
    """获取用户的所有文件记录（包括直接放入文件夹的文件）"""
    conn = get_db()
    
    # 清理孤儿记录
    cleanup_orphan_records(conn)
    
    # 从数据库获取记录
    records = conn.execute(
        "SELECT * FROM upload_records WHERE user_id = ? ORDER BY uploaded_at DESC",
        (user_id,)
    ).fetchall()
    
    records_list = [dict(r) for r in records]
    
    # 检查用户文件夹中是否有未记录的文件
    try:
        user_folder = get_user_folder(user_id)
        if user_folder and os.path.isdir(user_folder):
            for item in os.listdir(user_folder):
                item_path = os.path.join(user_folder, item)
                if os.path.isfile(item_path):
                    # 检查是否已在数据库中
                    exists = any(r['stored_path'] == item_path for r in records_list)
                    if not exists:
                        # 添加未记录的文件
                        file_size = os.path.getsize(item_path)
                        records_list.append({
                            'id': None,
                            'user_id': user_id,
                            'link_id': None,
                            'original_name': item,
                            'stored_name': item,
                            'stored_path': item_path,
                            'file_size': file_size,
                            'file_size_display': format_file_size(file_size),
                            'uploader_ip': '直接上传',
                            'uploaded_at': datetime.fromtimestamp(os.path.getmtime(item_path)).strftime('%Y-%m-%d %H:%M:%S'),
                            'download_count': 0
                        })
    except Exception as e:
        logger.error(f"扫描用户文件夹失败: {e}")
    
    # 按上传时间排序
    records_list.sort(key=lambda x: x['uploaded_at'], reverse=True)
    conn.close()
    return records_list

def admin_required(f):
    """管理员认证装饰器"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('is_admin'):
            if request.path.startswith('/api/'):
                return jsonify({'error': '未登录'}), 401
            return redirect(url_for('admin_login'))
        if request.method == 'POST':
            if not validate_csrf():
                return redirect(request.referrer or url_for('admin_dashboard'))
        admin_hash = get_setting('admin_password_hash', '')
        # 默认密码强制改密（除设置页和登出外，强制跳转到设置页）
        if check_password_hash(admin_hash, DEFAULT_ADMIN_PASS):
            if request.endpoint not in ('admin_settings', 'admin_logout'):
                if request.path.startswith('/api/'):
                    return jsonify({'error': '请先修改默认密码'}), 403
                flash('安全警告：您仍在使用默认密码，请立即修改！', 'error')
                return redirect(url_for('admin_settings'))
        # 昵称未设置提醒
        if session.get('is_admin'):
            conn = get_db()
            user = conn.execute("SELECT nickname FROM users WHERE id = ?", (session['user_id'],)).fetchone()
            conn.close()
            if user and not user['nickname'] and request.endpoint not in ('admin_settings', 'admin_logout'):
                flash('请设置您的昵称，将显示在收集页和分享页中')
        return f(*args, **kwargs)
    return decorated

def login_required(f):
    """登录认证装饰器（管理员和普通用户均可访问）"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user_id'):
            if request.path.startswith('/api/'):
                return jsonify({'error': '未登录'}), 401
            return redirect(url_for('admin_login'))
        if request.method == 'POST':
            if not validate_csrf():
                return redirect(request.referrer or url_for('admin_dashboard'))
        # 管理员默认密码强制改密（除设置页和登出外，强制跳转到设置页）
        if session.get('is_admin'):
            admin_hash = get_setting('admin_password_hash', '')
            if check_password_hash(admin_hash, DEFAULT_ADMIN_PASS):
                if request.endpoint not in ('admin_settings', 'admin_logout', 'user_settings'):
                    if request.path.startswith('/api/'):
                        return jsonify({'error': '请先修改默认密码'}), 403
                    flash('安全警告：您仍在使用默认密码，请立即修改！', 'error')
                    return redirect(url_for('admin_settings'))
            # 昵称未设置提醒
            conn = get_db()
            user = conn.execute("SELECT nickname FROM users WHERE id = ?", (session['user_id'],)).fetchone()
            conn.close()
            if user and not user['nickname'] and request.endpoint not in ('admin_settings', 'admin_logout', 'user_settings'):
                flash('请设置您的昵称，将显示在收集页和分享页中')
        return f(*args, **kwargs)
    return decorated

def _check_record_ownership(record_id):
    """校验当前用户是否有权访问指定上传记录（非管理员只能访问自己链接下的记录）"""
    if session.get('is_admin'):
        return True
    user_id = session.get('user_id')
    conn = get_db()
    row = conn.execute(
        """SELECT r.id FROM upload_records r
           INNER JOIN links l ON r.link_id = l.id
           WHERE r.id = ? AND l.user_id = ?""",
        (record_id, user_id)
    ).fetchone()
    conn.close()
    return row is not None

def _can_preview_attachment_ext(filename):
    """判断附件是否支持预览"""
    if not filename:
        return False
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    jit_exts = {'doc','docx','xls','xlsx','ppt','pptx','pdf','ofd','txt','md','markdown','csv','dxf','dwg'}
    img_exts = {'jpg','jpeg','png','gif','webp','bmp','svg','heic','heif'}
    vid_exts = {'mp4','webm','ogg','mov','avi','mkv'}
    aud_exts = {'mp3','wav','flac','aac','m4a'}
    return ext in jit_exts or ext in img_exts or ext in vid_exts or ext in aud_exts


def _attachment_preview_type(filename):
    """返回附件预览类型: image / video / audio / office"""
    if not filename:
        return 'office'
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    img_exts = {'jpg','jpeg','png','gif','webp','bmp','svg','heic','heif'}
    vid_exts = {'mp4','webm','ogg','mov','avi','mkv'}
    aud_exts = {'mp3','wav','flac','aac','m4a'}
    if ext in img_exts:
        return 'image'
    if ext in vid_exts:
        return 'video'
    if ext in aud_exts:
        return 'audio'
    return 'office'


def _check_link_ownership(link_id):
    """校验当前用户是否有权访问指定链接（非管理员只能访问自己的链接）"""
    if session.get('is_admin'):
        return True
    user_id = session.get('user_id')
    conn = get_db()
    row = conn.execute(
        "SELECT id FROM links WHERE id = ? AND user_id = ?",
        (link_id, user_id)
    ).fetchone()
    conn.close()
    return row is not None


def _resolve_public_url(user_id=None):
    """
    解析有效的公网地址：
    1. 反代运行中 → 使用反代域名
    2. 手动设置的 public_url
    3. 空（前端回退到 window.location.origin）
    """
    rp_public = RPROXY_PM.get_public_url()
    if rp_public:
        return rp_public
    if user_id:
        return get_user_setting(user_id, 'public_url', get_setting('public_url', ''))
    return get_setting('public_url', '')


@app.after_request
def add_security_headers(response):
    """添加安全响应头 + 缓存策略"""
    # HTML 页面不缓存（含动态内容：CSRF token、用户信息等）
    ct = (response.content_type or '')
    if 'text/html' in ct:
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    # 安全头
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'camera=(), microphone=(), geolocation=()'
    csp = "default-src 'self' blob:; script-src 'self' 'unsafe-inline' blob:; style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; font-src 'self' https://fonts.gstatic.com; img-src 'self' data: blob:; connect-src 'self' https://api.github.com https://github.com blob:; worker-src 'self' blob:"
    response.headers['Content-Security-Policy'] = csp
    return response


@app.after_request
def minify_html_response(response):
    """压缩 HTML 响应：去除注释和多余空白，减小传输体积"""
    ct = (response.content_type or '')
    if 'text/html' not in ct:
        return response
    # 仅压缩成功响应（200/304），跳过错误页
    if response.status_code not in (200, 304):
        return response
    try:
        data = response.get_data(as_text=True)
        if not data:
            return response
        minified = _minify_html(data)
        response.set_data(minified)
    except Exception:
        pass  # 压缩失败不影响正常响应
    return response


# 静态文件 MIME 类型修复 ——
# X-Content-Type-Options: nosniff 要求 Content-Type 必须精确，
# 而某些部署环境（fnOS 内嵌代理等）可能丢失或错误设置 Content-Type。
@app.after_request
def fix_static_content_type(response):
    """根据文件后缀补全/修正静态资源的 Content-Type"""
    path = request.path or ''
    # 仅处理 /static/ 路径下的请求
    if '/static/' not in path:
        return response
    ext = os.path.splitext(path)[1].lower()
    known = _ESSENTIAL_MIMETYPES.get(ext)
    if not known:
        return response
    current = response.content_type or ''
    # 如果当前类型为空、未设置，或变成了 text/plain/octet-stream，则修正
    if not current or 'text/plain' in current or 'application/octet-stream' in current:
        response.headers['Content-Type'] = known
        return response
    # 为 CSS/JS 补充 charset（某些环境可能缺失）
    if ext in ('.css',) and 'charset' not in current:
        response.headers['Content-Type'] = 'text/css; charset=utf-8'
    elif ext in ('.js', '.mjs') and 'charset' not in current:
        response.headers['Content-Type'] = 'application/javascript; charset=utf-8'
    # 静态资源强缓存：CSS/JS/字体/图片等长期缓存，HTML 不缓存
    if ext in _ESSENTIAL_MIMETYPES and ext not in ('.html', '.htm',):
        response.cache_control.public = True
        response.cache_control.max_age = 3600 * 24
        response.headers['Vary'] = 'Accept-Encoding'
    return response

@app.context_processor
def inject_globals():
    """注入全局模板变量"""
    title = get_setting('site_title', '文件收集器')
    return {
        'csrf_token': generate_csrf_token(),
        'site_title': title,
        'config_title': title,
        'collect_footer_text': get_setting('collect_footer_text', ''),
    }

@app.template_filter('parse_json')
def parse_json_filter(s):
    """Jinja2 过滤器：安全解析 JSON 字符串为 dict，失败返回 None"""
    if not s:
        return None
    try:
        return json.loads(s)
    except (json.JSONDecodeError, TypeError):
        return None

@app.template_filter('file_size')
def file_size_filter(size_bytes):
    """Jinja2 过滤器：格式化文件大小，支持 int/float/str"""
    try:
        size_bytes = int(size_bytes)
    except (ValueError, TypeError):
        return str(size_bytes)
    return format_file_size(size_bytes)

def _safe_delete(stored_path):
    """安全删除文件：校验 stored_path 在 UPLOAD_BASE 范围内，防止路径遍历攻击"""
    upload_base = get_upload_base()
    real_path = os.path.realpath(stored_path)
    real_base = os.path.realpath(upload_base)
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        logger.warning(f"路径遍历删除拦截: stored_path={stored_path}, upload_base={real_base}")
        return False
    if os.path.exists(real_path) and os.path.isfile(real_path):
        os.remove(real_path)
        return True
    return False


def _safe_delete_dir(dir_path):
    """安全删除目录：校验 dir_path 在 UPLOAD_BASE 范围内，防止路径遍历攻击"""
    upload_base = get_upload_base()
    real_path = os.path.realpath(dir_path)
    real_base = os.path.realpath(upload_base)
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        logger.warning(f"路径遍历删除拦截(目录): dir_path={dir_path}, upload_base={real_base}")
        return False
    if os.path.exists(real_path) and os.path.isdir(real_path):
        shutil.rmtree(real_path)
        return True
    return False

def safe_filename_unicode(filename):
    """安全文件名处理，保留Unicode字符（中文等），仅移除危险字符"""
    if not filename:
        return 'unnamed_file'
    # Unicode 规范化（防止 homoglyph 攻击）
    filename = unicodedata.normalize('NFC', filename)
    # 移除路径分隔符和空字符等危险字符
    filename = filename.replace('/', '_').replace('\\', '_').replace('\x00', '')
    # 移除首尾空白和点（防止隐藏文件）
    filename = filename.strip().strip('.')
    if not filename:
        return 'unnamed_file'
    # 限制长度（保留扩展名）
    if len(filename) > 255:
        name, ext = os.path.splitext(filename)
        filename = name[:251 - len(ext)] + ext
    return filename

def allowed_file(filename):
    """允许所有文件类型上传，但禁止用户配置的危险扩展名"""
    blocked = get_blocked_extensions()
    if not blocked:
        return True
    _, ext = os.path.splitext(filename.lower())
    if ext in blocked:
        return False
    return True

def get_blocked_extensions():
    """从设置中获取禁止上传的扩展名列表"""
    raw = get_setting('blocked_extensions', '').strip()
    if not raw:
        # 默认：禁止危险脚本和可执行文件
        default_blocked = {'.exe', '.php', '.jsp', '.asp', '.aspx', '.sh', '.bash', '.bat',
                           '.cmd', '.ps1', '.py', '.rb', '.pl', '.cgi', '.so', '.dll',
                           '.jspx', '.php3', '.php4', '.php5', '.phtml', '.shtml'}
        return default_blocked
    # 用户自定义：逗号或空格分隔
    exts = set()
    for part in raw.replace(',', ' ').split():
        part = part.strip().lower()
        if part and not part.startswith('.'):
            part = '.' + part
        if part:
            exts.add(part)
    return exts

def _safe_download(stored_path, original_name):
    """安全下载：校验 stored_path 在 UPLOAD_BASE 范围内，防止路径遍历攻击"""
    upload_base = get_upload_base()
    real_path = os.path.realpath(stored_path)
    real_base = os.path.realpath(upload_base)
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        logger.warning(f"路径遍历攻击拦截: stored_path={stored_path}, upload_base={real_base}")
        abort(403)
    directory = os.path.dirname(real_path)
    filename = os.path.basename(real_path)
    if not os.path.isfile(real_path):
        abort(404)
    return send_from_directory(
        directory, filename,
        download_name=original_name,
        as_attachment=True
    )

def _serve_heic_as_jpeg(stored_path):
    """将 HEIC 文件转换为 JPEG 并返回 bytesio"""
    if not HEIC_SUPPORT:
        return None
    
    # 路径遍历防护
    upload_base = get_upload_base()
    real_path = os.path.realpath(stored_path)
    real_base = os.path.realpath(upload_base)
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        logger.warning(f"HEIC 路径遍历攻击拦截: stored_path={stored_path}")
        return None
    
    try:
        from PIL import Image
        import io
        img = Image.open(stored_path)
        buf = io.BytesIO()
        # 转换 RGB（去掉 alpha 通道，JPEG 不支持）
        if img.mode in ('RGBA', 'LA', 'P'):
            img = img.convert('RGB')
        img.save(buf, format='JPEG', quality=92)
        buf.seek(0)
        return buf
    except Exception:
        return None


# ========== TXT 阅读器工具函数 ==========

_TXT_CHAPTER_CACHE = {}  # {filepath: (mtime, chapters, encoding)}

def _detect_txt_encoding(filepath, sample_size=4096):
    """检测文本文件编码，返回编码名称"""
    with open(filepath, 'rb') as f:
        raw = f.read(sample_size)
    if not raw:
        return 'utf-8'
    # BOM 检测
    if raw.startswith(b'\xef\xbb\xbf'):
        return 'utf-8-sig'
    if raw.startswith(b'\xff\xfe'):
        return 'utf-16-le'
    if raw.startswith(b'\xfe\xff'):
        return 'utf-16-be'
    # 尝试 UTF-8
    try:
        raw.decode('utf-8')
        return 'utf-8'
    except UnicodeDecodeError:
        pass
    # 尝试 GBK（中文 Windows 常用）
    try:
        raw.decode('gbk')
        return 'gbk'
    except UnicodeDecodeError:
        pass
    # 尝试 GB2312
    try:
        raw.decode('gb2312')
        return 'gb2312'
    except UnicodeDecodeError:
        pass
    # 尝试 UTF-16
    try:
        raw.decode('utf-16')
        return 'utf-16'
    except UnicodeDecodeError:
        pass
    # 最后尝试 latin-1（兜底，不会失败）
    return 'latin-1'


_CHAPTER_PATTERNS = [
    re.compile(r'^[第序][\s]*[0-9零一二三四五六七八九十百千万]+[\s]*[章节卷部篇集]'),
    re.compile(r'^[卷][\s]*[0-9零一二三四五六七八九十百千万]+'),
    re.compile(r'^Chapter\s+\d+', re.IGNORECASE),
    re.compile(r'^(序言|前言|楔子|后记|尾声|番外|终章|大结局)'),
    re.compile(r'^[第][\s]*[0-9零一二三四五六七八九十百千万]+[\s]*(节|回|幕)'),
]


def _scan_txt_chapters(filepath):
    """扫描 TXT 文件章节，返回 (chapters, encoding)
    
    使用缓存：如文件未修改则复用上次结果。
    chapters 格式: [{'title': '第一章 xxx', 'offset': 1234}, ...]
    """
    mtime = os.path.getmtime(filepath)
    if filepath in _TXT_CHAPTER_CACHE:
        cached_mtime, cached_chapters, cached_encoding = _TXT_CHAPTER_CACHE[filepath]
        if cached_mtime == mtime:
            return cached_chapters, cached_encoding

    encoding = _detect_txt_encoding(filepath)
    chapters = []
    offset = 0
    max_scan = 50 * 1024 * 1024  # 最多扫描 50MB

    try:
        with open(filepath, 'rb') as f:
            # 跳过 BOM
            first_bytes = f.read(4)
            if first_bytes.startswith(b'\xef\xbb\xbf'):
                offset = 3
            elif first_bytes.startswith(b'\xff\xfe') or first_bytes.startswith(b'\xfe\xff'):
                offset = 2
            f.seek(0)

            leftover = b''
            line_count = 0
            while offset < max_scan:
                chunk = f.read(65536)
                if not chunk:
                    break
                data = leftover + chunk
                if b'\n' not in data:
                    leftover = data
                    continue
                # 按行分割，最后一段不完整的留到下次
                *lines, leftover = data.split(b'\n')
                for line in lines:
                    line_count += 1
                    try:
                        text = line.decode(encoding, errors='replace').strip()
                    except Exception:
                        text = line.decode('latin-1', errors='replace').strip()
                    if not text:
                        continue
                    for pat in _CHAPTER_PATTERNS:
                        if pat.match(text):
                            chapters.append({
                                'title': text[:80],
                                'offset': offset
                            })
                            break
                    # 只扫描前 5 万行找章节（小说通常前几万行就有所有章节标题）
                    if line_count > 50000 and len(chapters) > 0:
                        break
                offset += len(data) - len(leftover)
                if line_count > 50000 and len(chapters) > 0:
                    break
    except Exception:
        pass

    # 去重：相邻相同标题合并
    seen = set()
    unique_chapters = []
    for ch in chapters:
        key = ch['title'].lower()
        if key not in seen:
            seen.add(key)
            unique_chapters.append(ch)

    _TXT_CHAPTER_CACHE[filepath] = (mtime, unique_chapters, encoding)
    # 限制缓存大小
    if len(_TXT_CHAPTER_CACHE) > 200:
        oldest = next(iter(_TXT_CHAPTER_CACHE))
        del _TXT_CHAPTER_CACHE[oldest]

    return unique_chapters, encoding


def _read_txt_chunk(filepath, offset, size, encoding=None):
    """读取 TXT 文件指定偏移的文本块，返回解码后的文本"""
    if encoding is None:
        encoding = _detect_txt_encoding(filepath)
    file_size = os.path.getsize(filepath)
    with open(filepath, 'rb') as f:
        # 如果不是从 0 开始，向前多读一点找到最近的换行符，保证从完整行开始
        if offset > 0:
            f.seek(max(0, offset - 256))
            prefix = f.read(min(256, offset))
            # 找最后一个换行符
            last_nl = prefix.rfind(b'\n')
            if last_nl >= 0:
                offset = max(0, offset - 256) + last_nl + 1
            else:
                offset = max(0, offset - 256)
        f.seek(offset)
        raw = f.read(size)
    text = raw.decode(encoding, errors='replace')
    # 如果带 BOM 且从开头读，去掉 BOM 字符
    if offset == 0 and encoding == 'utf-8-sig' and text.startswith('\ufeff'):
        text = text[1:]
        encoding = 'utf-8'
    return text, file_size


# ============================================================

def create_upload_dir(link_id, uploader_name=''):
    """为链接创建上传目录（用户隔离：UPLOAD_BASE/<username>/<folder_name>/）
    如果提供了 uploader_name，则在其下创建子文件夹：<...>/<folder_name>/<uploader_name>/"""
    conn = get_db()
    link = conn.execute(
        "SELECT id, title, user_id, folder_name FROM links WHERE (collect_slug = ? OR share_slug = ? OR id = ?)",
        (link_id, link_id, link_id)
    ).fetchone()
    conn.close()

    canonical_id = link['id'] if link else link_id

    if not link:
        upload_dir = os.path.join(UPLOAD_BASE, 'unnamed', canonical_id)
    elif link['user_id']:
        # 有用户归属：UPLOAD_BASE/<username>/<folder_name>/
        user = get_user_by_id(link['user_id'])
        if user:
            safe_username = re.sub(r'[^\w\-]', '_', user['username'])
            folder_name = link['folder_name'] or canonical_id
            upload_dir = os.path.join(UPLOAD_BASE, safe_username, folder_name)
        else:
            upload_dir = os.path.join(UPLOAD_BASE, 'unnamed', canonical_id)
    else:
        # 无用户归属（旧数据）：使用标题命名
        folder_name = re.sub(r'[<>:"/\\|?*]', '_', link['title'].strip()) if link['title'] else 'unnamed'
        if not folder_name:
            folder_name = 'unnamed'
        folder_name = os.path.normpath(folder_name).lstrip(os.sep).lstrip('.')
        if not folder_name:
            folder_name = 'unnamed'
        upload_dir = os.path.join(UPLOAD_BASE, folder_name)

    # 双重确保：realpath 必须在 UPLOAD_BASE 范围内
    real_dir = os.path.realpath(upload_dir)
    real_base = os.path.realpath(UPLOAD_BASE)
    if not real_dir.startswith(real_base + os.sep) and real_dir != real_base:
        upload_dir = os.path.join(UPLOAD_BASE, 'unnamed', canonical_id)
        real_dir = os.path.realpath(upload_dir)

    # 上传者子文件夹
    if uploader_name:
        safe_uploader = _sanitize_uploader_name(uploader_name)
        real_dir = os.path.join(real_dir, safe_uploader)
        # 再次安全检查
        if not os.path.realpath(real_dir).startswith(real_base + os.sep) and os.path.realpath(real_dir) != real_base:
            real_dir = os.path.realpath(os.path.join(upload_dir, 'unknown'))

    os.makedirs(real_dir, mode=0o755, exist_ok=True)

    # 确保目录可写（修复 NAS 上权限不一致的问题）
    if not os.access(real_dir, os.W_OK):
        try:
            os.chmod(real_dir, 0o755)
        except Exception:
            pass
        if not os.access(real_dir, os.W_OK):
            raise PermissionError(f'上传目录无写入权限: {real_dir}')

    return real_dir

DEFAULT_ATTACHMENT_MAX_MB = 1000  # 默认附件上限 1000 MB (≈1GB)

def get_attachment_max_size(user_id=None):
    """获取附件大小上限（字节），优先用户设置，回退默认值"""
    mb = DEFAULT_ATTACHMENT_MAX_MB
    if user_id:
        val = get_user_setting(user_id, 'attachment_max_mb', str(DEFAULT_ATTACHMENT_MAX_MB))
        try:
            mb = float(val)
            if mb <= 0:
                mb = DEFAULT_ATTACHMENT_MAX_MB
        except (ValueError, TypeError):
            mb = DEFAULT_ATTACHMENT_MAX_MB
    return int(mb * 1024 * 1024)

def _save_link_attachment(link_id, file_obj, user_id=None):
    """保存链接附件（老师上传的作业等），安全存储在链接目录的 _attachment 子目录下"""
    # 验证大小
    file_obj.seek(0, 2)
    file_size = file_obj.tell()
    file_obj.seek(0)
    max_size = get_attachment_max_size(user_id)
    if file_size > max_size:
        raise ValueError(f'附件大小不能超过 {format_file_size(max_size)}')
    if file_size == 0:
        return

    # 安全文件名
    ext = os.path.splitext(file_obj.filename)[1]
    safe_name = re.sub(r'[^\w.\-]', '_', file_obj.filename)
    if not safe_name:
        safe_name = 'attachment' + ext

    # 存储到链接目录的 _attachment 子目录
    upload_dir = create_upload_dir(link_id, '')
    attach_dir = os.path.join(upload_dir, '_attachment')
    os.makedirs(attach_dir, mode=0o755, exist_ok=True)

    # 如果有旧附件，先删除
    conn = get_db()
    old = conn.execute("SELECT attachment_path FROM links WHERE id=?", (link_id,)).fetchone()
    if old and old['attachment_path']:
        _safe_delete(old['attachment_path'])

    # 保存新附件
    stored_path = os.path.join(attach_dir, safe_name)
    file_obj.save(stored_path)

    conn.execute(
        "UPDATE links SET attachment_name=?, attachment_path=?, attachment_size=? WHERE id=?",
        (file_obj.filename, stored_path, file_size, link_id)
    )
    conn.commit()
    conn.close()
    logger.info(f"链接 {link_id} 附件已保存: {file_obj.filename} ({format_file_size(file_size)})")


# ============================================================
# 路由 - 文件收集页
# ============================================================
@app.route('/')
def index():
    """根路径 - 根据设置显示介绍页或 404"""
    if get_setting('landing_page_enabled', '1') == '1':
        return render_template('landing.html',
            site_title=get_setting('site_title', '文件收集器'))
    return render_template('error.html',
        error_code=404,
        error_title='页面未找到',
        error_message='您访问的页面不存在或已被移除。'), 404

@app.route('/collect/<link_id>')
def collect_page(link_id):
    """文件收集页面"""
    if not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$', link_id):
        return render_template('error.html',
            error_code=404,
            error_title='链接无效',
            error_message='链接格式不正确'), 404
    conn = get_db()
    row = conn.execute(
        "SELECT * FROM links WHERE (collect_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()
    conn.close()
    link = dict(row) if row else None

    if not link:
        return render_template('error.html',
            error_code=404,
            error_title='链接失效',
            error_message='链接不存在或已被停用'), 404

    # 统一使用数据库规范 ID（支持自定义 slug 访问）
    link_id = link['id']

    # 检查收集页是否启用
    if not link.get('collect_enabled', 1):
        share_hint = ''
        if link.get('share_enabled', 0):
            share_hint = '收集已关闭，但您仍可通过分享链接查看文件。'
        return render_template('error.html',
            error_code=410,
            error_title='收集已关闭',
            error_message=share_hint or '该收集链接已被创建者关闭。'), 410

    if link['expires_at']:
        try:
            expire_time = datetime.strptime(link['expires_at'], '%Y-%m-%dT%H:%M')
            if datetime.now() > expire_time:
                return render_template('error.html',
                    error_code=410,
                    error_title='链接已过期',
                    error_message='该收集链接已超过有效期，不再接受文件上传。'), 410
        except (ValueError, TypeError):
            pass

    has_passcode = bool(link['passcode_plain'] and link['passcode_plain'].strip())
    require_uploader = bool(link.get('require_uploader', 0))
    passcode_empty = bool(link.get('passcode_empty', 0))
    
    # 上传者身份：从 session 获取（不依赖 passcode_empty）
    uploader_name = ''
    if require_uploader:
        uploader_name = (session.get(f'uploader_{link_id}') or '').strip()
    
    # verified 仅取决于通行证验证，上传者不再作为门禁
    verified = is_verified(link_id) if has_passcode else True
    
    # 自动扫描文件夹中的新文件
    scan_link_folder(link_id)
    
    link_owner_id = link.get('user_id', '')
    ttl_minutes = int(get_user_setting(link_owner_id, 'passcode_ttl_minutes', '120'))
    if ttl_minutes >= 60 and ttl_minutes % 60 == 0:
        ttl_display = f'{ttl_minutes // 60} 小时'
    elif ttl_minutes >= 60:
        ttl_display = f'{ttl_minutes // 60} 小时 {ttl_minutes % 60} 分钟'
    else:
        ttl_display = f'{ttl_minutes} 分钟'

    expire_display = '永不过期'
    expire_days_left = -1  # -1 表示永不过期
    expire_text = '有效期不限'
    expire_level = 'forever'
    if link['expires_at']:
        try:
            expire_time = datetime.strptime(link['expires_at'], '%Y-%m-%dT%H:%M')
            expire_display = expire_time.strftime('%Y-%m-%d %H:%M')
            remain = expire_time - datetime.now()
            expire_days_left = max(0, remain.days)
            total_seconds = remain.total_seconds()
            if total_seconds <= 0:
                expire_text = '已过期'
                expire_level = 'danger'
            elif total_seconds < 60:
                expire_text = '即将过期'
                expire_level = 'danger'
            elif total_seconds < 3600:
                expire_text = f'还剩 {int(total_seconds // 60)} 分钟'
                expire_level = 'danger'
            elif total_seconds < 86400:
                hour_s = int(total_seconds // 3600)
                min_s = int((total_seconds % 3600) // 60)
                expire_text = f'还剩 {hour_s} 小时' if min_s == 0 else f'还剩 {hour_s} 小时 {min_s} 分钟'
                expire_level = 'warn'
            else:
                days = remain.days
                expire_text = f'还剩 {days} 天'
                expire_level = 'normal' if days > 1 else 'warn'
        except (ValueError, TypeError):
            pass

    # 查询创建者名称（优先昵称）
    creator_name = ''
    try:
        conn2 = get_db()
        user_row = conn2.execute("SELECT username, nickname FROM users WHERE id = ?", (link_owner_id,)).fetchone()
        conn2.close()
        if user_row:
            creator_name = user_row['nickname'] or user_row['username']
    except Exception:
        pass

    csrf_token = generate_csrf_token()
    # 公开链接下载令牌
    dl_token = ''
    dl_token_expires = 0
    if not has_passcode:
        dl_token, dl_token_expires = _generate_download_token(link_id)
    return render_template('collect.html',
        link_id=link_id,
        task_title=link['title'],
        description=link['description'],
        verified=verified,
        has_passcode=has_passcode,
        passcode_empty=passcode_empty,
        require_uploader=require_uploader,
        uploader_name=uploader_name,
        in_wechat=is_wechat_browser(),
        max_file_size_gb=link['max_file_size_gb'],
        max_files=link['max_files'],
        allow_delete=bool(link['allow_delete']),
        allow_preview_download=bool(link['allow_preview_download']),
        site_title=get_user_setting(link_owner_id, 'site_title', '文件收集器'),
        collect_footer_text=get_user_setting(link_owner_id, 'collect_footer_text', ''),
        public_url=_resolve_public_url(link_owner_id),
        version=VERSION,
        passcode_ttl_display=ttl_display,
        expire_display=expire_display,
        expire_days_left=expire_days_left,
        expire_text=expire_text,
        expire_level=expire_level,
        creator_name=creator_name,
        csrf_token=csrf_token,
        dl_token=dl_token,
        dl_token_expires=dl_token_expires,
        attachment_name=link.get('attachment_name', ''),
        attachment_size=link.get('attachment_size', 0),
        attachment_size_display=format_file_size(link['attachment_size']) if link.get('attachment_size') else '',
        attachment_can_preview=_can_preview_attachment_ext(link.get('attachment_name', '')),
        attachment_preview_type=_attachment_preview_type(link.get('attachment_name', '')),
        blocked_extensions=sorted(list(get_blocked_extensions())), upload_batch_limit=get_upload_batch_limit(link.get('user_id')))

@app.route('/collect/<link_id>/verify', methods=['POST'])
def verify_passcode(link_id):
    """验证上传通行证"""
    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败，请刷新页面重试'}), 403

    client_ip = _get_client_ip()
    if not rate_limit(f'verify_{link_id}_{client_ip}', max_attempts=5, window_seconds=60):
        return jsonify({'success': False, 'message': '验证过于频繁，请1分钟后再试'}), 429

    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (collect_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()
    conn.close()

    if not link:
        return jsonify({'success': False, 'message': '链接不存在或已失效'}), 404

    # 统一使用数据库规范 ID（支持自定义 slug 访问）
    link_id = link['id']

    has_passcode = bool(link['passcode_plain'] and link['passcode_plain'].strip())
    # 空通行证：直接放行
    if not has_passcode:
        session[f'verified_{link_id}'] = time.time()
        return jsonify({'success': True})

    passcode = request.form.get('passcode', '').strip()
    if passcode and check_password_hash(link['passcode'], passcode):
        session[f'verified_{link_id}'] = time.time()
        return jsonify({'success': True})
    return jsonify({'success': False, 'message': '通行证错误'}), 403

@app.route('/collect/<link_id>/logout', methods=['POST'])
def logout_passcode(link_id):
    """退出通行证，清除当前链接的验证缓存"""
    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败'}), 403
    # 解析链接（支持自定义 slug）
    conn = get_db()
    link = conn.execute(
        "SELECT id FROM links WHERE (collect_slug = ? OR id = ?)", (link_id, link_id)
    ).fetchone()
    conn.close()
    if link:
        link_id = link['id']
    session.pop(f'verified_{link_id}', None)
    return jsonify({'success': True, 'message': '已退出通行证'})

def _sanitize_uploader_name(name):
    """安全化上传者名称，去除危险字符，用于文件夹名"""
    # 去除路径遍历字符和危险特殊字符
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', name.strip())
    # 去除首尾的点和空格
    name = name.strip('. ')
    # 限制长度
    if len(name) > 60:
        name = name[:60]
    if not name:
        name = 'unknown'
    return name

@app.route('/collect/<link_id>/set_uploader', methods=['POST'])
def set_uploader(link_id):
    """设置上传者身份（require_uploader 开启时使用）"""
    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败'}), 403

    if not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$', link_id):
        return jsonify({'success': False, 'message': '链接格式无效'}), 400

    uploader_name = (request.form.get('uploader_name') or '').strip()
    if not uploader_name:
        return jsonify({'success': False, 'message': '请输入您的身份'}), 400

    # 字数限制
    chinese_count = 0
    other_count = 0
    for c in uploader_name:
        if '\u4e00' <= c <= '\u9fff' or '\u3400' <= c <= '\u4dbf':
            chinese_count += 1
        else:
            other_count += 1
    if chinese_count > 0:
        # 含中文：中文至少2字，最多8字；英文/数字/符号最多20字
        if chinese_count < 2:
            return jsonify({'success': False, 'message': '中文昵称至少填写2个字'}), 400
        if chinese_count > 8:
            return jsonify({'success': False, 'message': '中文字符最多8个'}), 400
        if other_count > 20:
            return jsonify({'success': False, 'message': '英文/数字/符号最多20个'}), 400
    else:
        # 纯英文/数字/符号：至少3字，最多20字
        if other_count < 3:
            return jsonify({'success': False, 'message': '昵称至少填写3个字符'}), 400
        if other_count > 20:
            return jsonify({'success': False, 'message': '英文/数字/符号最多20个'}), 400
    if chinese_count + other_count > 60:
        return jsonify({'success': False, 'message': '总字符数不能超过60个'}), 400

    # 安全化处理
    safe_name = _sanitize_uploader_name(uploader_name)

    conn = get_db()
    link = conn.execute(
        "SELECT id, passcode_plain, require_uploader FROM links WHERE (collect_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()
    conn.close()

    if not link:
        return jsonify({'success': False, 'message': '链接不存在或已失效'}), 404

    # 统一使用数据库规范 ID（支持自定义 slug 访问）
    link_id = link['id']

    if not bool(link['require_uploader']):
        return jsonify({'success': False, 'message': '该链接不需要上传者身份'}), 400

    # 无需二次验证通行证：能到达此步骤的用户已通过通行证验证（或无需通行证）
    session[f'uploader_{link_id}'] = safe_name
    return jsonify({'success': True, 'uploader_name': safe_name})

@app.route('/collect/<link_id>/logout_uploader', methods=['POST'])
def logout_uploader(link_id):
    """退出上传者身份"""
    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败'}), 403
    # 解析链接（支持自定义 slug）
    conn = get_db()
    link = conn.execute(
        "SELECT id FROM links WHERE (collect_slug = ? OR id = ?)", (link_id, link_id)
    ).fetchone()
    conn.close()
    if link:
        link_id = link['id']
    session.pop(f'uploader_{link_id}', None)
    return jsonify({'success': True, 'message': '已退出身份'})

# ============================================================
# 路由 - 分享页面（仅下载，无上传）
# ============================================================
@app.route('/share/<link_id>')
def share_page(link_id):
    """文件分享页面（仅查看和下载，无上传功能）"""
    if not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$', link_id):
        return render_template('error.html', error_code=400, error_title='链接格式无效', error_message='链接格式不合法'), 400
    conn = get_db()
    row = conn.execute(
        "SELECT * FROM links WHERE (share_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()
    conn.close()
    link = dict(row) if row else None

    if not link:
        return render_template('error.html',
            error_code=404,
            error_title='链接失效',
            error_message='链接不存在或已被停用'), 404

    # 统一使用数据库规范 ID（支持自定义 slug 访问）
    link_id = link['id']

    # 检查分享页是否启用
    if not link.get('share_enabled', 0):
        return render_template('error.html',
            error_code=404,
            error_title='分享未启用',
            error_message='该链接的分享功能未开启。'), 404

    # 分享页有效期：独立设置，不依赖收集页（留空表示永不过期）
    share_expires_at = link.get('share_expires_at')
    if share_expires_at:
        try:
            expire_time = datetime.strptime(share_expires_at, '%Y-%m-%dT%H:%M')
            if datetime.now() > expire_time:
                return render_template('error.html',
                    error_code=410,
                    error_title='链接已过期',
                    error_message='该分享链接已超过有效期。'), 410
        except (ValueError, TypeError):
            pass

    # 判断分享页有效通行证
    _sp = link.get('share_passcode', '')
    _spe = link.get('share_passcode_empty', 0)
    if _spe:
        # 分享页空通行证：无需验证
        has_passcode = False
    elif _sp and _sp.strip():
        # 分享页独立通行证
        has_passcode = True
    else:
        # 复用收集页通行证
        has_passcode = bool(link['passcode_plain'] and link['passcode_plain'].strip())

    # 使用独立的 share 验证 session key
    share_session_key = f'share_verified_{link_id}'
    if has_passcode:
        ts = session.get(share_session_key, 0)
        verified = isinstance(ts, (int, float)) and (time.time() - ts) < get_passcode_ttl()
    else:
        verified = True
    
    # 自动扫描文件夹中的新文件
    scan_link_folder(link_id)
    
    share_owner_id = link.get('user_id', '')
    ttl_minutes = int(get_user_setting(share_owner_id, 'passcode_ttl_minutes', '120'))
    if ttl_minutes >= 60 and ttl_minutes % 60 == 0:
        ttl_display = f'{ttl_minutes // 60} 小时'
    elif ttl_minutes >= 60:
        ttl_display = f'{ttl_minutes // 60} 小时 {ttl_minutes % 60} 分钟'
    else:
        ttl_display = f'{ttl_minutes} 分钟'

    expire_display = '永不过期'
    expire_days_left = -1
    expire_text = '有效期不限'
    expire_level = 'forever'
    if share_expires_at:
        try:
            expire_time = datetime.strptime(share_expires_at, '%Y-%m-%dT%H:%M')
            expire_display = expire_time.strftime('%Y-%m-%d %H:%M')
            remain = expire_time - datetime.now()
            expire_days_left = max(0, remain.days)
            total_seconds = remain.total_seconds()
            if total_seconds <= 0:
                expire_text = '已过期'
                expire_level = 'danger'
            elif total_seconds < 60:
                expire_text = '即将过期'
                expire_level = 'danger'
            elif total_seconds < 3600:
                expire_text = f'还剩 {int(total_seconds // 60)} 分钟'
                expire_level = 'danger'
            elif total_seconds < 86400:
                hour_s = int(total_seconds // 3600)
                min_s = int((total_seconds % 3600) // 60)
                expire_text = f'还剩 {hour_s} 小时' if min_s == 0 else f'还剩 {hour_s} 小时 {min_s} 分钟'
                expire_level = 'warn'
            else:
                days = remain.days
                expire_text = f'还剩 {days} 天'
                expire_level = 'normal' if days > 1 else 'warn'
        except (ValueError, TypeError):
            pass

    # 分享页描述：优先独立描述，否则回退到收集页描述
    share_description_text = link.get('share_description') or link.get('description', '')

    # 查询创建者名称（优先昵称）
    creator_name = ''
    try:
        conn2 = get_db()
        user_row = conn2.execute("SELECT username, nickname FROM users WHERE id = ?", (share_owner_id,)).fetchone()
        conn2.close()
        if user_row:
            creator_name = user_row['nickname'] or user_row['username']
    except Exception:
        pass

    csrf_token = generate_csrf_token()
    share_page_title = get_user_setting(share_owner_id, 'share_page_title', '')
    share_footer_text = get_user_setting(share_owner_id, 'share_footer_text', '')
    # 公开链接下载令牌
    dl_token = ''
    dl_token_expires = 0
    if not has_passcode:
        dl_token, dl_token_expires = _generate_download_token(link_id)
    return render_template('share.html',
        link_id=link_id,
        task_title=link['title'],
        description=share_description_text,
        verified=verified,
        has_passcode=has_passcode,
        in_wechat=is_wechat_browser(),
        site_title=get_user_setting(share_owner_id, 'site_title', '文件收集器'),
        share_page_title=share_page_title,
        collect_footer_text=get_user_setting(share_owner_id, 'collect_footer_text', ''),
        share_footer_text=share_footer_text,
        public_url=_resolve_public_url(share_owner_id),
        version=VERSION,
        passcode_ttl_display=ttl_display,
        expire_display=expire_display,
        expire_days_left=expire_days_left,
        expire_text=expire_text,
        expire_level=expire_level,
        creator_name=creator_name,
        require_uploader=bool(link.get('require_uploader', 0)),
        allow_preview_download=bool(link.get('allow_preview_download', 0)),
        csrf_token=csrf_token,
        dl_token=dl_token,
        dl_token_expires=dl_token_expires)

@app.route('/share/<link_id>/verify', methods=['POST'])
def share_verify_passcode(link_id):
    """分享页验证通行证（独立分享通行证 + 复用收集页通行证 fallback）"""
    if not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$', link_id):
        return jsonify({'success': False, 'message': '链接格式无效'}), 400
    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败，请刷新页面重试'}), 403

    client_ip = _get_client_ip()
    if not rate_limit(f'verify_{link_id}_{client_ip}', max_attempts=5, window_seconds=60):
        return jsonify({'success': False, 'message': '验证过于频繁，请1分钟后再试'}), 429

    conn = get_db()
    row = conn.execute(
        "SELECT * FROM links WHERE (share_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()
    conn.close()
    link = dict(row) if row else None

    if not link:
        return jsonify({'success': False, 'message': '链接不存在或已失效'}), 404

    # 统一使用数据库规范 ID（支持自定义 slug 访问）
    link_id = link['id']

    # 判断分享页有效通行证（与 share_page 逻辑一致）
    _sp = link.get('share_passcode', '')
    _spe = link.get('share_passcode_empty', 0)
    share_session_key = f'share_verified_{link_id}'

    if _spe:
        # 分享页空通行证：无需验证
        session[share_session_key] = time.time()
        return jsonify({'success': True})
    elif _sp and _sp.strip():
        # 分享页独立通行证
        passcode = request.form.get('passcode', '').strip()
        if passcode and check_password_hash(_sp, passcode):
            session[share_session_key] = time.time()
            return jsonify({'success': True})
        return jsonify({'success': False, 'message': '通行证错误'}), 403
    else:
        # 复用收集页通行证
        has_passcode = bool(link.get('passcode_plain', '') and link.get('passcode_plain', '').strip())
        if not has_passcode:
            session[share_session_key] = time.time()
            return jsonify({'success': True})
        passcode = request.form.get('passcode', '').strip()
        if passcode and check_password_hash(link.get('passcode', ''), passcode):
            session[share_session_key] = time.time()
            return jsonify({'success': True})
        return jsonify({'success': False, 'message': '通行证错误'}), 403

@app.route('/share/<link_id>/logout', methods=['POST'])
def share_logout_passcode(link_id):
    """分享页退出通行证"""
    if not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$', link_id):
        return jsonify({'success': False, 'message': '链接格式无效'}), 400
    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败'}), 403
    # 解析链接（支持自定义 slug）
    conn = get_db()
    link = conn.execute(
        "SELECT id FROM links WHERE (share_slug = ? OR id = ?)", (link_id, link_id)
    ).fetchone()
    conn.close()
    if link:
        link_id = link['id']
    session.pop(f'share_verified_{link_id}', None)
    return jsonify({'success': True, 'message': '已退出通行证'})

@app.route('/api/download-token/<link_id>', methods=['GET'])
def api_download_token(link_id):
    """刷新下载令牌（仅对公开链接有效）"""
    if not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$', link_id):
        return jsonify({'success': False, 'message': '链接格式无效'}), 400

    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (collect_slug = ? OR share_slug = ? OR id = ?) AND status = 'active'",
        (link_id, link_id, link_id)
    ).fetchone()
    conn.close()

    if not link:
        return jsonify({'success': False, 'message': '链接不存在'}), 404

    link_id = link['id']

    # 仅公开链接可获取令牌
    if not _is_link_public(link, for_share=True) and not _is_link_public(link, for_share=False):
        return jsonify({'success': False, 'message': '此链接无需令牌'}), 400

    token, expires = _generate_download_token(link_id)
    return jsonify({'success': True, 'token': token, 'expires': expires})


@app.route('/api/convert/docx/<link_id>/<int:record_id>', methods=['GET'])
def api_convert_docx(link_id, record_id):
    """将 docx 文件转换为 HTML"""
    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (collect_slug = ? OR share_slug = ? OR id = ?) AND status = 'active'",
        (link_id, link_id, link_id)
    ).fetchone()
    if not link:
        conn.close()
        return jsonify({'success': False, 'message': '链接不存在'}), 404

    link_id = link['id']
    if not is_verified(link_id, link) and not is_share_verified(link_id, link):
        conn.close()
        return jsonify({'success': False, 'message': '未授权'}), 403

    record = conn.execute(
        "SELECT stored_path, original_name FROM upload_records WHERE id = ? AND link_id = ?",
        (record_id, link_id)
    ).fetchone()
    conn.close()
    if not record:
        return jsonify({'success': False, 'message': '文件不存在'}), 404

    upload_base = get_upload_base()
    real_path = os.path.realpath(record['stored_path'])
    real_base = os.path.realpath(upload_base)
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        logger.warning(f"路径遍历攻击拦截: stored_path={record['stored_path']}")
        return jsonify({'success': False, 'message': '访问拒绝'}), 403
    if not os.path.isfile(real_path):
        return jsonify({'success': False, 'message': '文件不存在'}), 404

    ext = record['original_name'].split('.')[-1].lower()
    if ext not in ['doc', 'docx']:
        return jsonify({'success': False, 'message': '不支持的文件格式'}), 400

    try:
        import mammoth
        with open(real_path, 'rb') as f:
            result = mammoth.convert_to_html(f)
            html = result.value
            messages = result.messages
            return jsonify({'success': True, 'html': html, 'messages': [str(m) for m in messages]})
    except ImportError:
        return jsonify({'success': False, 'message': 'mammoth 库未安装'}), 500
    except Exception as e:
        logger.error(f"docx 转换失败: {e}")
        return jsonify({'success': False, 'message': '转换失败: ' + str(e)}), 500


@app.route('/api/convert/xlsx/<link_id>/<int:record_id>', methods=['GET'])
def api_convert_xlsx(link_id, record_id):
    """将 xlsx 文件转换为 HTML"""
    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (collect_slug = ? OR share_slug = ? OR id = ?) AND status = 'active'",
        (link_id, link_id, link_id)
    ).fetchone()
    if not link:
        conn.close()
        return jsonify({'success': False, 'message': '链接不存在'}), 404

    link_id = link['id']
    if not is_verified(link_id, link) and not is_share_verified(link_id, link):
        conn.close()
        return jsonify({'success': False, 'message': '未授权'}), 403

    record = conn.execute(
        "SELECT stored_path, original_name FROM upload_records WHERE id = ? AND link_id = ?",
        (record_id, link_id)
    ).fetchone()
    conn.close()
    if not record:
        return jsonify({'success': False, 'message': '文件不存在'}), 404

    upload_base = get_upload_base()
    real_path = os.path.realpath(record['stored_path'])
    real_base = os.path.realpath(upload_base)
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        logger.warning(f"路径遍历攻击拦截: stored_path={record['stored_path']}")
        return jsonify({'success': False, 'message': '访问拒绝'}), 403
    if not os.path.isfile(real_path):
        return jsonify({'success': False, 'message': '文件不存在'}), 404

    ext = record['original_name'].split('.')[-1].lower()
    if ext not in ['xls', 'xlsx']:
        return jsonify({'success': False, 'message': '不支持的文件格式'}), 400

    try:
        import openpyxl
        wb = openpyxl.load_workbook(real_path, read_only=True)
        html = '<table class="xlsx-table">'
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            html += f'<caption>{sheet_name}</caption>'
            html += '<thead><tr>'
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col).value
                html += f'<th>{escape_html(str(val) if val else "")}</th>'
            html += '</tr></thead><tbody>'
            for row in range(2, ws.max_row + 1):
                html += '<tr>'
                for col in range(1, ws.max_column + 1):
                    val = ws.cell(row=row, column=col).value
                    html += f'<td>{escape_html(str(val) if val else "")}</td>'
                html += '</tr>'
            html += '</tbody>'
        html += '</table>'
        wb.close()
        return jsonify({'success': True, 'html': html})
    except ImportError:
        return jsonify({'success': False, 'message': 'openpyxl 库未安装'}), 500
    except Exception as e:
        logger.error(f"xlsx 转换失败: {e}")
        return jsonify({'success': False, 'message': '转换失败: ' + str(e)}), 500

@app.route('/share/<link_id>/records', methods=['GET'])
def share_get_records(link_id):
    """获取分享页文件列表（支持分页）"""
    if not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$', link_id):
        return jsonify({'success': False, 'message': '链接格式无效'}), 400
    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (share_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()

    if not link:
        conn.close()
        return jsonify({'success': False, 'message': '链接不存在或已失效'}), 404

    # 统一使用数据库规范 ID（支持自定义 slug 访问）
    link_id = link['id']

    if not is_share_verified(link_id, link):
        conn.close()
        return jsonify({'success': False, 'message': '请先验证通行证'}), 403

    # 清理孤儿记录（文件已被删除的数据库记录）+ _attachment 脏记录
    cleanup_orphan_records_for_link(conn, link_id)
    conn.execute("DELETE FROM upload_records WHERE link_id = ? AND uploader_name = '_attachment'", (link_id,))

    require_uploader = bool(dict(link).get('require_uploader', 0))

    # 分页参数
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))
    except (ValueError, TypeError):
        page = 1
        per_page = 20
    page = max(1, page)
    per_page = max(5, min(per_page, 100))
    offset = (page - 1) * per_page
    
    # 按上传者过滤
    uploader_filter = request.args.get('uploader', '').strip()

    # 根据是否有上传者过滤来构建查询
    if uploader_filter:
        # 按上传者过滤时，支持分页
        total_uploaded = conn.execute(
            "SELECT COUNT(*) FROM upload_records WHERE link_id = ? AND uploader_name = ? AND uploader_name != '_attachment'",
            (link_id, uploader_filter)
        ).fetchone()[0]
        records = conn.execute(
            "SELECT id, original_name, file_size, file_size_display, uploaded_at, download_count, uploader_name FROM upload_records WHERE link_id = ? AND uploader_name = ? AND uploader_name != '_attachment' ORDER BY uploaded_at DESC LIMIT ? OFFSET ?",
            (link_id, uploader_filter, per_page, offset)
        ).fetchall()
        total_pages = max(1, (total_uploaded + per_page - 1) // per_page)
    else:
        total_uploaded = conn.execute(
            "SELECT COUNT(*) FROM upload_records WHERE link_id = ? AND uploader_name != '_attachment'", (link_id,)
        ).fetchone()[0]
        records = conn.execute(
            "SELECT id, original_name, file_size, file_size_display, uploaded_at, download_count, uploader_name FROM upload_records WHERE link_id = ? AND uploader_name != '_attachment' ORDER BY uploaded_at DESC LIMIT ? OFFSET ?",
            (link_id, per_page, offset)
        ).fetchall()
        total_pages = max(1, (total_uploaded + per_page - 1) // per_page)
    
    # 获取上传者统计（每个上传者的总文件数）
    uploader_stats = conn.execute(
        "SELECT uploader_name, COUNT(*) as count FROM upload_records WHERE link_id = ? AND uploader_name != '_attachment' GROUP BY uploader_name",
        (link_id,)
    ).fetchall()
    uploader_stats_dict = {row['uploader_name'] or '未署名': row['count'] for row in uploader_stats}
    conn.close()

    total_pages = max(1, (total_uploaded + per_page - 1) // per_page)

    return jsonify({
        'success': True,
        'require_uploader': require_uploader,
        'total_uploaded': total_uploaded,
        'total_pages': total_pages,
        'page': page,
        'per_page': per_page,
        'records': [dict(r) for r in records],
        'uploader_stats': uploader_stats_dict
    })

@app.route('/share/<link_id>/preview/<int:record_id>', methods=['GET'])
def share_preview_record(link_id, record_id):
    """预览分享页的文件（内联显示；支持 HEIC 转 JPEG）"""
    if not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$', link_id):
        return '链接格式无效', 400
    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (share_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()
    if not link:
        conn.close()
        return '链接不存在或已失效', 404

    # 统一使用数据库规范 ID（支持自定义 slug 访问）
    link_id = link['id']

    if not is_share_verified(link_id, link):
        conn.close()
        return '请先验证通行证', 403

    ok, err = _check_preview_rate_limit()
    if not ok:
        conn.close()
        return err, 429

    ok, err = _check_public_link_token(link_id, link, for_share=True)
    if not ok:
        conn.close()
        return err, 401

    if not link['allow_preview_download']:
        conn.close()
        return '预览功能未开启', 403

    record = conn.execute(
        "SELECT stored_path, original_name FROM upload_records WHERE id = ? AND link_id = ?",
        (record_id, link_id)
    ).fetchone()
    conn.close()
    if not record:
        return '文件不存在', 404

    upload_base = get_upload_base()
    real_path = os.path.realpath(record['stored_path'])
    real_base = os.path.realpath(upload_base)
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        logger.warning(f"路径遍历攻击拦截: stored_path={record['stored_path']}, upload_base={real_base}")
        abort(403)
    directory = os.path.dirname(real_path)
    filename = os.path.basename(real_path)
    if not os.path.isfile(real_path):
        abort(404)

    ext = os.path.splitext(record['original_name'])[1].lower()
    image_exts = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp', '.svg', '.heic', '.heif'}
    _IMAGE_MIME = {
        '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.png': 'image/png',
        '.gif': 'image/gif', '.webp': 'image/webp', '.bmp': 'image/bmp',
        '.svg': 'image/svg+xml', '.heic': 'image/heic', '.heif': 'image/heif'
    }
    pdf_ext = '.pdf'
    video_exts = {'.mp4', '.webm', '.ogg', '.mov', '.avi', '.mkv'}
    audio_exts = {'.mp3', '.wav', '.flac', '.aac', '.m4a'}

    if ext in image_exts:
        if ext in ('.heic', '.heif') and HEIC_SUPPORT:
            buf = _serve_heic_as_jpeg(real_path)
            if buf:
                resp = make_response(buf.read())
                resp.headers['Content-Type'] = 'image/jpeg'
                resp.headers['Cache-Control'] = 'public, max-age=3600'
                return resp
        return send_from_directory(directory, filename, mimetype=_IMAGE_MIME.get(ext, 'application/octet-stream'), as_attachment=False)
    elif ext == pdf_ext:
        return send_from_directory(directory, filename, mimetype='application/pdf', as_attachment=False)
    elif ext in video_exts:
        return send_from_directory(directory, filename, as_attachment=False)
    elif ext in audio_exts:
        return send_from_directory(directory, filename, as_attachment=False)
    else:
        return send_from_directory(directory, filename, as_attachment=True)





@app.route('/share/<link_id>/download/<int:record_id>', methods=['GET'])
def share_download_record(link_id, record_id):
    """分享页下载文件"""
    if not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$', link_id):
        return '链接格式无效', 400
    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (share_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()
    if not link:
        conn.close()
        return '链接不存在或已失效', 404

    # 统一使用数据库规范 ID（支持自定义 slug 访问）
    link_id = link['id']

    if not is_share_verified(link_id, link):
        conn.close()
        return '请先验证通行证', 403

    ok, err = _check_download_rate_limit()
    if not ok:
        conn.close()
        return err, 429

    ok, err = _check_public_link_token(link_id, link, for_share=True)
    if not ok:
        conn.close()
        return err, 401

    if not link['allow_preview_download']:
        conn.close()
        return '下载功能未开启', 403

    record = conn.execute(
        "SELECT stored_path, original_name FROM upload_records WHERE id = ? AND link_id = ?",
        (record_id, link_id)
    ).fetchone()

    if not record:
        conn.close()
        return '文件不存在', 404

    conn.execute("UPDATE upload_records SET download_count = download_count + 1 WHERE id = ?", (record_id,))
    conn.commit()
    conn.close()

    _log_download(record_id, 'share')
    return _safe_download(record['stored_path'], record['original_name'])

@app.route('/share/<link_id>/preview_file/<int:record_id>', methods=['GET'])
def share_preview_file(link_id, record_id):
    """分享页预览文件（用于 flyfish file-viewer 获取文件内容）"""
    if not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$', link_id):
        return '链接格式无效', 400
    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (share_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()
    if not link:
        conn.close()
        return '链接不存在或已失效', 404

    # 统一使用数据库规范 ID（支持自定义 slug 访问）
    link_id = link['id']

    if not is_share_verified(link_id, link):
        conn.close()
        return '请先验证通行证', 403

    ok, err = _check_preview_rate_limit()
    if not ok:
        conn.close()
        return err, 429

    ok, err = _check_public_link_token(link_id, link, for_share=True)
    if not ok:
        conn.close()
        return err, 401

    if not link['allow_preview_download']:
        conn.close()
        return '预览功能未开启', 403

    record = conn.execute(
        "SELECT stored_path, original_name FROM upload_records WHERE id = ? AND link_id = ?",
        (record_id, link_id)
    ).fetchone()

    if not record:
        conn.close()
        return '文件不存在', 404

    conn.close()

    upload_base = get_upload_base()
    real_path = os.path.realpath(record['stored_path'])
    real_base = os.path.realpath(upload_base)
    
    # 安全检查
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        logger.warning(f"路径遍历攻击拦截: stored_path={record['stored_path']}, upload_base={real_base}")
        abort(403)
    
    if not os.path.isfile(real_path):
        abort(404)

    directory = os.path.dirname(real_path)
    filename = os.path.basename(real_path)
    
    # 不使用 as_attachment，允许SDK读取文件内容
    response = send_from_directory(
        directory, filename,
        download_name=record['original_name'],
        as_attachment=False
    )
    
    # 添加CORS响应头，允许JIT Viewer SDK跨域访问
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    
    return response

@app.route('/share/<link_id>/txt_info/<int:record_id>', methods=['GET'])
def share_txt_info(link_id, record_id):
    """返回 TXT 文件章节目录和编码信息"""
    if not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$', link_id):
        return jsonify({'error': '链接格式无效'}), 400
    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (share_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()
    if not link:
        conn.close()
        return jsonify({'error': '链接不存在或已失效'}), 404
    link_id = link['id']
    if not is_share_verified(link_id, link):
        conn.close()
        return jsonify({'error': '请先验证通行证'}), 403
    ok, err = _check_preview_rate_limit()
    if not ok:
        conn.close()
        return jsonify({'error': err}), 429
    ok, err = _check_public_link_token(link_id, link, for_share=True)
    if not ok:
        conn.close()
        return jsonify({'error': err}), 401
    record = conn.execute(
        "SELECT stored_path, original_name FROM upload_records WHERE id = ? AND link_id = ?",
        (record_id, link_id)
    ).fetchone()
    if not record:
        conn.close()
        return jsonify({'error': '文件不存在'}), 404
    conn.close()
    upload_base = get_upload_base()
    real_path = os.path.realpath(record['stored_path'])
    real_base = os.path.realpath(upload_base)
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        abort(403)
    if not os.path.isfile(real_path):
        abort(404)
    chapters, encoding = _scan_txt_chapters(real_path)
    total_size = os.path.getsize(real_path)
    return jsonify({
        'chapters': chapters,
        'encoding': encoding,
        'total_size': total_size
    })


@app.route('/share/<link_id>/txt_chunk/<int:record_id>', methods=['GET'])
def share_txt_chunk(link_id, record_id):
    """返回 TXT 文件的文本块（分页用）"""
    if not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$', link_id):
        return jsonify({'error': '链接格式无效'}), 400
    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (share_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()
    if not link:
        conn.close()
        return jsonify({'error': '链接不存在或已失效'}), 404
    link_id = link['id']
    if not is_share_verified(link_id, link):
        conn.close()
        return jsonify({'error': '请先验证通行证'}), 403
    ok, err = _check_preview_rate_limit()
    if not ok:
        conn.close()
        return jsonify({'error': err}), 429
    ok, err = _check_public_link_token(link_id, link, for_share=True)
    if not ok:
        conn.close()
        return jsonify({'error': err}), 401
    record = conn.execute(
        "SELECT stored_path, original_name FROM upload_records WHERE id = ? AND link_id = ?",
        (record_id, link_id)
    ).fetchone()
    if not record:
        conn.close()
        return jsonify({'error': '文件不存在'}), 404
    conn.close()
    upload_base = get_upload_base()
    real_path = os.path.realpath(record['stored_path'])
    real_base = os.path.realpath(upload_base)
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        abort(403)
    if not os.path.isfile(real_path):
        abort(404)
    try:
        offset = int(request.args.get('offset', 0))
        size = min(int(request.args.get('size', 65536)), 524288)
    except (ValueError, TypeError):
        return jsonify({'error': '参数无效'}), 400
    text, file_size = _read_txt_chunk(real_path, offset, size)
    return jsonify({
        'text': text,
        'offset': offset,
        'size': len(text.encode('utf-8')),
        'total_size': file_size,
        'has_more': (offset + size) < file_size
    })


@app.route('/share/<link_id>/download_all', methods=['POST'])
def share_download_all(link_id):
    """分享页一键下载所有文件"""
    if not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$', link_id):
        return jsonify({'success': False, 'message': '链接格式无效'}), 400

    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (share_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()

    if not link:
        conn.close()
        return jsonify({'success': False, 'message': '链接不存在或已失效'}), 404

    link_id_db = link['id']

    if not is_share_verified(link_id_db, link):
        conn.close()
        return jsonify({'success': False, 'message': '请先验证通行证'}), 403

    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败，请刷新页面重试'}), 403

    # 查询该链接下所有文件（最多 200 个）
    records = conn.execute(
        "SELECT id, stored_path, original_name, file_size FROM upload_records WHERE link_id = ? ORDER BY uploaded_at ASC LIMIT 200",
        (link_id_db,)
    ).fetchall()
    conn.close()

    if not records:
        return jsonify({'success': False, 'message': '没有可下载的文件'}), 404

    import io, zipfile, tempfile

    link_title = (link['title'] or '').strip()
    safe_title = re.sub(r'[<>:"/\\|?*]', '_', link_title) if link_title else link_id
    zip_filename = f'{safe_title}.zip'

    tmp = tempfile.NamedTemporaryFile(suffix='.zip', delete=False)
    tmp_path = tmp.name
    try:
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zf:
            used_names = {}
            for rec in records:
                stored_path = rec['stored_path']
                if not os.path.isfile(stored_path):
                    continue
                original_name = rec['original_name']
                arcname = f"{safe_title}/{original_name}"
                if arcname in used_names:
                    used_names[arcname] += 1
                    base, ext = os.path.splitext(arcname)
                    arcname = f"{base}({used_names[arcname]}){ext}"
                else:
                    used_names[arcname] = 0
                compress_type = _get_zip_compress_type(original_name)
                zf.write(stored_path, arcname=arcname, compress_type=compress_type)

        tmp.close()

        @after_this_request
        def cleanup_download_all_zip(response):
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
            return response

        return send_file(
            tmp_path,
            download_name=zip_filename,
            as_attachment=True,
            mimetype='application/zip'
        )

    except Exception as e:
        logger.error(f"share一键下载打包失败: {e}")
        try:
            tmp.close()
            os.unlink(tmp_path)
        except Exception:
            pass
        return jsonify({'success': False, 'message': '文件打包失败，请稍后重试'}), 500


@app.route('/share/<link_id>/download_group', methods=['POST'])
def share_download_group(link_id):
    """分享页按上传者打包下载"""
    if not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$', link_id):
        return jsonify({'success': False, 'message': '链接格式无效'}), 400

    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (share_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()

    if not link:
        conn.close()
        return jsonify({'success': False, 'message': '链接不存在或已失效'}), 404

    link_id_db = link['id']

    if not is_share_verified(link_id_db, link):
        conn.close()
        return jsonify({'success': False, 'message': '请先验证通行证'}), 403

    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败，请刷新页面重试'}), 403

    data = request.get_json(silent=True) or {}
    uploader_name = (data.get('uploader_name') or '').strip()
    if not uploader_name:
        conn.close()
        return jsonify({'success': False, 'message': '缺少上传者名称'}), 400

    records = conn.execute(
        "SELECT id, stored_path, original_name, file_size FROM upload_records WHERE link_id = ? AND uploader_name = ? ORDER BY uploaded_at ASC",
        (link_id_db, uploader_name)
    ).fetchall()
    conn.close()

    if not records:
        return jsonify({'success': False, 'message': '该上传者没有文件'}), 404

    import io, zipfile, tempfile

    safe_name = re.sub(r'[<>:"/\\|?*]', '_', uploader_name)
    zip_filename = f'{safe_name}.zip'

    tmp = tempfile.NamedTemporaryFile(suffix='.zip', delete=False)
    tmp_path = tmp.name
    try:
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zf:
            used_names = {}
            for rec in records:
                stored_path = rec['stored_path']
                if not os.path.isfile(stored_path):
                    continue
                original_name = rec['original_name']
                arcname = f"{safe_name}/{original_name}"
                if arcname in used_names:
                    used_names[arcname] += 1
                    base, ext = os.path.splitext(arcname)
                    arcname = f"{base}({used_names[arcname]}){ext}"
                else:
                    used_names[arcname] = 0
                compress_type = _get_zip_compress_type(original_name)
                zf.write(stored_path, arcname=arcname, compress_type=compress_type)

        tmp.close()

        @after_this_request
        def cleanup_download_group_zip(response):
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
            return response

        return send_file(
            tmp_path,
            download_name=zip_filename,
            as_attachment=True,
            mimetype='application/zip'
        )

    except Exception as e:
        logger.error(f"share按上传者打包失败: {e}")
        try:
            tmp.close()
            os.unlink(tmp_path)
        except Exception:
            pass
        return jsonify({'success': False, 'message': '文件打包失败，请稍后重试'}), 500


@app.route('/share/<link_id>/batch_download', methods=['POST'])
def share_batch_download(link_id):
    """分享页批量下载：将所选文件打包为 zip 返回"""
    if not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$', link_id):
        return jsonify({'success': False, 'message': '链接格式无效'}), 400

    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (share_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()

    if not link:
        conn.close()
        return jsonify({'success': False, 'message': '链接不存在或已失效'}), 404

    link_id_db = link['id']

    if not is_share_verified(link_id_db, link):
        conn.close()
        return jsonify({'success': False, 'message': '请先验证通行证'}), 403

    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败，请刷新页面重试'}), 403

    data = request.get_json(silent=True) or {}
    record_ids = data.get('record_ids', [])
    if not record_ids or not isinstance(record_ids, list):
        return jsonify({'success': False, 'message': '请提供要下载的记录ID列表'}), 400

    # 限制每次最多下载 50 个文件
    record_ids = record_ids[:50]

    records = conn.execute(
        "SELECT id, stored_path, original_name, file_size FROM upload_records WHERE id IN ({}) AND link_id = ?".format(
            ','.join('?' * len(record_ids))),
        record_ids + [link_id_db]
    ).fetchall()
    conn.close()

    if not records:
        return jsonify({'success': False, 'message': '未找到可下载的文件'}), 404

    import io, zipfile, tempfile

    # 获取链接标题，用于 zip 文件名和内部文件夹
    link_title = (link['title'] or '').strip()
    safe_title = re.sub(r'[<>:"/\\|?*]', '_', link_title) if link_title else link_id
    zip_filename = f'{safe_title}.zip'

    # 使用临时文件，避免大文件撑爆内存
    tmp = tempfile.NamedTemporaryFile(suffix='.zip', delete=False)
    tmp_path = tmp.name
    try:
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zf:
            used_names = {}
            for rec in records:
                stored_path = rec['stored_path']
                if not os.path.isfile(stored_path):
                    continue
                original_name = rec['original_name']
                arcname = f"{safe_title}/{original_name}"
                if arcname in used_names:
                    used_names[arcname] += 1
                    base, ext = os.path.splitext(arcname)
                    arcname = f"{base}({used_names[arcname]}){ext}"
                else:
                    used_names[arcname] = 0
                # 智能压缩：已压缩格式只存储不重新压缩
                compress_type = _get_zip_compress_type(original_name)
                zf.write(stored_path, arcname=arcname, compress_type=compress_type)

        tmp.close()

        @after_this_request
        def cleanup_share_zip(response):
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
            return response

        return send_file(
            tmp_path,
            download_name=zip_filename,
            as_attachment=True,
            mimetype='application/zip'
        )

    except Exception as e:
        logger.error(f"share批量下载打包失败: {e}")
        try:
            tmp.close()
            os.unlink(tmp_path)
        except Exception:
            pass
        return jsonify({'success': False, 'message': '文件打包失败，请稍后重试'}), 500

@app.route('/share/<link_id>/delete_record/<int:record_id>', methods=['POST'])
def share_delete_record(link_id, record_id):
    """分享页不允许删除文件"""
    if not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$', link_id):
        return jsonify({'success': False, 'message': '链接格式无效'}), 400
    return jsonify({'success': False, 'message': '分享页面不支持删除操作'}), 403

@app.route('/collect/<link_id>/records', methods=['GET'])
def get_upload_records(link_id):
    """获取上传历史记录（支持分页）"""
    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (collect_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()

    if not link:
        conn.close()
        return jsonify({'success': False, 'message': '链接不存在或已失效'}), 404

    # 统一使用数据库规范 ID（支持自定义 slug 访问）
    link_id = link['id']

    if not is_verified(link_id, link):
        conn.close()
        return jsonify({'success': False, 'message': '请先验证通行证'}), 403

    # 清理孤儿记录（文件已被删除的数据库记录）
    cleanup_orphan_records_for_link(conn, link_id)

    # 静默清理超过24小时的 .part 断点文件
    cleanup_stale_part_files(conn, link_id)

    require_uploader = bool(dict(link).get('require_uploader', 0))
    uploader_name = (session.get(f'uploader_{link_id}') or '').strip()

    # 分页参数
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))
    except (ValueError, TypeError):
        page = 1
        per_page = 20
    page = max(1, page)
    per_page = max(5, min(per_page, 100))
    offset = (page - 1) * per_page

    if require_uploader and uploader_name:
        # 上传者已设置身份 → 只看自己上传的文件
        total_uploaded = conn.execute(
            "SELECT COUNT(*) FROM upload_records WHERE link_id = ? AND uploader_name = ?", (link_id, uploader_name)
        ).fetchone()[0]
        records = conn.execute(
            "SELECT id, original_name, file_size, file_size_display, uploaded_at, download_count, uploader_name FROM upload_records WHERE link_id = ? AND uploader_name = ? ORDER BY uploaded_at DESC LIMIT ? OFFSET ?",
            (link_id, uploader_name, per_page, offset)
        ).fetchall()
    elif require_uploader:
        # 开启了上传者但未设置身份 → 什么也看不到
        total_uploaded = 0
        records = []
    else:
        total_uploaded = conn.execute(
            "SELECT COUNT(*) FROM upload_records WHERE link_id = ?", (link_id,)
        ).fetchone()[0]
        records = conn.execute(
            "SELECT id, original_name, file_size, file_size_display, uploaded_at, download_count, uploader_name FROM upload_records WHERE link_id = ? ORDER BY uploaded_at DESC LIMIT ? OFFSET ?",
            (link_id, per_page, offset)
        ).fetchall()
    conn.close()

    total_pages = max(1, (total_uploaded + per_page - 1) // per_page)

    return jsonify({
        'success': True,
        'allow_delete': bool(link['allow_delete']),
        'allow_preview_download': bool(link['allow_preview_download']),
        'max_files': link['max_files'],
        'total_uploaded': total_uploaded,
        'require_uploader': require_uploader,
        'page': page,
        'per_page': per_page,
        'total_pages': total_pages,
        'records': [dict(r) for r in records]
    })

@app.route('/collect/<link_id>/preview/<int:record_id>', methods=['GET'])
def preview_record(link_id, record_id):
    """预览上传的文件（内联显示；支持 HEIC 转 JPEG）"""
    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (collect_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()
    if not link:
        conn.close()
        return '链接不存在或已失效', 404

    # 统一使用数据库规范 ID（支持自定义 slug 访问）
    link_id = link['id']

    if not is_verified(link_id, link):
        conn.close()
        return '请先验证通行证', 403

    ok, err = _check_preview_rate_limit()
    if not ok:
        conn.close()
        return err, 429

    ok, err = _check_public_link_token(link_id, link, for_share=False)
    if not ok:
        conn.close()
        return err, 401

    if not link['allow_preview_download']:
        conn.close()
        return '预览功能未开启', 403

    record = conn.execute(
        "SELECT stored_path, original_name, uploader_name FROM upload_records WHERE id = ? AND link_id = ?",
        (record_id, link_id)
    ).fetchone()
    conn.close()
    if not record:
        return '文件不存在', 404

    # require_uploader 模式下只允许预览自己的文件
    if link['require_uploader']:
        uploader_name = (session.get(f'uploader_{link_id}') or '').strip()
        rec_uploader = (record['uploader_name'] or '').strip()
        if uploader_name and rec_uploader and rec_uploader != uploader_name:
            return '无权访问此文件', 403

    upload_base = get_upload_base()
    real_path = os.path.realpath(record['stored_path'])
    real_base = os.path.realpath(upload_base)
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        logger.warning(f"路径遍历攻击拦截: stored_path={record['stored_path']}, upload_base={real_base}")
        abort(403)
    directory = os.path.dirname(real_path)
    filename = os.path.basename(real_path)
    if not os.path.isfile(real_path):
        abort(404)

    ext = os.path.splitext(record['original_name'])[1].lower()
    image_exts = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp', '.svg', '.heic', '.heif'}
    _IMAGE_MIME = {
        '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.png': 'image/png',
        '.gif': 'image/gif', '.webp': 'image/webp', '.bmp': 'image/bmp',
        '.svg': 'image/svg+xml', '.heic': 'image/heic', '.heif': 'image/heif'
    }
    pdf_ext = '.pdf'
    video_exts = {'.mp4', '.webm', '.ogg', '.mov', '.avi', '.mkv'}
    audio_exts = {'.mp3', '.wav', '.flac', '.aac', '.m4a'}

    if ext in image_exts:
        if ext in ('.heic', '.heif') and HEIC_SUPPORT:
            buf = _serve_heic_as_jpeg(real_path)
            if buf:
                resp = make_response(buf.read())
                resp.headers['Content-Type'] = 'image/jpeg'
                resp.headers['Cache-Control'] = 'public, max-age=3600'
                return resp
        return send_from_directory(directory, filename, mimetype=_IMAGE_MIME.get(ext, 'application/octet-stream'), as_attachment=False)
    elif ext == pdf_ext:
        return send_from_directory(directory, filename, mimetype='application/pdf', as_attachment=False)
    elif ext in video_exts:
        return send_from_directory(directory, filename, as_attachment=False)
    elif ext in audio_exts:
        return send_from_directory(directory, filename, as_attachment=False)
    else:
        return send_from_directory(directory, filename, as_attachment=True)


@app.route('/collect/<link_id>/download/<int:record_id>', methods=['GET'])
def download_record(link_id, record_id):
    """下载上传历史记录中的文件"""
    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (collect_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()
    if not link:
        conn.close()
        return '链接不存在或已失效', 404

    # 统一使用数据库规范 ID（支持自定义 slug 访问）
    link_id = link['id']

    if not is_verified(link_id, link):
        conn.close()
        return '请先验证通行证', 403

    ok, err = _check_download_rate_limit()
    if not ok:
        conn.close()
        return err, 429

    ok, err = _check_public_link_token(link_id, link, for_share=False)
    if not ok:
        conn.close()
        return err, 401

    if not link['allow_preview_download']:
        conn.close()
        return '下载功能未开启', 403

    record = conn.execute(
        "SELECT stored_path, original_name, uploader_name FROM upload_records WHERE id = ? AND link_id = ?",
        (record_id, link_id)
    ).fetchone()

    if not record:
        conn.close()
        return '文件不存在', 404

    # require_uploader 模式下只允许下载自己的文件
    if link['require_uploader']:
        uploader_name = (session.get(f'uploader_{link_id}') or '').strip()
        rec_uploader = (record['uploader_name'] or '').strip()
        if uploader_name and rec_uploader and rec_uploader != uploader_name:
            conn.close()
            return '无权下载此文件', 403

    conn.execute("UPDATE upload_records SET download_count = download_count + 1 WHERE id = ?", (record_id,))
    conn.commit()
    conn.close()

    _log_download(record_id, 'collect')
    return _safe_download(record['stored_path'], record['original_name'])

@app.route('/collect/<link_id>/preview_file/<int:record_id>', methods=['GET'])
def collect_preview_file(link_id, record_id):
    """收集页预览文件（用于 flyfish file-viewer 获取文件内容）"""
    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (collect_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()
    if not link:
        conn.close()
        return '链接不存在或已失效', 404

    # 统一使用数据库规范 ID（支持自定义 slug 访问）
    link_id = link['id']

    if not is_verified(link_id, link):
        conn.close()
        return '请先验证通行证', 403

    ok, err = _check_preview_rate_limit()
    if not ok:
        conn.close()
        return err, 429

    ok, err = _check_public_link_token(link_id, link, for_share=False)
    if not ok:
        conn.close()
        return err, 401

    if not link['allow_preview_download']:
        conn.close()
        return '预览功能未开启', 403

    record = conn.execute(
        "SELECT stored_path, original_name, uploader_name FROM upload_records WHERE id = ? AND link_id = ?",
        (record_id, link_id)
    ).fetchone()

    if not record:
        conn.close()
        return '文件不存在', 404

    # require_uploader 模式下只允许预览自己的文件
    if link['require_uploader']:
        uploader_name = (session.get(f'uploader_{link_id}') or '').strip()
        rec_uploader = (record['uploader_name'] or '').strip()
        if uploader_name and rec_uploader and rec_uploader != uploader_name:
            conn.close()
            return '无权访问此文件', 403

    conn.close()

    upload_base = get_upload_base()
    real_path = os.path.realpath(record['stored_path'])
    real_base = os.path.realpath(upload_base)
    
    # 安全检查
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        logger.warning(f"路径遍历攻击拦截: stored_path={record['stored_path']}, upload_base={real_base}")
        abort(403)
    
    if not os.path.isfile(real_path):
        abort(404)

    directory = os.path.dirname(real_path)
    filename = os.path.basename(real_path)
    
    # 不使用 as_attachment，允许SDK读取文件内容
    response = send_from_directory(
        directory, filename,
        download_name=record['original_name'],
        as_attachment=False
    )
    
    # 添加CORS响应头，允许JIT Viewer SDK跨域访问
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    
    return response

@app.route('/collect/<link_id>/txt_info/<int:record_id>', methods=['GET'])
def collect_txt_info(link_id, record_id):
    """返回 TXT 文件章节目录和编码信息"""
    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (collect_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()
    if not link:
        conn.close()
        return jsonify({'error': '链接不存在或已失效'}), 404
    link_id = link['id']
    if not is_verified(link_id, link):
        conn.close()
        return jsonify({'error': '请先验证通行证'}), 403
    ok, err = _check_preview_rate_limit()
    if not ok:
        conn.close()
        return jsonify({'error': err}), 429
    ok, err = _check_public_link_token(link_id, link, for_share=False)
    if not ok:
        conn.close()
        return jsonify({'error': err}), 401
    if not link['allow_preview_download']:
        conn.close()
        return jsonify({'error': '预览功能未开启'}), 403
    record = conn.execute(
        "SELECT stored_path, original_name FROM upload_records WHERE id = ? AND link_id = ?",
        (record_id, link_id)
    ).fetchone()
    if not record:
        conn.close()
        return jsonify({'error': '文件不存在'}), 404
    conn.close()
    upload_base = get_upload_base()
    real_path = os.path.realpath(record['stored_path'])
    real_base = os.path.realpath(upload_base)
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        abort(403)
    if not os.path.isfile(real_path):
        abort(404)
    chapters, encoding = _scan_txt_chapters(real_path)
    total_size = os.path.getsize(real_path)
    return jsonify({
        'chapters': chapters,
        'encoding': encoding,
        'total_size': total_size
    })


@app.route('/collect/<link_id>/txt_chunk/<int:record_id>', methods=['GET'])
def collect_txt_chunk(link_id, record_id):
    """返回 TXT 文件的文本块（分页用）"""
    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (collect_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()
    if not link:
        conn.close()
        return jsonify({'error': '链接不存在或已失效'}), 404
    link_id = link['id']
    if not is_verified(link_id, link):
        conn.close()
        return jsonify({'error': '请先验证通行证'}), 403
    ok, err = _check_preview_rate_limit()
    if not ok:
        conn.close()
        return jsonify({'error': err}), 429
    ok, err = _check_public_link_token(link_id, link, for_share=False)
    if not ok:
        conn.close()
        return jsonify({'error': err}), 401
    if not link['allow_preview_download']:
        conn.close()
        return jsonify({'error': '预览功能未开启'}), 403
    record = conn.execute(
        "SELECT stored_path, original_name FROM upload_records WHERE id = ? AND link_id = ?",
        (record_id, link_id)
    ).fetchone()
    if not record:
        conn.close()
        return jsonify({'error': '文件不存在'}), 404
    conn.close()
    upload_base = get_upload_base()
    real_path = os.path.realpath(record['stored_path'])
    real_base = os.path.realpath(upload_base)
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        abort(403)
    if not os.path.isfile(real_path):
        abort(404)
    try:
        offset = int(request.args.get('offset', 0))
        size = min(int(request.args.get('size', 65536)), 524288)
    except (ValueError, TypeError):
        return jsonify({'error': '参数无效'}), 400
    text, file_size = _read_txt_chunk(real_path, offset, size)
    return jsonify({
        'text': text,
        'offset': offset,
        'size': len(text.encode('utf-8')),
        'total_size': file_size,
        'has_more': (offset + size) < file_size
    })


@app.route('/collect/<link_id>/delete_record/<int:record_id>', methods=['POST'])
def delete_upload_record(link_id, record_id):
    """删除单条上传记录及文件"""
    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (collect_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()

    if not link:
        conn.close()
        return jsonify({'success': False, 'message': '链接不存在或已失效'}), 404

    # 统一使用数据库规范 ID（支持自定义 slug 访问）
    link_id = link['id']

    if not is_verified(link_id, link):
        conn.close()
        return jsonify({'success': False, 'message': '请先验证通行证'}), 403

    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败，请刷新页面重试'}), 403

    if not link['allow_delete']:
        conn.close()
        return jsonify({'success': False, 'message': '该链接不允许删除文件'}), 403

    record = conn.execute(
        "SELECT id, stored_path, uploader_name FROM upload_records WHERE id = ? AND link_id = ?",
        (record_id, link_id)
    ).fetchone()

    if not record:
        conn.close()
        return jsonify({'success': False, 'message': '记录不存在'}), 404

    # require_uploader 模式下只允许删除自己的文件
    if link['require_uploader']:
        uploader_name = (session.get(f'uploader_{link_id}') or '').strip()
        rec_uploader = (record['uploader_name'] or '').strip()
        if uploader_name and rec_uploader and rec_uploader != uploader_name:
            conn.close()
            return jsonify({'success': False, 'message': '无法删除他人的文件'}), 403

    try:
        _safe_delete(record['stored_path'])
    except Exception:
        pass

    conn.execute("DELETE FROM upload_records WHERE id = ?", (record_id,))
    conn.commit()
    conn.close()

    return jsonify({'success': True, 'message': '已删除'})

@app.route('/collect/<link_id>/batch_delete', methods=['POST'])
def batch_delete_records(link_id):
    """批量删除上传记录及文件"""
    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (collect_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()

    if not link:
        conn.close()
        return jsonify({'success': False, 'message': '链接不存在或已失效'}), 404

    link_id_db = link['id']

    if not is_verified(link_id_db, link):
        conn.close()
        return jsonify({'success': False, 'message': '请先验证通行证'}), 403

    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败，请刷新页面重试'}), 403

    if not link['allow_delete']:
        conn.close()
        return jsonify({'success': False, 'message': '该链接不允许删除文件'}), 403

    data = request.get_json(silent=True) or {}
    record_ids = data.get('record_ids', [])
    if not record_ids or not isinstance(record_ids, list):
        return jsonify({'success': False, 'message': '请提供要删除的记录ID列表'}), 400

    # 限制每次最多删除 100 条
    record_ids = record_ids[:100]

    uploader_name = session.get(f'uploader_{link_id_db}', '').strip()
    require_uploader = bool(link['require_uploader'])

    deleted = 0
    skipped = 0
    for rid in record_ids:
        try:
            rid = int(rid)
        except (ValueError, TypeError):
            skipped += 1
            continue

        record = conn.execute(
            "SELECT id, stored_path, uploader_name FROM upload_records WHERE id = ? AND link_id = ?",
            (rid, link_id_db)
        ).fetchone()

        if not record:
            skipped += 1
            continue

        # require_uploader 模式下只允许删除自己的文件
        if require_uploader and uploader_name:
            rec_uploader = (record['uploader_name'] or '').strip()
            if rec_uploader and rec_uploader != uploader_name:
                skipped += 1
                continue

        try:
            _safe_delete(record['stored_path'])
        except Exception:
            pass

        conn.execute("DELETE FROM upload_records WHERE id = ?", (rid,))
        deleted += 1

    conn.commit()
    conn.close()

    return jsonify({
        'success': True,
        'message': f'成功删除 {deleted} 个文件' + (f'，跳过 {skipped} 个' if skipped else ''),
        'deleted': deleted,
        'skipped': skipped
    })

@app.route('/collect/<link_id>/delete_all', methods=['POST'])
def delete_all_records(link_id):
    """删除该链接下所有上传记录及文件"""
    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (collect_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
    ).fetchone()

    if not link:
        conn.close()
        return jsonify({'success': False, 'message': '链接不存在或已失效'}), 404

    link_id_db = link['id']

    if not is_verified(link_id_db, link):
        conn.close()
        return jsonify({'success': False, 'message': '请先验证通行证'}), 403

    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败，请刷新页面重试'}), 403

    if not link['allow_delete']:
        conn.close()
        return jsonify({'success': False, 'message': '该链接不允许删除文件'}), 403

    uploader_name = session.get(f'uploader_{link_id_db}', '').strip()
    require_uploader = bool(link['require_uploader'])

    records = conn.execute(
        "SELECT id, stored_path, uploader_name FROM upload_records WHERE link_id = ?",
        (link_id_db,)
    ).fetchall()

    deleted = 0
    skipped = 0
    for record in records:
        if require_uploader and uploader_name:
            rec_uploader = (record['uploader_name'] or '').strip()
            if rec_uploader and rec_uploader != uploader_name:
                skipped += 1
                continue

        try:
            _safe_delete(record['stored_path'])
        except Exception:
            pass

        conn.execute("DELETE FROM upload_records WHERE id = ?", (record['id'],))
        deleted += 1

    conn.commit()
    conn.close()

    return jsonify({
        'success': True,
        'message': f'成功删除 {deleted} 个文件' + (f'，跳过 {skipped} 个' if skipped else ''),
        'deleted': deleted,
        'skipped': skipped
    })

@app.route('/collect/<link_id>/upload', methods=['POST'])
def upload_file(link_id):
    """处理文件上传"""
    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败，请刷新页面重试'}), 403

    conn = None
    try:
        conn = get_db()
        link = conn.execute(
            "SELECT * FROM links WHERE (collect_slug = ? OR id = ?) AND status = 'active'", (link_id, link_id)
        ).fetchone()

        if not link:
            return jsonify({'success': False, 'message': '链接不存在或已失效'}), 404

        # 上传频率限制：基于链接 max_files 动态计算（最少 30，不超过 200）
        client_ip = _get_client_ip()
        link_max_files = link['max_files']
        if link_max_files > 0:
            rate_limit_max = max(30, min(link_max_files * 3, 200))
        else:
            rate_limit_max = 200
        if not rate_limit(f'upload_{link_id}_{client_ip}', max_attempts=rate_limit_max, window_seconds=60):
            return jsonify({'success': False, 'message': '上传过于频繁，请稍后再试'}), 429

        # 统一使用数据库规范 ID（支持自定义 slug 访问）
        link_id = link['id']

        # 空通行证直接放行，有通行证才需要验证
        if not is_verified(link_id, link):
            return jsonify({'success': False, 'message': '请先验证通行证'}), 403

        uploaded_files = request.files.getlist('file')
        if not uploaded_files:
            uploaded_files = request.files.getlist('files')

        if not uploaded_files:
            return jsonify({'success': False, 'message': '没有接收到文件'}), 400

        max_files = link['max_files']
        # max_files=0 表示不限制
        if max_files > 0 and len(uploaded_files) > max_files:
            return jsonify({
                'success': False,
                'message': f'单次最多上传 {max_files} 个文件'
            }), 400

        uploader_name = (session.get(f'uploader_{link_id}') or '').strip()

        # require_uploader 强制校验：后端拦截绕过前端检查的请求
        if link['require_uploader'] and not uploader_name:
            return jsonify({'success': False, 'message': '请先填写您的身份'}), 403

        # 检查是否已超过上传总数上限（max_files=0 不限制）
        # 开启上传者时：每个上传者独立配额；未开启时：全局配额
        if max_files > 0:
            if link['require_uploader']:
                current_count = conn.execute(
                    "SELECT COUNT(*) FROM upload_records WHERE link_id = ? AND uploader_name = ?",
                    (link_id, uploader_name)
                ).fetchone()[0]
            else:
                current_count = conn.execute(
                    "SELECT COUNT(*) FROM upload_records WHERE link_id = ?", (link_id,)
                ).fetchone()[0]
            if current_count >= max_files:
                return jsonify({
                    'success': False,
                    'message': f'已达到最大上传数 {max_files} 个，无法继续上传'
                }), 400
        upload_dir = create_upload_dir(link_id, uploader_name)
        max_size_bytes = round(link['max_file_size_gb'] * 1024 * 1024 * 1024)
        results = []

        for file in uploaded_files:
            if not file or not file.filename:
                continue

            result = {'filename': file.filename, 'success': True}

            if not allowed_file(file.filename):
                _, ext = os.path.splitext(file.filename.lower())
                result.update({'success': False, 'message': f'文件类型 {ext or "（无扩展名）"} 被禁止上传'})
                results.append(result)
                continue

            file.seek(0, os.SEEK_END)
            size = file.tell()
            file.seek(0)

            if size > max_size_bytes:
                limit_display = f'{link["max_file_size_gb"]} GB'
                if link['max_file_size_gb'] < 1:
                    limit_mb = round(link['max_file_size_gb'] * 1024, 1)
                    limit_display = f'{limit_mb} MB'
                result.update({
                    'success': False,
                    'message': f'文件大小 {format_file_size(size)} 超过限制（上限 {limit_display}）'
                })
                results.append(result)
                continue

            safe_name = safe_filename_unicode(file.filename)
            if not safe_name:
                safe_name = 'unnamed_file'
            stored_name = safe_name
            stored_path = os.path.join(upload_dir, stored_name)
            if os.path.exists(stored_path):
                result.update({
                    'success': False,
                    'message': f'「{file.filename}」已存在，请改名后上传'
                })
                results.append(result)
                continue

            file.save(stored_path)
            logger.info(f"文件已保存: {stored_path} ({format_file_size(size)})")

            conn.execute(
                """INSERT INTO upload_records
                   (link_id, user_id, original_name, stored_name, stored_path, file_size,
                    file_size_display, uploader_ip, uploader_name, uploaded_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (link_id, link['user_id'] or '', file.filename, stored_name, stored_path, size,
                 format_file_size(size), _get_client_ip(), uploader_name,
                 datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            )
            record_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            conn.commit()
            _log_upload(link_id, 'upload_complete', record_id=record_id,
                        details={'name': file.filename, 'size': size, 'stored': stored_name},
                        uploader_name=uploader_name)

            result['size'] = format_file_size(size)
            results.append(result)

        success_count = sum(1 for r in results if r['success'])
        fail_count = sum(1 for r in results if not r['success'])

        logger.info(f"上传完成: link={link_id}, 成功={success_count}, 失败={fail_count}")

        return jsonify({
            'success': True,
            'results': results,
            'summary': f'成功 {success_count} 个' + (f'，失败 {fail_count} 个' if fail_count else '')
        })

    except PermissionError as e:
        logger.error(f"上传目录无权限: link={link_id}, error={e}")
        return jsonify({'success': False, 'message': '请在飞牛应用设置中添加自定义文件夹的读写权限'}), 500
    except Exception as e:
        logger.error(f"上传异常: link={link_id}, error={e}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'message': '服务器内部错误，请稍后重试'}), 500

    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass

# ============================================================
# 路由 - Tus 协议断点续传上传
# ============================================================

TUS_VERSION = '1.0.0'
TUS_EXTENSIONS = 'creation,creation-with-upload,termination'
TUS_CHUNK_SIZE = 5 * 1024 * 1024  # 5MB per PATCH (前端分片大小)

# 保留旧常量用于清理旧数据
CHUNK_SIZE = 5 * 1024 * 1024

def _get_tus_dir(upload_dir):
    """获取 Tus 临时文件目录，确保路径安全"""
    tus_dir = os.path.join(upload_dir, '.tus')
    real_tus_dir = os.path.realpath(tus_dir)
    real_base = os.path.realpath(upload_dir)
    if not real_tus_dir.startswith(real_base + os.sep) and real_tus_dir != real_base:
        raise ValueError('Invalid tus path')
    os.makedirs(real_tus_dir, mode=0o755, exist_ok=True)
    return real_tus_dir

def _cleanup_abandoned_tus(upload_dir, max_age_hours=24):
    """清理超时的 Tus 临时文件"""
    tus_dir = os.path.join(upload_dir, '.tus')
    if not os.path.isdir(tus_dir):
        return
    now = time.time()
    cutoff = now - (max_age_hours * 3600)
    try:
        for fname in os.listdir(tus_dir):
            fpath = os.path.join(tus_dir, fname)
            if os.path.isfile(fpath):
                mtime = os.path.getmtime(fpath)
                if mtime < cutoff:
                    try:
                        os.remove(fpath)
                        logger.info(f"清理过期 Tus 临时文件: {fpath}")
                    except Exception:
                        pass
    except Exception:
        pass

def _cleanup_abandoned_chunks(upload_dir, max_age_hours=24):
    """清理超时的分片临时文件（兼容旧版 .chunks 目录）"""
    chunks_root = os.path.join(upload_dir, '.chunks')
    if not os.path.isdir(chunks_root):
        return
    now = time.time()
    cutoff = now - (max_age_hours * 3600)
    try:
        for uid_dir in os.listdir(chunks_root):
            uid_path = os.path.join(chunks_root, uid_dir)
            if os.path.isdir(uid_path):
                mtime = os.path.getmtime(uid_path)
                if mtime < cutoff:
                    try:
                        shutil.rmtree(uid_path)
                        logger.info(f"清理过期分片目录: {uid_path}")
                    except Exception:
                        pass
    except Exception:
        pass

def _cleanup_chunk_dir(chunk_dir):
    """安全删除分片目录（兼容旧版）"""
    try:
        if os.path.isdir(chunk_dir):
            shutil.rmtree(chunk_dir)
    except Exception:
        pass

def _get_chunk_dir(upload_dir, upload_id):
    """获取分片临时目录（兼容旧版）"""
    safe_uid = re.sub(r'[^\w\-]', '_', upload_id)
    chunk_dir = os.path.join(upload_dir, '.chunks', safe_uid)
    real_chunk_dir = os.path.realpath(chunk_dir)
    real_base = os.path.realpath(upload_dir)
    if not real_chunk_dir.startswith(real_base + os.sep) and real_chunk_dir != real_base:
        raise ValueError('Invalid chunk path')
    os.makedirs(real_chunk_dir, mode=0o755, exist_ok=True)
    return real_chunk_dir

def _tus_response(code=204, extra_headers=None):
    """构造 Tus 协议响应"""
    resp = make_response('', code)
    resp.headers['Tus-Resumable'] = TUS_VERSION
    if extra_headers:
        for k, v in extra_headers.items():
            resp.headers[k] = v
    return resp

def _tus_error(code, message):
    """构造 Tus 错误响应"""
    logger.info(f"Tus error {code}: {message} [link_id={request.view_args.get('link_id','?')}]")
    resp = make_response(json.dumps({'error': message}), code)
    resp.headers['Content-Type'] = 'application/json'
    resp.headers['Tus-Resumable'] = TUS_VERSION
    return resp

def _parse_tus_metadata(metadata_str):
    """解析 Tus Upload-Metadata 头，返回 dict（值已 base64 解码）"""
    result = {}
    if not metadata_str:
        return result
    import base64
    for pair in metadata_str.split(','):
        pair = pair.strip()
        if not pair:
            continue
        parts = pair.split(' ', 1)
        key = parts[0]
        val = ''
        if len(parts) > 1 and parts[1]:
            try:
                val = base64.b64decode(parts[1]).decode('utf-8')
            except Exception:
                val = ''
        result[key] = val
    return result

def _tus_resolve_link(link_id):
    """解析链接，返回 (link_dict, canonical_id) 或 (None, None)"""
    if not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$', link_id):
        return None, None
    conn = get_db()
    try:
        link = conn.execute(
            "SELECT * FROM links WHERE (collect_slug = ? OR id = ?) AND status = 'active'",
            (link_id, link_id)
        ).fetchone()
        if not link:
            return None, None
        return dict(link), link['id']
    finally:
        conn.close()

def _tus_complete_upload(conn, upload_id, link, link_id, uploader_name, temp_path, upload_dir, stored_name, original_name, total_size):
    """完成 Tus 上传：移动文件、写入记录、日志"""
    stored_path = os.path.join(upload_dir, stored_name)

    # 文件名冲突检查（创建会话时已预检，此处为并发兜底）
    if os.path.exists(stored_path):
        raise FileExistsError(f'「{original_name}」已存在，请改名后上传')

    # 移动临时文件到最终位置
    import shutil as _shutil
    _shutil.move(temp_path, stored_path)

    # 写入 upload_records
    cursor = conn.execute(
        """INSERT INTO upload_records
           (link_id, user_id, original_name, stored_name, stored_path, file_size,
            file_size_display, uploader_ip, uploader_name, uploaded_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (link_id, link.get('user_id', '') or '', original_name,
         stored_name, stored_path, total_size,
         format_file_size(total_size), _get_client_ip(), uploader_name,
         datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    )
    record_id = cursor.lastrowid

    # 标记完成
    conn.execute(
        "UPDATE tus_uploads SET status = 'completed', stored_path = ?, updated_at = CURRENT_TIMESTAMP WHERE upload_id = ?",
        (stored_path, upload_id)
    )
    conn.commit()

    # 记录日志
    _log_upload(link_id, 'tus_upload', record_id=record_id,
                details={'name': original_name, 'size': total_size, 'stored': stored_name},
                uploader_name=uploader_name)

    logger.info(f"Tus 上传完成: link={link_id}, file={stored_name}, record_id={record_id}")
    return record_id, stored_name, stored_path


@app.route('/collect/<link_id>/tus', methods=['OPTIONS'])
def tus_options(link_id):
    """Tus 协议能力发现"""
    return _tus_response(204, {
        'Tus-Version': TUS_VERSION,
        'Tus-Extension': TUS_EXTENSIONS,
        'Tus-Max-Size': str(round(2048 * 1024 * 1024 * 1024)),  # 2TB 上限
        'Access-Control-Allow-Methods': 'POST,HEAD,PATCH,DELETE,OPTIONS',
        'Access-Control-Allow-Headers': 'Tus-Resumable,Upload-Length,Upload-Offset,Upload-Metadata,Upload-Concat,Content-Type,X-CSRFToken,X-HTTP-Method-Override',
        'Access-Control-Max-Age': '86400',
    })


@app.route('/collect/<link_id>/tus', methods=['POST'])
def tus_create(link_id):
    """Tus 创建上传会话（POST）"""
    # 检查 Tus-Resumable 头
    tus_resumable = request.headers.get('Tus-Resumable', '')
    if not tus_resumable:
        return _tus_error(412, 'Missing Tus-Resumable header')
    if tus_resumable != TUS_VERSION:
        return _tus_error(412, f'Unsupported Tus version: {tus_resumable}')

    if not validate_csrf():
        return _tus_error(403, 'CSRF validation failed')

    link, canonical_id = _tus_resolve_link(link_id)
    if not link:
        return _tus_error(404, 'Link not found')
    link_id = canonical_id

    if not is_verified(link_id, link):
        return _tus_error(403, 'Passcode verification required')

    # 解析 Upload-Length
    upload_length_str = request.headers.get('Upload-Length', '')
    try:
        total_size = int(upload_length_str)
    except (ValueError, TypeError):
        return _tus_error(400, 'Invalid or missing Upload-Length')

    if total_size <= 0:
        return _tus_error(400, 'Invalid file size')

    # 解析 Upload-Metadata
    metadata = _parse_tus_metadata(request.headers.get('Upload-Metadata', ''))
    filename = metadata.get('filename', '')
    if not filename:
        return _tus_error(400, 'Missing filename in metadata')

    uploader_name = (session.get(f'uploader_{link_id}') or '').strip()

    # require_uploader 强制校验：后端拦截绕过前端检查的请求
    if link['require_uploader'] and not uploader_name:
        return _tus_error(403, '请先填写您的身份')

    conn = get_db()
    try:
        # 检查数量限制
        max_files = link['max_files']

        # 上传频率限制：基于链接 max_files 动态计算
        client_ip = _get_client_ip()
        if max_files > 0:
            tus_rate_limit_max = max(30, min(max_files * 3, 200))
        else:
            tus_rate_limit_max = 200
        if not rate_limit(f'tus_create_{link_id}_{client_ip}', max_attempts=tus_rate_limit_max, window_seconds=60):
            return _tus_error(429, '上传过于频繁，请稍后再试')

        if max_files > 0:
            if link['require_uploader']:
                current_count = conn.execute(
                    "SELECT COUNT(*) FROM upload_records WHERE link_id = ? AND uploader_name = ?",
                    (link_id, uploader_name)
                ).fetchone()[0]
            else:
                current_count = conn.execute(
                    "SELECT COUNT(*) FROM upload_records WHERE link_id = ?", (link_id,)
                ).fetchone()[0]
            if current_count >= max_files:
                return _tus_error(400, f'已达到最大上传数 {max_files} 个，无法继续上传')

        # 检查大小限制
        max_size_bytes = round(link['max_file_size_gb'] * 1024 * 1024 * 1024)
        if total_size > max_size_bytes:
            limit_display = f'{link["max_file_size_gb"]} GB'
            if link['max_file_size_gb'] < 1:
                limit_mb = round(link['max_file_size_gb'] * 1024, 1)
                limit_display = f'{limit_mb} MB'
            return _tus_error(413, f'文件大小 {format_file_size(total_size)} 超过限制（上限 {limit_display}）')

        # 检查文件类型
        if not allowed_file(filename):
            _, ext = os.path.splitext(filename.lower())
            return _tus_error(400, f'文件类型 {ext or "（无扩展名）"} 被禁止上传')

        safe_name = safe_filename_unicode(filename)
        if not safe_name:
            safe_name = 'unnamed_file'
        upload_dir = create_upload_dir(link_id, uploader_name)

        # 检查文件名是否已存在（提前拦截，避免传完大文件才发现冲突）
        if os.path.exists(os.path.join(upload_dir, safe_name)):
            return _tus_error(409, f'「{filename}」已存在，请改名后上传')

        # 清理过期临时文件
        _cleanup_abandoned_tus(upload_dir, max_age_hours=24)
        _cleanup_abandoned_chunks(upload_dir, max_age_hours=24)

        # 生成 upload_id
        upload_id = secrets.token_hex(16)

        # 检查并发上传会话数
        client_ip = _get_client_ip()
        active_count = conn.execute(
            """SELECT COUNT(*) FROM tus_uploads
               WHERE uploader_ip = ? AND status = 'uploading'
               AND updated_at > datetime('now', '-2 hours')""",
            (client_ip,)
        ).fetchone()[0]
        if active_count >= 10:
            return _tus_error(429, '同时上传的任务过多，请等待已有上传完成后再试')

        # 创建临时文件路径
        tus_dir = _get_tus_dir(upload_dir)
        safe_uid = re.sub(r'[^\w\-]', '_', upload_id)
        temp_path = os.path.join(tus_dir, f'{safe_uid}.part')

        # 创建空临时文件
        with open(temp_path, 'wb') as f:
            pass

        # 插入 DB 记录
        conn.execute(
            """INSERT INTO tus_uploads
               (upload_id, link_id, user_id, original_name, stored_name,
                total_size, uploaded_offset, status, temp_path,
                uploader_ip, uploader_name)
               VALUES (?, ?, ?, ?, ?, ?, 0, 'uploading', ?, ?, ?)""",
            (upload_id, link_id, link.get('user_id', '') or '', filename, safe_name,
             total_size, temp_path, client_ip, uploader_name)
        )
        conn.commit()

        _log_upload(link_id, 'tus_create',
                    details={'name': filename, 'size': total_size, 'upload_id': upload_id},
                    uploader_name=uploader_name)

        # 返回 201 + Location
        # 使用 url_for 生成 Location，确保在带路径前缀的反代场景下路径正确
        # （ProxyFix x_prefix=1 仅对 url_for 生效，手动拼接字符串会丢失前缀）
        location = url_for('tus_patch', link_id=link_id, upload_id=upload_id)
        return _tus_response(201, {
            'Location': location,
            'Upload-Expires': (datetime.now(timezone.utc).strftime('%a, %d %b %Y %H:%M:%S GMT')),
        })

    except sqlite3.OperationalError as e:
        logger.error(f"tus_create DB错误: {e}")
        return _tus_error(503, '服务器繁忙，请重试')
    except PermissionError as e:
        logger.error(f"tus_create 目录无权限: {e}")
        return _tus_error(500, '请在飞牛应用设置中添加自定义文件夹的读写权限')
    except Exception as e:
        logger.error(f"tus_create error: {type(e).__name__}: {e}\n{traceback.format_exc()}")
        return _tus_error(500, f'创建失败: {type(e).__name__}')
    finally:
        conn.close()


@app.route('/collect/<link_id>/tus/<upload_id>', methods=['HEAD'])
def tus_head(link_id, upload_id):
    """Tus 查询上传偏移（HEAD）"""
    tus_resumable = request.headers.get('Tus-Resumable', '')
    if not tus_resumable or tus_resumable != TUS_VERSION:
        return _tus_error(412, 'Unsupported Tus version')

    link, canonical_id = _tus_resolve_link(link_id)
    if not link:
        return _tus_error(404, 'Link not found')
    link_id = canonical_id

    if not is_verified(link_id, link):
        return _tus_error(403, 'Passcode verification required')

    conn = get_db()
    try:
        row = conn.execute(
            "SELECT * FROM tus_uploads WHERE upload_id = ? AND link_id = ?",
            (upload_id, link_id)
        ).fetchone()

        if not row:
            return _tus_error(404, 'Upload not found')

        if row['status'] == 'completed':
            return _tus_response(200, {
                'Upload-Offset': str(row['total_size']),
                'Upload-Length': str(row['total_size']),
                'Cache-Control': 'no-store',
            })

        if row['status'] == 'cancelled':
            return _tus_error(404, 'Upload was cancelled')

        # 返回实际文件大小作为 offset（而非 DB 存储值，确保部分写入也能正确恢复）
        temp_path = row['temp_path']
        actual_offset = 0
        if temp_path and os.path.exists(temp_path):
            actual_offset = os.path.getsize(temp_path)

        # 如果 DB offset 与实际不一致，更新 DB
        if actual_offset != row['uploaded_offset']:
            conn.execute(
                "UPDATE tus_uploads SET uploaded_offset = ?, updated_at = CURRENT_TIMESTAMP WHERE upload_id = ?",
                (actual_offset, upload_id)
            )
            conn.commit()

        return _tus_response(200, {
            'Upload-Offset': str(actual_offset),
            'Upload-Length': str(row['total_size']),
            'Cache-Control': 'no-store',
        })
    except Exception as e:
        logger.error(f"tus_head error: {e}")
        return _tus_error(500, 'Internal error')
    finally:
        conn.close()


@app.route('/collect/<link_id>/tus/<upload_id>', methods=['PATCH'])
def tus_patch(link_id, upload_id):
    """Tus 上传分片数据（PATCH）"""
    tus_resumable = request.headers.get('Tus-Resumable', '')
    if not tus_resumable or tus_resumable != TUS_VERSION:
        return _tus_error(412, 'Unsupported Tus version')

    if not validate_csrf():
        return _tus_error(403, 'CSRF validation failed')

    # 检查 Content-Type
    content_type = request.headers.get('Content-Type', '')
    if content_type != 'application/offset+octet-stream':
        return _tus_error(415, 'Unsupported Media Type, expected application/offset+octet-stream')

    link, canonical_id = _tus_resolve_link(link_id)
    if not link:
        return _tus_error(404, 'Link not found')
    link_id = canonical_id

    if not is_verified(link_id, link):
        return _tus_error(403, 'Passcode verification required')

    # 解析 Upload-Offset
    offset_str = request.headers.get('Upload-Offset', '')
    try:
        client_offset = int(offset_str)
    except (ValueError, TypeError):
        return _tus_error(400, 'Invalid or missing Upload-Offset')

    conn = get_db()
    try:
        row = conn.execute(
            "SELECT * FROM tus_uploads WHERE upload_id = ? AND link_id = ? AND status = 'uploading'",
            (upload_id, link_id)
        ).fetchone()

        if not row:
            # 检查是否已完成
            completed = conn.execute(
                "SELECT * FROM tus_uploads WHERE upload_id = ? AND link_id = ? AND status = 'completed'",
                (upload_id, link_id)
            ).fetchone()
            if completed:
                return _tus_response(204, {
                    'Upload-Offset': str(completed['total_size']),
                })
            return _tus_error(404, 'Upload not found or already finished')

        temp_path = row['temp_path']
        total_size = row['total_size']

        # 获取实际文件大小
        actual_offset = 0
        if os.path.exists(temp_path):
            actual_offset = os.path.getsize(temp_path)

        # 验证 offset 匹配
        if client_offset != actual_offset:
            logger.warning(f"tus_patch offset mismatch: client={client_offset} actual={actual_offset} upload_id={upload_id[:16]}...")
            return _tus_error(409, f'Offset mismatch: client sent {client_offset}, actual is {actual_offset}')

        # 检查是否会超出总大小
        content_length = request.content_length or 0
        if actual_offset + content_length > total_size:
            return _tus_error(413, 'Upload exceeds declared size')

        # 追加写入数据（流式读取，避免大内存占用）
        bytes_written = 0
        try:
            with open(temp_path, 'ab') as f:
                while True:
                    buf = request.stream.read(65536)
                    if not buf:
                        break
                    f.write(buf)
                    bytes_written += len(buf)
        except IOError as e:
            logger.error(f"tus_patch 写入失败: {e}")
            return _tus_error(500, 'File write error')

        new_offset = actual_offset + bytes_written

        # 更新 DB
        conn.execute(
            "UPDATE tus_uploads SET uploaded_offset = ?, updated_at = CURRENT_TIMESTAMP WHERE upload_id = ?",
            (new_offset, upload_id)
        )
        conn.commit()

        # 检查是否上传完成
        if new_offset >= total_size:
            uploader_name = (session.get(f'uploader_{link_id}') or '').strip()
            upload_dir = create_upload_dir(link_id, uploader_name)
            stored_name = row['stored_name']
            original_name = row['original_name']

            try:
                record_id, final_stored_name, stored_path = _tus_complete_upload(
                    conn, upload_id, link, link_id, uploader_name,
                    temp_path, upload_dir, stored_name, original_name, total_size
                )
            except FileExistsError as e:
                # 文件名冲突：回滚状态，提示用户改名后重新上传
                logger.warning(f"tus_patch 文件名冲突: {e}")
                conn.execute(
                    "UPDATE tus_uploads SET status = 'uploading', updated_at = CURRENT_TIMESTAMP WHERE upload_id = ?",
                    (upload_id,)
                )
                conn.commit()
                return _tus_error(409, str(e))
            except Exception as e:
                logger.error(f"tus_patch 完成上传失败: {e}\n{traceback.format_exc()}")
                # 回滚状态
                conn.execute(
                    "UPDATE tus_uploads SET status = 'uploading', updated_at = CURRENT_TIMESTAMP WHERE upload_id = ?",
                    (upload_id,)
                )
                conn.commit()
                return _tus_error(500, f'完成上传失败: {type(e).__name__}')

            return _tus_response(204, {
                'Upload-Offset': str(new_offset),
            })

        return _tus_response(204, {
            'Upload-Offset': str(new_offset),
        })

    except sqlite3.OperationalError as e:
        logger.error(f"tus_patch DB错误: {e}")
        return _tus_error(503, '服务器繁忙，请重试')
    except PermissionError as e:
        logger.error(f"tus_patch 目录无权限: {e}")
        return _tus_error(500, '请在飞牛应用设置中添加自定义文件夹的读写权限')
    except Exception as e:
        logger.error(f"tus_patch error: {type(e).__name__}: {e}\n{traceback.format_exc()}")
        return _tus_error(500, f'上传失败: {type(e).__name__}')
    finally:
        conn.close()


@app.route('/collect/<link_id>/tus/<upload_id>', methods=['DELETE'])
def tus_delete(link_id, upload_id):
    """Tus 取消上传（DELETE）"""
    tus_resumable = request.headers.get('Tus-Resumable', '')
    if not tus_resumable or tus_resumable != TUS_VERSION:
        return _tus_error(412, 'Unsupported Tus version')

    if not validate_csrf():
        return _tus_error(403, 'CSRF validation failed')

    link, canonical_id = _tus_resolve_link(link_id)
    if not link:
        return _tus_error(404, 'Link not found')
    link_id = canonical_id

    if not is_verified(link_id, link):
        return _tus_error(403, 'Passcode verification required')

    conn = get_db()
    try:
        row = conn.execute(
            "SELECT * FROM tus_uploads WHERE upload_id = ? AND link_id = ?",
            (upload_id, link_id)
        ).fetchone()

        if not row:
            return _tus_error(404, 'Upload not found')

        # 标记取消
        conn.execute(
            "UPDATE tus_uploads SET status = 'cancelled', updated_at = CURRENT_TIMESTAMP WHERE upload_id = ?",
            (upload_id,)
        )
        conn.commit()

        # 删除临时文件
        temp_path = row['temp_path']
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass

        return _tus_response(204)
    except Exception as e:
        logger.error(f"tus_delete error: {e}")
        return _tus_error(500, 'Cancel failed')
    finally:
        conn.close()

# ============================================================
# 路由 - 用户注册和邀请码管理
# ============================================================

@app.route('/register', methods=['GET', 'POST'])
def register():
    """用户注册（需要邀请码）"""
    allow_reg = get_setting('allow_registration', '0') == '1'
    if not allow_reg:
        flash('当前未开放注册')
        return redirect(url_for('admin_login'))

    if request.method == 'POST':
        if not validate_csrf():
            flash('安全验证失败')
            return render_template('register.html', allow_registration=True)

        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        email_addr = request.form.get('email', '').strip().lower()
        invite_code = request.form.get('invite_code', '').strip()

        # 验证用户名
        if not username or len(username) < 3 or len(username) > 32:
            flash('用户名必须在3-32个字符之间')
            return render_template('register.html', allow_registration=True)
        if not re.match(r'^[a-zA-Z0-9_]{3,20}$', username):
            flash('用户名需由3-20个字母、数字或下划线组成')
            return render_template('register.html', allow_registration=True)

        # 验证邮箱（必须填写）
        if not email_addr or '@' not in email_addr:
            flash('请输入有效的邮箱地址')
            return render_template('register.html', allow_registration=True)

        # 验证密码（8位以上且包含字母和数字）
        if len(password) < 8:
            flash('密码长度至少需要8位')
            return render_template('register.html', allow_registration=True)
        if not re.search(r'[a-zA-Z]', password) or not re.search(r'[0-9]', password):
            flash('密码必须同时包含字母和数字')
            return render_template('register.html', allow_registration=True)

        if password != confirm_password:
            flash('两次输入的密码不一致')
            return render_template('register.html', allow_registration=True)

        # 验证邀请码
        invite = validate_invite_code(invite_code)
        if not invite:
            flash('邀请码无效或已过期')
            return render_template('register.html', allow_registration=True)

        # 检查邮箱是否已被使用
        conn = get_db()
        existing_email = conn.execute("SELECT id FROM users WHERE email = ?", (email_addr,)).fetchone()
        conn.close()
        if existing_email:
            flash('该邮箱已被注册')
            return render_template('register.html', allow_registration=True)

        # 创建用户
        user_id = create_user(username, password, email_addr)
        if not user_id:
            flash('注册失败，请重试（用户名可能已存在）')
            return render_template('register.html', allow_registration=True)

        # 标记邀请码已使用
        mark_invite_code_used(invite_code, user_id)

        # 创建用户文件夹
        user_folder = get_user_folder(user_id)
        if user_folder:
            os.makedirs(user_folder, exist_ok=True)

        # 发送注册欢迎邮件
        try:
            site_title = get_setting('site_title', '文件收集器')
            smtp_config = get_smtp_config()
            if smtp_config['host'] and smtp_config['from_email']:
                welcome_body = f'''<!DOCTYPE html>
<html>
<head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f5f7fa;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f5f7fa;padding:30px 0;">
<tr><td align="center">
<table width="480" cellpadding="0" cellspacing="0" style="background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.06);">
  <tr><td style="background:#6366f1;padding:24px 32px;text-align:center;">
    <h2 style="color:#fff;margin:0;font-size:18px;">{site_title}</h2>
  </td></tr>
  <tr><td style="padding:32px;">
    <h3 style="margin:0 0 12px;color:#111827;font-size:15px;">🎉 欢迎注册 {site_title}</h3>
    <p style="color:#6b7280;font-size:14px;line-height:1.6;margin:0 0 20px;">亲爱的 <strong>{username}</strong>，您好！</p>
    <p style="color:#6b7280;font-size:14px;line-height:1.6;margin:0 0 20px;">您的账号已成功注册，以下是您的账户信息：</p>
    <div style="background:#f0f4ff;border-radius:8px;padding:16px 24px;margin-bottom:20px;">
      <table cellpadding="0" cellspacing="0" style="width:100%;">
        <tr><td style="padding:4px 0;color:#4b5563;font-size:14px;">用户名</td><td style="padding:4px 0;color:#111827;font-size:14px;font-weight:500;">{username}</td></tr>
        <tr><td style="padding:4px 0;color:#4b5563;font-size:14px;">邮箱</td><td style="padding:4px 0;color:#111827;font-size:14px;font-weight:500;">{email_addr}</td></tr>
        <tr><td style="padding:4px 0;color:#4b5563;font-size:14px;">注册时间</td><td style="padding:4px 0;color:#111827;font-size:14px;font-weight:500;">{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</td></tr>
      </table>
    </div>
    <p style="color:#6b7280;font-size:14px;line-height:1.6;margin:0 0 8px;">您现在可以使用此账号登录并使用文件收集服务。</p>
    <p style="color:#9ca3af;font-size:13px;line-height:1.6;margin:0;">如非本人操作，请联系管理员。</p>
  </td></tr>
  <tr><td style="border-top:1px solid #e5e7eb;padding:16px 32px;text-align:center;">
    <p style="color:#9ca3af;font-size:12px;margin:0;">此邮件由 {site_title} 系统自动发送</p>
  </td></tr>
</table>
</td></tr>
</table>
</body>
</html>'''
                send_email(email_addr, f'[{site_title}] 注册成功 - 欢迎加入', welcome_body)
                logger.info(f"注册欢迎邮件已发送至 {email_addr}")
        except Exception as e:
            logger.error(f"发送注册欢迎邮件失败: {e}")

        flash('注册成功，请登录')
        return redirect(url_for('admin_login'))

    return render_template('register.html', allow_registration=True)

@admin_required
@app.route('/admin/invite_codes')
def admin_invite_codes():
    """邀请码管理页面"""
    conn = get_db()
    codes = []
    
    # 检查表是否存在
    try:
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='invite_codes'")
        if cursor.fetchone():
            codes = conn.execute(
                """SELECT ic.*, u.username as used_by_name, creator.username as created_by_name
                   FROM invite_codes ic
                   LEFT JOIN users u ON ic.used_by = u.id
                   LEFT JOIN users creator ON ic.created_by = creator.id
                   ORDER BY ic.created_at DESC"""
            ).fetchall()
    except Exception as e:
        logger.error(f"查询邀请码失败: {e}")
    
    conn.close()
    
    codes_list = [dict(code) for code in codes]
    return render_template('admin_invite_codes.html', invite_codes=codes_list, now=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

@admin_required
@app.route('/admin/invite_codes/generate', methods=['POST'])
def generate_invite_code_route():
    """生成邀请码"""
    if not validate_csrf():
        flash('安全验证失败')
        return redirect(url_for('admin_invite_codes'))

    expire_days = int(request.form.get('expire_days', 7))
    created_by = session.get('user_id')
    
    code = generate_invite_code(created_by, expire_days)
    if code:
        flash(f'邀请码已生成: {code}')
    else:
        flash('邀请码生成失败')
    return redirect(url_for('admin_invite_codes'))

@admin_required
@app.route('/admin/users')
def admin_users():
    """用户管理页面"""
    conn = get_db()
    users = []
    
    # 检查表是否存在
    try:
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='users'")
        if cursor.fetchone():
            users = conn.execute("SELECT * FROM users ORDER BY created_at DESC").fetchall()
    except Exception as e:
        logger.error(f"查询用户失败: {e}")
    
    conn.close()
    
    users_list = [dict(user) for user in users]
    return render_template('admin_users.html', users=users_list)

@admin_required
@app.route('/admin/users/new', methods=['GET', 'POST'])
def admin_user_create():
    """新增用户"""
    if request.method == 'POST':
        if not validate_csrf():
            flash('安全验证失败')
            return redirect(url_for('admin_user_create'))

        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        email_addr = request.form.get('email', '').strip().lower()
        nickname = request.form.get('nickname', '').strip()
        is_admin = 1 if request.form.get('is_admin') == '1' else 0
        status = request.form.get('status', 'active')

        # 验证用户名
        if not username or len(username) < 3 or len(username) > 20:
            flash('用户名必须在3-20个字符之间')
            return render_template('admin_user_create.html', form_data=request.form)
        if not re.match(r'^[a-zA-Z0-9_]{3,20}$', username):
            flash('用户名需由3-20个字母、数字或下划线组成')
            return render_template('admin_user_create.html', form_data=request.form)

        # 验证邮箱
        if not email_addr or '@' not in email_addr:
            flash('请输入有效的邮箱地址')
            return render_template('admin_user_create.html', form_data=request.form)

        # 验证密码
        if len(password) < 8:
            flash('密码长度至少需要8位')
            return render_template('admin_user_create.html', form_data=request.form)
        if not re.search(r'[a-zA-Z]', password) or not re.search(r'[0-9]', password):
            flash('密码必须同时包含字母和数字')
            return render_template('admin_user_create.html', form_data=request.form)

        if password != confirm_password:
            flash('两次输入的密码不一致')
            return render_template('admin_user_create.html', form_data=request.form)

        if status not in ('active', 'inactive'):
            status = 'active'

        # 检查用户名是否已存在
        conn = get_db()
        try:
            existing = conn.execute(
                "SELECT id FROM users WHERE username = ?", (username,)
            ).fetchone()
            if existing:
                flash('用户名已存在')
                conn.close()
                return render_template('admin_user_create.html', form_data=request.form)

            user_id = str(uuid.uuid4())
            password_hash = generate_password_hash(password)
            # 清理昵称 XSS 风险
            safe_nickname = bleach.clean(nickname) if nickname else ''
            conn.execute(
                """INSERT INTO users (id, username, password_hash, email, nickname, is_admin, status)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (user_id, username, password_hash, email_addr, safe_nickname, is_admin, status)
            )
            conn.commit()
            conn.close()
            flash('用户创建成功')
            return redirect(url_for('admin_users'))
        except sqlite3.IntegrityError:
            flash('用户名已存在')
            conn.close()
            return render_template('admin_user_create.html', form_data=request.form)
        except Exception as e:
            logger.error(f"创建用户失败: {e}")
            flash('创建失败，请稍后重试')
            conn.close()
            return render_template('admin_user_create.html', form_data=request.form)

    return render_template('admin_user_create.html', form_data={})

@admin_required
@app.route('/admin/users/<user_id>/toggle', methods=['POST'])
def toggle_user_status(user_id):
    """切换用户状态"""
    if not validate_csrf():
        flash('安全验证失败')
        return redirect(url_for('admin_users'))

    conn = get_db()
    try:
        user = conn.execute("SELECT status FROM users WHERE id = ?", (user_id,)).fetchone()
        if user:
            new_status = 'inactive' if user['status'] == 'active' else 'active'
            conn.execute("UPDATE users SET status = ? WHERE id = ?", (new_status, user_id))
            conn.commit()
    except Exception as e:
        logger.error(f"切换用户状态失败: {e}")
        flash('操作失败')
    conn.close()
    
    flash('用户状态已更新')
    return redirect(url_for('admin_users'))

@admin_required
@app.route('/admin/users/<user_id>/edit', methods=['POST'])
def edit_user(user_id):
    """编辑用户信息"""
    if not validate_csrf():
        flash('安全验证失败')
        return redirect(url_for('admin_users'))
    
    new_username = request.form.get('new_username', '').strip()
    new_password = request.form.get('new_password', '').strip()
    new_email = request.form.get('new_email', '').strip().lower()
    
    if not new_username:
        flash('用户名不能为空')
        return redirect(url_for('admin_users'))
    
    conn = get_db()
    try:
        user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not user:
            flash('用户不存在')
            conn.close()
            return redirect(url_for('admin_users'))
        
        # 检查用户名冲突
        existing = conn.execute("SELECT id FROM users WHERE username = ? AND id != ?", (new_username, user_id)).fetchone()
        if existing:
            flash('用户名已存在')
            conn.close()
            return redirect(url_for('admin_users'))
        
        # 检查邮箱冲突
        if new_email:
            email_conflict = conn.execute("SELECT id FROM users WHERE email = ? AND id != ?", (new_email, user_id)).fetchone()
            if email_conflict:
                flash('该邮箱已被其他用户使用')
                conn.close()
                return redirect(url_for('admin_users'))
        
        if new_password:
            password_hash = generate_password_hash(new_password)
            conn.execute("UPDATE users SET username = ?, password_hash = ?, email = ? WHERE id = ?",
                        (new_username, password_hash, new_email, user_id))
        else:
            conn.execute("UPDATE users SET username = ?, email = ? WHERE id = ?",
                        (new_username, new_email, user_id))
        conn.commit()
        # 如果用户登录了，更新 session
        if session.get('user_id') == user_id:
            session['username'] = new_username
        flash('用户信息已更新')
    except Exception as e:
        logger.error(f"编辑用户失败: {e}")
        flash('操作失败')
    conn.close()
    return redirect(url_for('admin_users'))

@admin_required
@app.route('/admin/users/<user_id>/delete', methods=['POST'])
def delete_user(user_id):
    """删除用户"""
    if not validate_csrf():
        flash('安全验证失败')
        return redirect(url_for('admin_users'))
    
    conn = get_db()
    try:
        user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not user:
            flash('用户不存在')
            conn.close()
            return redirect(url_for('admin_users'))
        if user['is_admin']:
            flash('不能删除管理员账户')
            conn.close()
            return redirect(url_for('admin_users'))
        
        conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
        # 清理关联数据
        conn.execute("UPDATE invite_codes SET used_by = NULL WHERE used_by = ?", (user_id,))
        conn.execute("UPDATE links SET user_id = '' WHERE user_id = ?", (user_id,))
        conn.execute("UPDATE upload_records SET user_id = '' WHERE user_id = ?", (user_id,))
        conn.commit()
        flash('用户已删除')
    except Exception as e:
        logger.error(f"删除用户失败: {e}")
        flash('操作失败')
    conn.close()
    return redirect(url_for('admin_users'))

@admin_required
@app.route('/admin/users/<user_id>/reset-password', methods=['POST'])
def reset_user_password(user_id):
    """重置用户密码"""
    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败'})
    
    conn = get_db()
    try:
        user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not user:
            conn.close()
            return jsonify({'success': False, 'message': '用户不存在'})
        
        new_password = secrets.token_hex(4)  # 8位随机密码
        password_hash = generate_password_hash(new_password)
        conn.execute("UPDATE users SET password_hash = ? WHERE id = ?", (password_hash, user_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': f'密码已重置为: {new_password}', 'new_password': new_password})
    except Exception as e:
        logger.error(f"重置密码失败: {e}")
        conn.close()
        return jsonify({'success': False, 'message': '操作失败'})

# ============================================================
# 路由 - 管理员后台
# ============================================================
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """管理员/用户登录"""
    # 已登录用户直接进入后台 dashboard
    if session.get('user_id'):
        return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        if not validate_csrf():
            flash('安全验证失败，请刷新页面重试')
            login_tip = get_setting('login_tip', '')
            allow_reg = get_setting('allow_registration', '0') == '1'
            return render_template('admin_login.html', login_tip=login_tip, allow_registration=allow_reg)

        client_ip = _get_client_ip()
        if not rate_limit(f'login_{client_ip}', max_attempts=5, window_seconds=60):
            flash('登录尝试过于频繁，请稍后再试')
            login_tip = get_setting('login_tip', '')
            allow_reg = get_setting('allow_registration', '0') == '1'
            return render_template('admin_login.html', login_tip=login_tip, allow_registration=allow_reg)

        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        # 使用新的多用户认证
        user = validate_user_login(username, password)
        
        if user:
            session.clear()
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['is_admin'] = user['is_admin'] == 1
            session['nickname'] = (user.get('nickname') or '').strip()
            session.permanent = True
            
            # 更新最后登录时间
            conn = get_db()
            conn.execute("UPDATE users SET last_login_at = CURRENT_TIMESTAMP WHERE id = ?", (user['id'],))
            conn.commit()
            conn.close()
            
            # 所有登录用户进入后台 dashboard
            return redirect(url_for('admin_dashboard'))
        else:
            flash('用户名或密码错误')
            login_tip = get_setting('login_tip', '')
            allow_reg = get_setting('allow_registration', '0') == '1'
            return render_template('admin_login.html', login_tip=login_tip, allow_registration=allow_reg)

    login_tip = get_setting('login_tip', '')
    allow_reg = get_setting('allow_registration', '0') == '1'
    return render_template('admin_login.html', login_tip=login_tip, allow_registration=allow_reg)

# ============================================================
# 忘记密码 / 重置密码
# ============================================================
@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    """忘记密码 - 输入邮箱发送验证码"""
    if session.get('user_id'):
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        if not validate_csrf():
            flash('安全验证失败')
            return render_template('forgot_password.html')

        email_addr = request.form.get('email', '').strip().lower()
        if not email_addr or '@' not in email_addr:
            flash('请输入有效的邮箱地址')
            return render_template('forgot_password.html')

        # 频率限制：同一邮箱 60 秒内只能发一次
        client_ip = _get_client_ip()
        if not rate_limit(f'forgot_{client_ip}', max_attempts=3, window_seconds=60):
            flash('操作过于频繁，请稍后再试')
            return render_template('forgot_password.html')

        user = get_user_by_email(email_addr)
        if not user:
            # 兼容：管理员初始创建时 users 表 email 为空
            # 若输入邮箱与 SMTP 发件人邮箱一致，允许管理员重置密码
            smtp_config = get_smtp_config()
            if smtp_config['from_email'] and smtp_config['from_email'].lower() == email_addr:
                admin_user = get_setting('admin_username', 'admin')
                conn_admin = get_db()
                user = conn_admin.execute(
                    "SELECT * FROM users WHERE username = ? AND is_admin = 1 AND status = 'active'",
                    (admin_user,)
                ).fetchone()
                conn_admin.close()
                if user:
                    user = dict(user)
                    # 同步 email 到 users 表，下次直接命中
                    try:
                        conn_sync = get_db()
                        conn_sync.execute("UPDATE users SET email = ? WHERE id = ? AND (email = '' OR email IS NULL)",
                                         (email_addr, user['id']))
                        conn_sync.commit()
                        conn_sync.close()
                    except Exception:
                        pass
            if not user:
                flash('该用户暂未注册，请仔细检查邮箱地址')
                return render_template('forgot_password.html')

        # 检查 SMTP 是否已配置
        smtp_config = get_smtp_config()
        if not smtp_config['host'] or not smtp_config['from_email']:
            flash('系统邮件服务未配置，请联系管理员')
            return render_template('forgot_password.html')

        # 生成验证码
        code = ''.join(secrets.choice('0123456789') for _ in range(6))
        expires_at = datetime.now() + timedelta(minutes=10)

        conn = get_db()
        conn.execute(
            "INSERT INTO verification_codes (email, code, purpose, expires_at) VALUES (?, ?, 'reset_password', ?)",
            (email_addr, code, expires_at.strftime('%Y-%m-%dT%H:%M:%S'))
        )
        conn.commit()
        conn.close()

        site_title = get_setting('site_title', '文件收集器')
        body = f'''<!DOCTYPE html>
<html>
<head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f5f7fa;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;">
<table width="100%%" cellpadding="0" cellspacing="0" style="background:#f5f7fa;padding:30px 0;">
<tr><td align="center">
<table width="480" cellpadding="0" cellspacing="0" style="background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.06);">
  <tr><td style="background:#4f46e5;padding:24px 32px;text-align:center;">
    <h2 style="color:#fff;margin:0;font-size:18px;">{site_title}</h2>
  </td></tr>
  <tr><td style="padding:32px;">
    <h3 style="margin:0 0 12px;color:#111827;font-size:15px;">密码重置验证码</h3>
    <p style="color:#6b7280;font-size:14px;line-height:1.6;margin:0 0 20px;">您正在请求重置密码，请使用以下验证码完成验证：</p>
    <div style="background:#f0f4ff;border-radius:8px;padding:16px 24px;text-align:center;margin-bottom:20px;">
      <span style="font-size:28px;font-weight:700;letter-spacing:6px;color:#4f46e5;font-family:'Courier New',monospace;">{code}</span>
    </div>
    <p style="color:#9ca3af;font-size:13px;line-height:1.6;margin:0 0 8px;">验证码 10 分钟内有效，请勿泄露给他人。</p>
    <p style="color:#9ca3af;font-size:13px;line-height:1.6;margin:0;">如非本人操作，请忽略此邮件。</p>
  </td></tr>
  <tr><td style="border-top:1px solid #e5e7eb;padding:16px 32px;text-align:center;">
    <p style="color:#9ca3af;font-size:12px;margin:0;">此邮件由 {site_title} 系统自动发送</p>
  </td></tr>
</table>
</td></tr>
</table>
</body>
</html>'''
        ok, msg = send_email(email_addr, f'[{site_title}] 密码重置验证码', body)

        if ok:
            flash('验证码已发送到您的邮箱，请查收')
            return redirect(url_for('reset_password', email=email_addr))
        else:
            logger.error(f"发送验证码邮件失败: {msg}")
            flash(f'邮件发送失败：{msg}')
        return render_template('forgot_password.html')

    return render_template('forgot_password.html')


@app.route('/reset-password', methods=['GET', 'POST'])
def reset_password():
    """重置密码 - 验证码验证后设置新密码"""
    if session.get('user_id'):
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        if not validate_csrf():
            flash('安全验证失败')
            return render_template('reset_password.html')

        email_addr = request.form.get('email', '').strip().lower()
        code = request.form.get('code', '').strip()
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')

        if not all([email_addr, code, new_password, confirm_password]):
            flash('请填写所有字段')
            return render_template('reset_password.html', email=email_addr)

        # 验证验证码
        conn = get_db()
        vcode = conn.execute(
            "SELECT * FROM verification_codes WHERE email = ? AND code = ? "
            "AND purpose = 'reset_password' AND used = 0 AND expires_at > CURRENT_TIMESTAMP "
            "ORDER BY created_at DESC LIMIT 1",
            (email_addr, code)
        ).fetchone()
        if not vcode:
            conn.close()
            flash('验证码无效或已过期')
            return render_template('reset_password.html', email=email_addr)

        # 标记验证码已使用
        conn.execute("UPDATE verification_codes SET used = 1 WHERE id = ?", (vcode['id'],))
        conn.commit()

        # 验证密码
        if len(new_password) < 8:
            conn.close()
            flash('密码长度至少需要8位')
            return render_template('reset_password.html', email=email_addr)

        if not re.search(r'[a-zA-Z]', new_password) or not re.search(r'[0-9]', new_password):
            conn.close()
            flash('密码必须同时包含字母和数字')
            return render_template('reset_password.html', email=email_addr)

        if new_password != confirm_password:
            conn.close()
            flash('两次输入的密码不一致')
            return render_template('reset_password.html', email=email_addr)

        # 更新密码
        password_hash = generate_password_hash(new_password)
        conn.execute("UPDATE users SET password_hash = ? WHERE email = ?", (password_hash, email_addr))
        updated = conn.execute("SELECT changes()").fetchone()[0]
        if updated == 0:
            # 邮箱匹配失败：可能是管理员邮箱刚同步，尝试通过 SMTP 发件人邮箱匹配管理员
            smtp_config = get_smtp_config()
            admin_username = get_setting('admin_username', 'admin')
            if smtp_config['from_email'] and smtp_config['from_email'].lower() == email_addr:
                conn.execute(
                    "UPDATE users SET password_hash = ? WHERE username = ? AND is_admin = 1",
                    (password_hash, admin_username)
                )
                updated = conn.execute("SELECT changes()").fetchone()[0]
                if updated:
                    # 同时同步邮箱
                    conn.execute("UPDATE users SET email = ? WHERE username = ? AND is_admin = 1 AND (email = '' OR email IS NULL)",
                                 (email_addr, admin_username))
        conn.commit()
        conn.close()

        if updated == 0:
            flash('未找到匹配的账户，请确认邮箱地址是否正确')
            return render_template('reset_password.html', email=email_addr)

        flash('密码重置成功，请使用新密码登录')
        return redirect(url_for('admin_login'))

    email_addr = request.args.get('email', '')
    return render_template('reset_password.html', email=email_addr)




@app.route('/admin/user-settings', methods=['GET', 'POST'])
@login_required
def user_settings():
    """普通用户个人设置（继承管理员全局默认值，覆盖保存到 user_settings 表）"""
    user_id = session.get('user_id')
    is_admin = session.get('is_admin', False)
    
    # 管理员去看完整系统设置
    if is_admin:
        return redirect(url_for('admin_settings'))
    
    if request.method == 'POST':
        if not validate_csrf():
            flash('安全验证失败')
            return redirect(url_for('user_settings'))
        
        action = request.form.get('action', '')
        
        if action == 'defaults':
            max_files = request.form.get('default_max_files', str(DEFAULT_MAX_FILES))
            max_size = request.form.get('default_max_size', str(DEFAULT_MAX_FILE_SIZE_GB))
            try:
                _mf = float(max_files)
                if _mf != int(_mf):
                    raise ValueError('默认最大文件数必须为整数')
                max_files = int(_mf)
                max_size = round(float(max_size), 6)
                if max_files < 0:
                    raise ValueError('默认最大文件数不能为负数')
                if max_size < 0.01 or max_size > 64:
                    raise ValueError('单文件上限必须在 0.01-64 GB 之间')
            except ValueError as e:
                flash(str(e) if '必须' in str(e) else '默认值格式错误')
                return redirect(url_for('user_settings'))
            set_user_setting(user_id, 'max_files', str(max_files))
            set_user_setting(user_id, 'max_file_size_gb', str(max_size))
            flash('个人默认设置已保存')
        
        elif action == 'attachment_max':
            raw = request.form.get('attachment_max_mb', str(DEFAULT_ATTACHMENT_MAX_MB))
            try:
                _v = float(raw)
                if _v < 0.1:
                    raise ValueError('附件上限最小为 0.1 MB')
                mb = round(_v, 1)
            except ValueError as e:
                flash(str(e) if '最小' in str(e) else '附件上限格式错误')
                return redirect(url_for('user_settings'))
            set_user_setting(user_id, 'attachment_max_mb', str(mb))
            flash('附件大小上限已保存')
        
        elif action == 'passcode_ttl':
            minutes = request.form.get('passcode_ttl_minutes', '120')
            try:
                _m = float(minutes)
                if _m != int(_m):
                    raise ValueError('通行证有效期必须为整数')
                val = int(_m)
                if val < 1 or val > 43200:
                    raise ValueError('通行证有效期必须在 1-43200 分钟之间')
            except ValueError as e:
                flash(str(e) if '必须' in str(e) else '通行证有效期格式错误')
                return redirect(url_for('user_settings'))
            set_user_setting(user_id, 'passcode_ttl_minutes', str(val))
            flash('通行证有效期已保存')
        
        elif action == 'default_link_expiry':
            raw = request.form.get('default_link_expire_days', '30')
            try:
                _v = float(raw)
                if _v != int(_v):
                    raise ValueError('链接有效期天数必须为整数')
                days = int(_v)
                if days < 1 or days > 3650:
                    raise ValueError('链接有效期天数必须在 1-3650 天之间')
            except ValueError as e:
                flash(str(e) if '必须' in str(e) else '链接有效期格式错误')
                return redirect(url_for('user_settings'))
            set_user_setting(user_id, 'default_link_expire_days', str(days))
            flash('个人默认链接有效期已保存')

        elif action == 'links_per_page':
            raw = request.form.get('links_per_page_val', '50')
            try:
                _v = float(raw)
                if _v != int(_v):
                    raise ValueError('每页显示数量必须为整数')
                val = int(_v)
                if val < 5 or val > 100:
                    raise ValueError('每页显示数量必须在 5-100 之间')
            except ValueError as e:
                flash(str(e) if '必须' in str(e) else '每页数量格式错误')
                return redirect(url_for('user_settings'))
            set_user_setting(user_id, 'links_per_page', str(val))
            flash('收集链接显示数量已保存')

        elif action == 'records_per_page':
            raw = request.form.get('records_per_page_val', '50')
            try:
                _v = float(raw)
                if _v != int(_v):
                    raise ValueError('每页显示数量必须为整数')
                val = int(_v)
                if val < 5 or val > 200:
                    raise ValueError('每页显示数量必须在 5-200 之间')
            except ValueError as e:
                flash(str(e) if '必须' in str(e) else '每页数量格式错误')
                return redirect(url_for('user_settings'))
            set_user_setting(user_id, 'records_per_page', str(val))
            flash('上传记录单页数量已保存（仅对自己生效）')
        
        elif action == 'account':
            new_username = request.form.get('new_username', '').strip()
            new_nickname = request.form.get('new_nickname', '').strip()
            new_email = request.form.get('new_email', '').strip().lower()
            old_pass = request.form.get('old_password', '')
            new_pass = request.form.get('new_password', '')
            confirm_pass = request.form.get('confirm_password', '')
            user = get_user_by_id(user_id)
            if not user or not check_password_hash(user['password_hash'], old_pass):
                flash('原密码错误')
            elif not new_username:
                flash('用户名不能为空')
            else:
                conn = get_db()
                # 检查用户名冲突
                conflict = conn.execute("SELECT id FROM users WHERE username = ? AND id != ?",
                                       (new_username, user_id)).fetchone()
                if conflict:
                    flash('该用户名已被使用')
                    conn.close()
                    return redirect(url_for('user_settings'))
                
                # 邮箱校验和冲突检查
                if new_email:
                    if '@' not in new_email:
                        flash('请输入有效的邮箱地址')
                        conn.close()
                        return redirect(url_for('user_settings'))
                    conflict = conn.execute("SELECT id FROM users WHERE email = ? AND id != ?",
                                           (new_email, user_id)).fetchone()
                    if conflict:
                        flash('该邮箱已被其他用户使用')
                        conn.close()
                        return redirect(url_for('user_settings'))
                
                # 构建动态 UPDATE
                updates_sql = "UPDATE users SET username = ?"
                updates_params = [new_username]
                if new_nickname:
                    updates_sql += ", nickname = ?"
                    updates_params.append(new_nickname)
                if new_email:
                    updates_sql += ", email = ?"
                    updates_params.append(new_email)
                if new_pass:
                    if len(new_pass) < 8:
                        flash('新密码至少8位，且需包含字母和数字')
                        conn.close()
                        return redirect(url_for('user_settings'))
                    if not re.search(r'[a-zA-Z]', new_pass) or not re.search(r'[0-9]', new_pass):
                        flash('新密码必须同时包含字母和数字')
                        conn.close()
                        return redirect(url_for('user_settings'))
                    if new_pass != confirm_pass:
                        flash('两次密码不一致')
                        conn.close()
                        return redirect(url_for('user_settings'))
                    updates_sql += ", password_hash = ?"
                    updates_params.append(generate_password_hash(new_pass))
                
                updates_sql += " WHERE id = ?"
                updates_params.append(user_id)
                conn.execute(updates_sql, updates_params)
                conn.commit()
                conn.close()
                
                if new_nickname and not new_pass:
                    session['nickname'] = new_nickname
                
                if new_pass:
                    flash('密码修改成功，请重新登录')
                    session.clear()
                    return redirect(url_for('admin_login'))
                flash('账号信息已更新')
        
        return redirect(url_for('user_settings'))
    
    # GET: 构建 defaults dict（用户设置优先，回退到全局设置）
    defaults = {
        'max_files': get_user_setting(user_id, 'max_files', str(DEFAULT_MAX_FILES)),
        'max_file_size_gb': get_user_setting(user_id, 'max_file_size_gb', str(DEFAULT_MAX_FILE_SIZE_GB)),
        'attachment_max_mb': get_user_setting(user_id, 'attachment_max_mb', str(DEFAULT_ATTACHMENT_MAX_MB)),
        'passcode_ttl_minutes': get_user_setting(user_id, 'passcode_ttl_minutes', '120'),
        'default_link_expire_days': get_user_setting(user_id, 'default_link_expire_days', '30'),
        'links_per_page': get_user_setting(user_id, 'links_per_page', '50'),
        'records_per_page': get_user_setting(user_id, 'records_per_page', '50'),
        'public_url': _resolve_public_url(user_id),
    }
    user = get_user_by_id(user_id)
    return render_template('admin_user_settings.html', defaults=defaults, user=user)


@app.route('/admin/logout')
def admin_logout():
    """退出登录"""
    session.clear()
    return redirect(url_for('admin_login'))

@app.route('/admin')
@login_required
def admin_dashboard():
    """管理后台首页"""
    user_id = session.get('user_id')
    is_admin = session.get('is_admin', False)
    conn = get_db()
    
    if is_admin:
        total_links = conn.execute("SELECT COUNT(*) FROM links").fetchone()[0]
        active_links = conn.execute(
            "SELECT COUNT(*) FROM links WHERE status = 'active'"
        ).fetchone()[0]
        total_uploads = conn.execute("SELECT COUNT(*) FROM upload_records").fetchone()[0]
        today = datetime.now().strftime('%Y-%m-%d')
        today_uploads = conn.execute(
            "SELECT COUNT(*) FROM upload_records WHERE date(uploaded_at) = ?", (today,)
        ).fetchone()[0]
        recent = conn.execute(
            """SELECT r.*, l.title as link_title, l.require_uploader, u.username, u.nickname
               FROM upload_records r
               LEFT JOIN links l ON r.link_id = l.id
               LEFT JOIN users u ON l.user_id = u.id
               ORDER BY r.uploaded_at DESC LIMIT 5"""
        ).fetchall()
    else:
        total_links = conn.execute(
            "SELECT COUNT(*) FROM links WHERE user_id = ?", (user_id,)
        ).fetchone()[0]
        active_links = conn.execute(
            "SELECT COUNT(*) FROM links WHERE user_id = ? AND status = 'active'", (user_id,)
        ).fetchone()[0]
        # 普通用户只统计自己链接下的上传记录
        total_uploads = conn.execute(
            """SELECT COUNT(*) FROM upload_records r
               INNER JOIN links l ON r.link_id = l.id
               WHERE l.user_id = ?""", (user_id,)
        ).fetchone()[0]
        today = datetime.now().strftime('%Y-%m-%d')
        today_uploads = conn.execute(
            """SELECT COUNT(*) FROM upload_records r
               INNER JOIN links l ON r.link_id = l.id
               WHERE l.user_id = ? AND date(r.uploaded_at) = ?""", (user_id, today)
        ).fetchone()[0]
        recent = conn.execute(
            """SELECT r.*, l.title as link_title, l.require_uploader, u.username, u.nickname
               FROM upload_records r
               INNER JOIN links l ON r.link_id = l.id
               LEFT JOIN users u ON l.user_id = u.id
               WHERE l.user_id = ?
               ORDER BY r.uploaded_at DESC LIMIT 5""", (user_id,)
        ).fetchall()
    
    conn.close()

    # 构建 creator_name（优先昵称）
    recent_list = []
    for r in recent:
        r = dict(r)
        r['creator_name'] = r.get('nickname') or r.get('username') or '未知用户'
        recent_list.append(r)

    return render_template('admin_dashboard.html',
        total_links=total_links,
        active_links_count=active_links,
        total_uploads=total_uploads,
        today_uploads=today_uploads,
        recent_uploads=recent_list)

@app.route('/admin/links')
@login_required
def admin_links():
    """收集链接管理（支持分页和筛选）"""
    user_id = session.get('user_id')
    is_admin = session.get('is_admin', False)
    
    # 获取筛选参数（仅管理员可用）
    creator_filter = request.args.get('creator', '')
    
    # 获取每页显示数量
    per_page = int(get_user_setting(user_id, 'links_per_page', '50'))
    try:
        page = int(request.args.get('page', '1'))
        if page < 1:
            page = 1
    except ValueError:
        page = 1
    
    conn = get_db()
    # 明确指定需要的列（避免拉取不需要的列，但 description 在 data-* 属性中需要）
    _link_cols = ("l.id, l.user_id, l.title, l.description, l.passcode, l.passcode_plain, l.passcode_empty, "
                  "l.max_files, l.max_file_size_gb, l.expires_at, l.created_at, l.updated_at, "
                  "l.status, l.allow_delete, l.allow_preview_download, l.share_enabled, l.share_passcode, l.share_passcode_plain, l.share_passcode_empty, "
                  "l.require_uploader, l.collect_slug, l.share_slug, "
                  "u.username, u.nickname")
    
    # 构建筛选条件（仅管理员可用）
    creator_where = ""
    if is_admin and creator_filter:
        creator_where = " AND l.user_id = ?"
    
    # 获取所有创建人列表（仅管理员）
    creators = []
    if is_admin:
        creators = conn.execute(
            "SELECT id, username, nickname FROM users ORDER BY nickname, username"
        ).fetchall()
    
    if is_admin:
        if creator_filter:
            total = conn.execute(f"SELECT COUNT(*) FROM links l WHERE 1=1{creator_where}", (creator_filter,)).fetchone()[0]
        else:
            total = conn.execute("SELECT COUNT(*) FROM links l").fetchone()[0]
        total_pages = max(1, (total + per_page - 1) // per_page)
        page = min(page, total_pages)
        offset = (page - 1) * per_page
        if creator_filter:
            links = conn.execute(
                f"SELECT {_link_cols} FROM links l LEFT JOIN users u ON l.user_id = u.id WHERE 1=1{creator_where} ORDER BY l.created_at DESC LIMIT ? OFFSET ?",
                (creator_filter, per_page, offset)
            ).fetchall()
        else:
            links = conn.execute(
                f"SELECT {_link_cols} FROM links l LEFT JOIN users u ON l.user_id = u.id ORDER BY l.created_at DESC LIMIT ? OFFSET ?",
                (per_page, offset)
            ).fetchall()
    else:
        total = conn.execute(
            "SELECT COUNT(*) FROM links l WHERE l.user_id = ?", (user_id,)
        ).fetchone()[0]
        total_pages = max(1, (total + per_page - 1) // per_page)
        page = min(page, total_pages)
        offset = (page - 1) * per_page
        links = conn.execute(
            f"SELECT {_link_cols} FROM links l LEFT JOIN users u ON l.user_id = u.id WHERE l.user_id = ? ORDER BY l.created_at DESC LIMIT ? OFFSET ?",
            (user_id, per_page, offset)
        ).fetchall()
    conn.close()

    # 为每个链接标记通行证状态（使用 passcode_empty 字段，避免 bcrypt 哈希）
    processed_links = []
    for link in links:
        d = dict(link)
        pp = d['passcode_plain']
        is_empty = d.get('passcode_empty', 0) == 1
        # _has_passcode: 是否有明文通行证（非空字符串）
        d['_has_passcode'] = bool(pp and pp.strip())
        # _is_legacy: 旧版加密存储（passcode 不是空哈希，但没有明文）
        # passcode_empty=0 且无明文 = 旧版（未迁移的旧数据）
        d['_is_legacy'] = not d['_has_passcode'] and not is_empty
        # 分享页通行证状态
        sp = d.get('share_passcode_plain', '')
        sp_empty = d.get('share_passcode_empty', 0) == 1
        sp_hash = d.get('share_passcode', '')
        d['_has_share_passcode'] = bool(sp and sp.strip())
        # _is_share_legacy: 旧版加密存储（share_passcode 有哈希，但无明文，且不是空通行证）
        # 注意：share_passcode 本身就是空字符串时，表示从未设置，不算 legacy
        d['_is_share_legacy'] = not d['_has_share_passcode'] and not sp_empty and bool(sp_hash)
        # 创建者名称（优先昵称）
        d['creator_name'] = d.get('nickname') or d.get('username') or '未知用户'
        processed_links.append(d)

    return render_template('admin_links.html', links=processed_links,
                           public_url=_resolve_public_url(user_id),
                           page=page, total_pages=total_pages, total=total,
                           per_page=per_page,
                           creators=creators,
                           creator_filter=creator_filter,
                           csrf_token=generate_csrf_token())

@app.route('/admin/links/batch')
@login_required
def admin_link_batch():
    """批量创建收集链接页面"""
    return render_template('admin_link_batch.html', csrf_token=session.get('csrf_token', ''))


@app.route('/admin/links/template/download')
@login_required
def download_link_template():
    """下载批量创建收集链接的CSV模板"""
    import csv
    from io import StringIO
    
    # CSV字段定义
    headers = [
        '收集名称*',          # title - 必填
        '描述',               # description - 可选
        '通行证',             # passcode - 空表示无需验证
        '最大文件数量',       # max_files - 空表示默认值(0=不限制)
        '单文件上限(GB)',     # max_file_size_gb - 空表示默认值
        '有效期(天)',         # expire_days - 空表示永不过期
        '允许删除',           # allow_delete - 空或"否"表示否，"是"表示是
        '显示预览按钮',       # allow_preview_download - 空或"否"表示否，"是"表示是
        '上传者登记',         # require_uploader - 空或"否"表示否，"是"表示是
        '自定义收集链接',     # collect_slug - 可选，自定义收集页链接ID
        '启用分享页',         # share_enabled - 空或"否"表示否，"是"表示是
        '分享页描述',         # share_description - 分享页描述（可选）
        '分享页通行证',       # share_passcode - 空表示无需验证（启用分享页后生效）
        '分享页有效期(天)',   # share_expire_days - 空表示永不过期（启用分享页后生效）
        '自定义分享链接'      # share_slug - 可选，自定义分享页链接ID
    ]
    
    # 演示数据（不会被导入）
    demo_data = [
        [
            '示例收集链接（此行不会导入）',
            '这是一个演示数据，用于展示如何填写模板',
            '123456',
            '50',
            '2',
            '7',
            '是',
            '是',
            '否',
            'my-collect-link',
            '是',
            '这是分享页的描述信息',
            '654321',
            '7',
            'my-share-link'
        ],
        [
            '示例收集链接2（此行不会导入）',
            '不启用分享页的示例',
            '',
            '',
            '',
            '',
            '否',
            '否',
            '否',
            '',
            '否',
            '',
            '',
            '',
            ''
        ]
    ]
    
    # 使用 StringIO 写入CSV
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in demo_data:
        writer.writerow(row)
    
    output.seek(0)
    # 添加 UTF-8 BOM 确保Excel正确识别中文
    csv_content = '\ufeff' + output.getvalue()
    response = make_response(csv_content)
    response.headers['Content-Disposition'] = 'attachment; filename=collect_links_template.csv'
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    return response

@app.route('/admin/links/batch/import', methods=['POST'])
@login_required
def batch_import_links():
    """批量导入收集链接"""
    import csv
    from io import StringIO
    
    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败，请刷新页面重试'})
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '请选择要上传的文件'})
    
    file = request.files['file']
    if not file.filename:
        return jsonify({'success': False, 'message': '请选择要上传的文件'})
    
    ext = file.filename.split('.').pop().lower()
    if ext != 'csv':
        return jsonify({'success': False, 'message': '仅支持CSV格式。请下载模板后，用Excel打开编辑，保存时选择"CSV（逗号分隔）"格式。'})
    
    try:
        # 读取CSV内容（支持 UTF-8 BOM）
        raw = file.read()
        try:
            content = raw.decode('utf-8-sig')
        except UnicodeDecodeError:
            # 可能是 GBK 编码，尝试自动检测
            try:
                content = raw.decode('gbk')
            except Exception:
                return jsonify({'success': False, 'message': '文件编码无法识别，请确保使用 UTF-8 或 GBK 编码保存CSV文件。'})
        reader = csv.reader(StringIO(content))
        
        rows = list(reader)
        if len(rows) < 2:
            return jsonify({'success': False, 'message': '文件内容为空或只有表头'})
        
        headers = rows[0]
        data_rows = rows[1:]
        
        created_count = 0
        skipped_count = 0
        error_messages = []
        
        user_id = session.get('user_id')
        
        for row_idx, row in enumerate(data_rows):
            row_num = row_idx + 2  # 行号（从2开始，1是表头）
            
            # 跳过演示数据行
            if row and len(row) > 0 and '不会导入' in str(row[0]):
                skipped_count += 1
                continue
            
            # 验证必填字段
            if len(row) < 1 or not row[0] or not str(row[0]).strip():
                error_messages.append(f'第{row_num}行：收集名称不能为空')
                continue
            
            try:
                # 解析字段
                title = str(row[0]).strip() if len(row) > 0 else ''
                description = str(row[1]).strip() if len(row) > 1 else ''
                passcode = str(row[2]).strip() if len(row) > 2 else ''
                max_files = str(row[3]).strip() if len(row) > 3 else ''
                max_file_size_gb = str(row[4]).strip() if len(row) > 4 else ''
                expire_days = str(row[5]).strip() if len(row) > 5 else ''
                allow_delete = str(row[6]).strip() if len(row) > 6 else ''
                allow_preview_download = str(row[7]).strip() if len(row) > 7 else ''
                require_uploader = str(row[8]).strip() if len(row) > 8 else ''
                collect_slug = str(row[9]).strip() if len(row) > 9 else ''
                share_enabled = str(row[10]).strip() if len(row) > 10 else ''
                share_description = str(row[11]).strip() if len(row) > 11 else ''
                share_passcode = str(row[12]).strip() if len(row) > 12 else ''
                share_expire_days = str(row[13]).strip() if len(row) > 13 else ''
                share_slug = str(row[14]).strip() if len(row) > 14 else ''
                
                # 转换字段类型（带友好错误提示）
                try:
                    max_files = int(max_files) if max_files else DEFAULT_MAX_FILES
                except (ValueError, TypeError):
                    raise ValueError(f'"最大文件数量"不是有效数字，请填写整数')
                try:
                    max_file_size_gb = float(max_file_size_gb) if max_file_size_gb else DEFAULT_MAX_FILE_SIZE_GB
                except (ValueError, TypeError):
                    raise ValueError(f'"单文件上限(GB)"不是有效数字，请填写数字')
                allow_delete = 1 if allow_delete in ('是', '1', 'true', 'True') else 0
                allow_preview_download = 1 if allow_preview_download in ('是', '1', 'true', 'True') else 0
                require_uploader = 1 if require_uploader in ('是', '1', 'true', 'True') else 0
                share_enabled = 1 if share_enabled in ('是', '1', 'true', 'True') else 0
                
                # 验证数字范围
                if max_files < 0:
                    raise ValueError('最大文件数量不能为负数')
                if max_file_size_gb < 0.01 or max_file_size_gb > 64:
                    raise ValueError('单文件上限必须在 0.01-64 GB 之间')
                
                # 计算有效期
                _max_expire_days = int(get_user_setting(user_id, 'default_link_expire_days', '30'))
                expires_at = None
                if expire_days:
                    expires_at, err = _parse_expire_input(expire_days, '天', _max_expire_days)
                    if err:
                        raise ValueError(f'"有效期(天)"填写错误: {err}')
                
                # 生成链接ID
                link_id = generate_link_id()
                
                # 生成文件夹名
                folder_name = re.sub(r'[<>:"/\\|?*]', '_', title.strip())
                folder_name = folder_name.strip().lstrip('.')
                if not folder_name:
                    folder_name = link_id
                
                # 检查文件夹重名
                user_folder = get_user_folder(user_id)
                if user_folder:
                    base_name = folder_name
                    counter = 1
                    while os.path.exists(os.path.join(user_folder, folder_name)):
                        counter += 1
                        folder_name = f"{base_name}_{counter}"
                
                # 处理通行证
                if passcode:
                    passcode_hash = generate_password_hash(passcode)
                    passcode_plain = passcode
                    passcode_empty = 0
                else:
                    passcode_hash = generate_password_hash('')
                    passcode_plain = ''
                    passcode_empty = 1
                
                # 处理分享页通行证
                if share_passcode:
                    share_passcode_hash = generate_password_hash(share_passcode)
                    share_passcode_plain = share_passcode
                    share_passcode_empty = 0
                else:
                    share_passcode_hash = generate_password_hash('')
                    share_passcode_plain = ''
                    share_passcode_empty = 1
                
                # 处理分享页有效期
                share_expires_at = None
                if share_expire_days:
                    share_expires_at, err = _parse_expire_input(share_expire_days, '天', _max_expire_days)
                    if err:
                        raise ValueError(f'"分享页有效期(天)"填写错误: {err}')
                
                # 创建数据库记录
                conn = get_db()
                now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                conn.execute('''
                    INSERT INTO links (
                        id, user_id, title, description, passcode, passcode_plain, passcode_empty,
                        max_files, max_file_size_gb, expires_at, allow_delete, allow_preview_download,
                        require_uploader, collect_slug, share_enabled, share_description, 
                        share_passcode, share_passcode_plain, share_passcode_empty, share_expires_at, share_slug,
                        status, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active', ?, ?)
                ''', (
                    link_id, user_id, title, description, passcode_hash, passcode_plain, passcode_empty,
                    max_files, max_file_size_gb, expires_at, allow_delete, allow_preview_download,
                    require_uploader, collect_slug, share_enabled, share_description,
                    share_passcode_hash, share_passcode_plain, share_passcode_empty, share_expires_at, share_slug,
                    now_str, now_str
                ))
                conn.commit()
                conn.close()
                
                # 创建文件夹
                link_folder = get_link_folder(link_id, user_id)
                os.makedirs(link_folder, exist_ok=True)
                
                created_count += 1
                
            except ValueError as e:
                error_messages.append(f'第{row_num}行：{str(e)}')
            except Exception as e:
                logger.error(f'批量导入第{row_num}行失败: {e}\n{traceback.format_exc()}')
                error_messages.append(f'第{row_num}行：创建失败')
        
        # 构建结果消息
        message = f'成功创建 {created_count} 个收集链接'
        if skipped_count > 0:
            message += f'，跳过 {skipped_count} 行演示数据'
        if error_messages:
            message += f'，{len(error_messages)} 行失败'
        
        return jsonify({
            'success': True,
            'message': message,
            'details': error_messages if error_messages else None
        })
        
    except Exception as e:
        logger.error(f'批量导入失败: {e}\n{traceback.format_exc()}')
        return jsonify({'success': False, 'message': f'导入失败：{str(e)}'})

@app.route('/admin/links/new')
@login_required
def admin_link_new():
    """新建链接页面"""
    user_id = session.get('user_id')
    return render_template('admin_link_form.html',
                           edit_link=None,
                           defaults={'max_files': get_user_setting(user_id, 'max_files', str(DEFAULT_MAX_FILES)),
                                     'max_file_size_gb': get_user_setting(user_id, 'max_file_size_gb', str(DEFAULT_MAX_FILE_SIZE_GB)),
                                     'attachment_max_mb': get_user_setting(user_id, 'attachment_max_mb', str(DEFAULT_ATTACHMENT_MAX_MB)),
                                     'expire_days': get_user_setting(user_id, 'default_link_expire_days', '30')})

@app.route('/admin/links/<link_id>/form')
@login_required
def admin_link_form(link_id):
    """编辑链接页面"""
    if not _check_link_ownership(link_id):
        flash('无权编辑该链接')
        return redirect(url_for('admin_links'))

    conn = get_db()
    link = conn.execute("SELECT * FROM links WHERE id = ?", (link_id,)).fetchone()
    conn.close()

    if not link:
        flash('链接不存在')
        return redirect(url_for('admin_links'))

    link_dict = dict(link)
    # 编辑表单回显：收集页有效期 → 反算天数+单位
    if link_dict.get('expires_at'):
        val_str, unit_str, display_str = _format_expire_for_edit(link_dict['expires_at'])
        link_dict['_expire_days_value'] = val_str
        link_dict['_expire_unit'] = unit_str
        link_dict['_expires_display'] = display_str
    else:
        link_dict['_expire_days_value'] = ''
        link_dict['_expire_unit'] = '天'
        link_dict['_expires_display'] = '永不过期'
    # 编辑表单回显：分享页有效期 → 反算天数+单位
    if link_dict.get('share_expires_at') and link_dict['share_expires_at']:
        val_str, unit_str, display_str = _format_expire_for_edit(link_dict['share_expires_at'])
        link_dict['_share_expire_days_value'] = val_str
        link_dict['_share_expire_unit'] = unit_str
        link_dict['_share_expires_display'] = display_str
    else:
        link_dict['_share_expire_days_value'] = ''
        link_dict['_share_expire_unit'] = '天'
        link_dict['_share_expires_display'] = '永不过期'

    user_id = session.get('user_id')
    return render_template('admin_link_form.html',
                           edit_link=link_dict,
                           defaults={'max_files': get_user_setting(user_id, 'max_files', str(DEFAULT_MAX_FILES)),
                                     'max_file_size_gb': get_user_setting(user_id, 'max_file_size_gb', str(DEFAULT_MAX_FILE_SIZE_GB)),
                                     'attachment_max_mb': get_user_setting(user_id, 'attachment_max_mb', str(DEFAULT_ATTACHMENT_MAX_MB)),
                                     'expire_days': get_user_setting(user_id, 'default_link_expire_days', '30')})

@app.route('/admin/links/create', methods=['POST'])
@login_required
def create_link():
    """创建新的收集链接"""
    if not validate_csrf():
        flash('安全验证失败，请刷新页面重试')
        return redirect(url_for('admin_links'))

    title = request.form.get('title', '').strip()
    try:
        description = sanitize_html(request.form.get('description', ''))
    except Exception as e:
        logger.error(f"create_link sanitize_html 失败: {e}\n{traceback.format_exc()}")
        description = request.form.get('description', '')  # 降级：保留原始内容
    passcode = request.form.get('passcode', '').strip()
    empty_passcode = request.form.get('empty_passcode') == '1'  # 空通行证复选框
    passcode_mode = request.form.get('passcode_mode', 'set')  # set/empty/keep
    max_files = request.form.get('max_files', DEFAULT_MAX_FILES)
    max_file_size_gb = request.form.get('max_file_size_gb', str(DEFAULT_MAX_FILE_SIZE_GB))

    if not title:
        flash('标题不能为空')
        return redirect(url_for('admin_links'))

    # 验证：选择了"需要通行证验证"但未输入密码
    if passcode_mode == 'set' and not passcode:
        flash('请输入通行证密码')
        return redirect(url_for('admin_links'))

    # 如果勾选了空通行证，强制清空 passcode
    if empty_passcode:
        passcode = ''

    try:
        _mf = float(max_files)
        if _mf != int(_mf):
            raise ValueError('最大文件数量必须为整数')
        max_files = int(_mf)
        max_file_size_gb = round(float(max_file_size_gb), 6)
        if max_files < 0:
            raise ValueError('最大文件数量不能为负数')
        if max_file_size_gb < 0.01 or max_file_size_gb > 64:
            raise ValueError('单文件上限必须在 0.01-64 GB 之间')
    except ValueError as e:
        flash(str(e) if '必须' in str(e) else '数字格式错误')
        return redirect(url_for('admin_links'))

    _max_expire_days = int(get_user_setting(session.get('user_id'), 'default_link_expire_days', '30'))
    expire_days_val = request.form.get('expire_days', '').strip()
    expire_unit = request.form.get('expire_unit', '天').strip()
    expires_at, err = _parse_expire_input(expire_days_val, expire_unit, _max_expire_days)
    if err:
        flash(err)
        return redirect(url_for('admin_links'))

    link_id = generate_link_id()
    
    # 获取当前用户ID（用于文件夹命名）
    user_id = session.get('user_id')
    
    # 根据收集名称生成文件夹名
    folder_name = re.sub(r'[<>:"/\\|?*]', '_', title.strip())
    folder_name = folder_name.strip().lstrip('.')
    if not folder_name:
        folder_name = link_id
    # 检查是否与已有文件夹重名，重名则加后缀
    user_folder = get_user_folder(user_id)
    if user_folder:
        base_name = folder_name
        counter = 1
        while os.path.exists(os.path.join(user_folder, folder_name)):
            counter += 1
            folder_name = f"{base_name}_{counter}"
    
    allow_delete = 1 if request.form.get('allow_delete') == '1' else 0
    allow_preview_download = 1 if request.form.get('allow_preview_download') == '1' else 0
    if passcode:
        passcode_hash = generate_password_hash(passcode)
        passcode_plain = passcode
    else:
        # 空通行证：使用空字符串的哈希，passcode_plain 为空
        passcode_hash = generate_password_hash('')
        passcode_plain = ''

    # 分享页设置
    share_enabled = 1 if request.form.get('share_enabled') == '1' else 0
    share_passcode_raw = request.form.get('share_passcode', '').strip()
    share_passcode_empty = 1 if request.form.get('share_passcode_empty') == '1' else 0
    if share_passcode_raw and not share_passcode_empty:
        share_passcode_hash = generate_password_hash(share_passcode_raw)
        share_passcode_plain = share_passcode_raw
    else:
        share_passcode_hash = ''
        share_passcode_plain = ''

    # 分享页独立描述
    try:
        share_description = sanitize_html(request.form.get('share_description', ''))
    except Exception:
        share_description = request.form.get('share_description', '')
    # 分享页独立有效期（天数+单位）
    share_expire_days = request.form.get('share_expire_days', '').strip()
    share_expire_unit = request.form.get('share_expire_unit', '天').strip()
    share_expires_at, _share_err = _parse_expire_input(share_expire_days, share_expire_unit, _max_expire_days)
    # 收集页开关
    collect_enabled = 0 if request.form.get('collect_disabled') == '1' else 1
    require_uploader = 1 if request.form.get('require_uploader') == '1' else 0

    # 自定义链接 ID
    collect_slug = request.form.get('collect_slug', '').strip()
    share_slug = request.form.get('share_slug', '').strip()
    SLUG_PATTERN = re.compile(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$')
    SLUG_RESERVED = {'admin', 'api', 'static', 'login', 'logout', 'register', 'setup',
                     'collect', 'share', 'links', 'settings', 'upload', 'download',
                     'preview', 'chunk', 'verify', 'records', 'delete', 'toggle'}
    for field_name, slug_val in [('收集页', collect_slug), ('分享页', share_slug)]:
        if not slug_val:
            continue
        slug_lower = slug_val.lower()
        if slug_lower in SLUG_RESERVED:
            flash(f'{field_name}自定义链接"{slug_val}"为系统保留字，请更换')
            return redirect(url_for('admin_links'))
        if not SLUG_PATTERN.match(slug_val):
            flash(f'{field_name}自定义链接只能使用3-32位字母数字及-_，且必须以字母或数字开头')
            return redirect(url_for('admin_links'))
        if not _check_slug_unique(slug_val):
            flash(f'{field_name}自定义链接"{slug_val}"已被占用，请更换')
            return redirect(url_for('admin_links'))
    if not collect_slug:
        collect_slug = None
    if not share_slug:
        share_slug = None

    try:
        conn = get_db()
        conn.execute(
            """INSERT INTO links (id, user_id, title, description, passcode, passcode_plain,
               max_file_size_gb, max_files, expires_at, allow_delete, allow_preview_download, passcode_empty,
               share_enabled, share_passcode, share_passcode_plain, share_passcode_empty,
               share_description, share_expires_at, collect_enabled, require_uploader, 
               folder_name, collect_slug, share_slug)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (link_id, user_id, title, description, passcode_hash, passcode_plain, max_file_size_gb, max_files, expires_at or None, allow_delete, allow_preview_download, empty_passcode,
             share_enabled, share_passcode_hash, share_passcode_plain, share_passcode_empty,
             share_description, share_expires_at, collect_enabled, require_uploader, 
             folder_name, collect_slug, share_slug)
        )
        conn.commit()
        conn.close()

        # 创建用户文件夹（如果不存在）
        user_folder = get_user_folder(user_id)
        if user_folder and not os.path.exists(user_folder):
            os.makedirs(user_folder, exist_ok=True)

        # 创建链接专属文件夹（UPLOAD_BASE/<username>/<folder_name>/）
        try:
            create_upload_dir(link_id)
        except Exception as e:
            logger.warning(f"创建链接文件夹失败（不影响链接创建）: {e}")

        # 处理附件上传（老师上传的作业等）
        try:
            att_file = request.files.get('attachment')
            if att_file and att_file.filename and att_file.filename.strip():
                _save_link_attachment(link_id, att_file, session.get('user_id'))
        except Exception as e:
            logger.warning(f"附件保存失败（不影响链接创建）: {e}")

        # 优先显示自定义slug
        display_slug = collect_slug if collect_slug else link_id
        flash(f'收集链接已创建: /collect/{display_slug}')
    except Exception as e:
        logger.error(f"创建链接失败: {e}\n{traceback.format_exc()}")
        flash(f'创建失败：{e}')
    return redirect(url_for('admin_links'))

@app.route('/admin/links/<link_id>/edit', methods=['POST'])
@login_required
def edit_link(link_id):
    """编辑收集链接"""
    if not validate_csrf():
        flash('安全验证失败，请刷新页面重试')
        return redirect(url_for('admin_links'))

    if not _check_link_ownership(link_id):
        flash('无权编辑该链接')
        return redirect(url_for('admin_links'))

    try:
        title = request.form.get('title', '').strip()
        try:
            description = sanitize_html(request.form.get('description', ''))
        except Exception as e:
            logger.error(f"edit_link sanitize_html 失败: {e}\n{traceback.format_exc()}")
            description = request.form.get('description', '')  # 降级：保留原始内容
        passcode = request.form.get('passcode', '').strip()
        empty_passcode = request.form.get('empty_passcode') == '1'  # 空通行证复选框（选中时值为 '1'）
        passcode_mode = request.form.get('passcode_mode', 'keep')  # set/empty/keep
        max_files = request.form.get('max_files', DEFAULT_MAX_FILES)
        max_file_size_gb = request.form.get('max_file_size_gb', str(DEFAULT_MAX_FILE_SIZE_GB))

        try:
            _mf = float(max_files)
            if _mf != int(_mf):
                raise ValueError('最大文件数量必须为整数')
            max_files = int(_mf)
            max_file_size_gb = round(float(max_file_size_gb), 6)
            if max_files < 0:
                raise ValueError('最大文件数量不能为负数')
            if max_file_size_gb < 0.01 or max_file_size_gb > 64:
                raise ValueError('单文件上限必须在 0.01-64 GB 之间')
        except ValueError as e:
            flash(str(e) if '必须' in str(e) else '数字格式错误')
            return redirect(url_for('admin_links'))

        if not title:
            flash('标题不能为空')
            return redirect(url_for('admin_links'))

        # 验证：选择了"需要通行证验证"但未输入密码
        if passcode_mode == 'set' and not passcode:
            flash('请输入通行证密码')
            return redirect(url_for('admin_links'))

        _max_expire_days = int(get_user_setting(session.get('user_id'), 'default_link_expire_days', '30'))
        clear_expiry = request.form.get('clear_expiry') == '1'
        if clear_expiry:
            expires_at = None
        else:
            expire_days_val = request.form.get('expire_days', '').strip()
            expire_unit = request.form.get('expire_unit', '天').strip()
            expires_at, err = _parse_expire_input(expire_days_val, expire_unit, _max_expire_days)
            if err:
                flash(err)
                return redirect(url_for('admin_links'))
            # 分享页独立有效期（天数+单位，独立设置不受收集页影响）
            share_expire_days_edit = request.form.get('share_expire_days', '').strip()
            share_expire_unit = request.form.get('share_expire_unit', '天').strip()
            share_expires_at, share_err = _parse_expire_input(share_expire_days_edit, share_expire_unit, _max_expire_days)
            if share_err:
                flash(share_err)
                return redirect(url_for('admin_links'))

        conn = get_db()
        allow_delete = 1 if request.form.get('allow_delete') == '1' else 0
        allow_preview_download = 1 if request.form.get('allow_preview_download') == '1' else 0

        # 分享页设置
        share_enabled = 1 if request.form.get('share_enabled') == '1' else 0
        share_passcode_mode = request.form.get('share_passcode_mode', 'keep')
        share_passcode_raw = request.form.get('share_passcode', '').strip()
        share_passcode_empty = 1 if request.form.get('share_passcode_empty') == '1' else 0
        # 分享页通行证是否被用户明确改动（填了新密码 或 勾了空通行证 或 选择了复用收集页通行证）
        share_changed = bool(share_passcode_raw) or bool(share_passcode_empty) or (share_passcode_mode == 'inherit')
        if share_passcode_empty:
            share_passcode_hash = ''
            share_passcode_plain = ''
            share_passcode_empty = 1
        elif share_passcode_raw:
            share_passcode_hash = generate_password_hash(share_passcode_raw)
            share_passcode_plain = share_passcode_raw
            share_passcode_empty = 0
        elif share_passcode_mode == 'inherit':
            # 复用收集页通行证：清空独立通行证，设置 empty=0 触发后端 fallback
            share_passcode_hash = ''
            share_passcode_plain = ''
            share_passcode_empty = 0
        else:
            share_passcode_hash = ''  # 不会被使用（share_changed=False）
            share_passcode_plain = ''

        # 分享页独立描述
        try:
            share_description = sanitize_html(request.form.get('share_description', ''))
        except Exception:
            share_description = request.form.get('share_description', '')
        # 收集页开关
        collect_enabled = 0 if request.form.get('collect_disabled') == '1' else 1
        require_uploader = 1 if request.form.get('require_uploader') == '1' else 0

        # 自定义链接 ID
        collect_slug = request.form.get('collect_slug', '').strip()
        share_slug = request.form.get('share_slug', '').strip()
        SLUG_PATTERN = re.compile(r'^[a-zA-Z0-9][a-zA-Z0-9_-]{2,31}$')
        SLUG_RESERVED = {'admin', 'api', 'static', 'login', 'logout', 'register', 'setup',
                         'collect', 'share', 'links', 'settings', 'upload', 'download',
                         'preview', 'chunk', 'verify', 'records', 'delete', 'toggle'}
        for field_name, slug_val in [('收集页', collect_slug), ('分享页', share_slug)]:
            if not slug_val:
                continue
            slug_lower = slug_val.lower()
            if slug_lower in SLUG_RESERVED:
                flash(f'{field_name}自定义链接"{slug_val}"为系统保留字，请更换')
                return redirect(url_for('admin_links'))
            if not SLUG_PATTERN.match(slug_val):
                flash(f'{field_name}自定义链接只能使用3-32位字母数字及-_，且必须以字母或数字开头')
                return redirect(url_for('admin_links'))
            if not _check_slug_unique(slug_val, exclude_id=link_id):
                flash(f'{field_name}自定义链接"{slug_val}"已被占用，请更换')
                return redirect(url_for('admin_links'))
        if not collect_slug:
            collect_slug = None
        if not share_slug:
            share_slug = None

        # 处理通行证更新逻辑（新的复选框方案）
        if empty_passcode:
            # 用户勾选了空通行证：清空通行证（允许任何人访问）
            passcode_hash = generate_password_hash('')
            passcode_plain = ''
            if share_changed:
                conn.execute(
                    """UPDATE links SET title=?, description=?, passcode=?, passcode_plain=?,
                       max_file_size_gb=?, max_files=?, expires_at=?, allow_delete=?, allow_preview_download=?, passcode_empty=1,
                       share_enabled=?, share_passcode=?, share_passcode_plain=?, share_passcode_empty=?,
                       share_description=?, share_expires_at=?, collect_enabled=?, require_uploader=?, collect_slug=?, share_slug=?, updated_at=CURRENT_TIMESTAMP
                       WHERE id=?""",
                    (title, description, passcode_hash, passcode_plain, max_file_size_gb, max_files, expires_at or None, allow_delete, allow_preview_download,
                     share_enabled, share_passcode_hash, share_passcode_plain, share_passcode_empty,
                     share_description, share_expires_at, collect_enabled, require_uploader, collect_slug, share_slug, link_id)
                )
            else:
                conn.execute(
                    """UPDATE links SET title=?, description=?, passcode=?, passcode_plain=?,
                       max_file_size_gb=?, max_files=?, expires_at=?, allow_delete=?, allow_preview_download=?, passcode_empty=1,
                       share_enabled=?, share_description=?, share_expires_at=?, collect_enabled=?, require_uploader=?, collect_slug=?, share_slug=?, updated_at=CURRENT_TIMESTAMP
                       WHERE id=?""",
                    (title, description, passcode_hash, passcode_plain, max_file_size_gb, max_files, expires_at or None, allow_delete, allow_preview_download,
                     share_enabled, share_description, share_expires_at, collect_enabled, require_uploader, collect_slug, share_slug, link_id)
                )
        elif passcode:
            # 用户输入了新通行证：设置新通行证
            passcode_hash = generate_password_hash(passcode)
            passcode_plain = passcode
            if share_changed:
                conn.execute(
                    """UPDATE links SET title=?, description=?, passcode=?, passcode_plain=?,
                       max_file_size_gb=?, max_files=?, expires_at=?, allow_delete=?, allow_preview_download=?, passcode_empty=0,
                       share_enabled=?, share_passcode=?, share_passcode_plain=?, share_passcode_empty=?,
                       share_description=?, share_expires_at=?, collect_enabled=?, require_uploader=?, collect_slug=?, share_slug=?, updated_at=CURRENT_TIMESTAMP
                       WHERE id=?""",
                    (title, description, passcode_hash, passcode_plain, max_file_size_gb, max_files, expires_at or None, allow_delete, allow_preview_download,
                     share_enabled, share_passcode_hash, share_passcode_plain, share_passcode_empty,
                     share_description, share_expires_at, collect_enabled, require_uploader, collect_slug, share_slug, link_id)
                )
            else:
                conn.execute(
                    """UPDATE links SET title=?, description=?, passcode=?, passcode_plain=?,
                       max_file_size_gb=?, max_files=?, expires_at=?, allow_delete=?, allow_preview_download=?, passcode_empty=0,
                       share_enabled=?, share_description=?, share_expires_at=?, collect_enabled=?, require_uploader=?, collect_slug=?, share_slug=?, updated_at=CURRENT_TIMESTAMP
                       WHERE id=?""",
                    (title, description, passcode_hash, passcode_plain, max_file_size_gb, max_files, expires_at or None, allow_delete, allow_preview_download,
                     share_enabled, share_description, share_expires_at, collect_enabled, require_uploader, collect_slug, share_slug, link_id)
                )
        else:
            # 保持原有通行证不变 - 只更新其他字段
            if share_changed:
                conn.execute(
                    """UPDATE links SET title=?, description=?,
                       max_file_size_gb=?, max_files=?, expires_at=?, allow_delete=?, allow_preview_download=?,
                       share_enabled=?, share_passcode=?, share_passcode_plain=?, share_passcode_empty=?,
                       share_description=?, share_expires_at=?, collect_enabled=?, require_uploader=?, collect_slug=?, share_slug=?, updated_at=CURRENT_TIMESTAMP
                       WHERE id=?""",
                    (title, description, max_file_size_gb, max_files, expires_at or None, allow_delete, allow_preview_download,
                     share_enabled, share_passcode_hash, share_passcode_plain, share_passcode_empty,
                     share_description, share_expires_at, collect_enabled, require_uploader, collect_slug, share_slug, link_id)
                )
            else:
                conn.execute(
                    """UPDATE links SET title=?, description=?,
                       max_file_size_gb=?, max_files=?, expires_at=?, allow_delete=?, allow_preview_download=?,
                       share_enabled=?, share_description=?, share_expires_at=?, collect_enabled=?, require_uploader=?, collect_slug=?, share_slug=?, updated_at=CURRENT_TIMESTAMP
                       WHERE id=?""",
                    (title, description, max_file_size_gb, max_files, expires_at or None, allow_delete, allow_preview_download,
                     share_enabled, share_description, share_expires_at, collect_enabled, require_uploader, collect_slug, share_slug, link_id)
                )

        # 处理附件（上传新附件 / 删除旧附件 / 保持不变）
        remove_attachment = request.form.get('remove_attachment') == '1'
        att_file = request.files.get('attachment')
        if remove_attachment:
            old = conn.execute("SELECT attachment_path FROM links WHERE id=?", (link_id,)).fetchone()
            if old and old['attachment_path']:
                _safe_delete(old['attachment_path'])
            conn.execute("UPDATE links SET attachment_name='', attachment_path='', attachment_size=0 WHERE id=?", (link_id,))
        elif att_file and att_file.filename and att_file.filename.strip():
            # 验证大小
            att_file.seek(0, 2)
            att_size = att_file.tell()
            att_file.seek(0)
            max_attach = get_attachment_max_size(session.get('user_id'))
            if att_size > max_attach:
                conn.close()
                flash(f'附件大小不能超过 {format_file_size(max_attach)}')
                return redirect(url_for('admin_links'))
            # 删除旧附件
            old = conn.execute("SELECT attachment_path FROM links WHERE id=?", (link_id,)).fetchone()
            if old and old['attachment_path']:
                _safe_delete(old['attachment_path'])
            try:
                upload_dir = create_upload_dir(link_id, '')
                attach_dir = os.path.join(upload_dir, '_attachment')
                os.makedirs(attach_dir, mode=0o755, exist_ok=True)
                ext = os.path.splitext(att_file.filename)[1]
                safe_name = re.sub(r'[^\w.\-]', '_', att_file.filename)
                if not safe_name:
                    safe_name = 'attachment' + ext
                stored_path = os.path.join(attach_dir, safe_name)
                att_file.save(stored_path)
                conn.execute(
                    "UPDATE links SET attachment_name=?, attachment_path=?, attachment_size=? WHERE id=?",
                    (att_file.filename, stored_path, att_size, link_id)
                )
            except Exception as e:
                logger.warning(f"编辑时附件保存失败: {e}")

        conn.commit()
        conn.close()

        flash('收集已更新')
        return redirect(url_for('admin_links'))

    except Exception as e:
        logger.error(f"编辑链接 {link_id} 失败: {e}\n{traceback.format_exc()}")
        flash(f'编辑失败：{e}')
        return redirect(url_for('admin_links'))

# ============================================================
# 附件下载路由（老师上传的作业等）
# ============================================================

@app.route('/collect/<link_id>/attachment')
def collect_attachment(link_id):
    """下载收集页附件"""
    return _serve_link_attachment(link_id)





def _serve_link_attachment(link_id, as_attachment=True):
    """安全下载/内联链接附件，使用路径遍历防护"""
    conn = get_db()
    link = conn.execute(
        "SELECT * FROM links WHERE (collect_slug = ? OR id = ?) AND status = 'active'",
        (link_id, link_id)
    ).fetchone()
    if not link:
        conn.close()
        abort(404)

    # 统一使用数据库规范 ID
    link_id = link['id']

    if not is_verified(link_id, link):
        conn.close()
        return '请先验证通行证', 403

    ok, err = _check_public_link_token(link_id, link, for_share=False)
    if not ok:
        conn.close()
        return err, 401

    if not link['attachment_path']:
        conn.close()
        abort(404)
    conn.close()

    stored_path = link['attachment_path']
    upload_base = get_upload_base()
    real_path = os.path.realpath(stored_path)
    real_base = os.path.realpath(upload_base)
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        logger.warning(f"附件路径遍历攻击拦截: stored_path={stored_path}")
        abort(403)

    if not os.path.exists(real_path) or not os.path.isfile(real_path):
        abort(404)

    original_name = link['attachment_name'] or 'attachment'
    if as_attachment:
        mimetype = 'application/octet-stream'
    else:
        ext = os.path.splitext(original_name)[1].lower()
        # HEIC/HEIF 浏览器不原生支持，服务端转为 JPEG
        if ext in ('.heic', '.heif'):
            if HEIC_SUPPORT:
                buf = _serve_heic_as_jpeg(real_path)
                if buf:
                    resp = make_response(buf.read())
                    resp.headers['Content-Type'] = 'image/jpeg'
                    resp.headers['Cache-Control'] = 'public, max-age=3600'
                    return resp
            # HEIC 不支持时，返回提示而非原始文件
            resp = make_response('HEIC 图片暂不支持预览，请点击下载', 415)
            resp.headers['Content-Type'] = 'text/plain; charset=utf-8'
            return resp
        mimetypes_map = {
            '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.png': 'image/png',
            '.gif': 'image/gif', '.webp': 'image/webp', '.bmp': 'image/bmp',
            '.svg': 'image/svg+xml',
            '.mp4': 'video/mp4', '.webm': 'video/webm', '.ogg': 'video/ogg',
            '.mov': 'video/quicktime', '.avi': 'video/x-msvideo',
            '.mp3': 'audio/mpeg', '.wav': 'audio/wav', '.flac': 'audio/flac',
            '.aac': 'audio/aac', '.m4a': 'audio/mp4',
            # Office / PDF 格式（flyfish file-viewer 需要正确 Content-Type）
            '.pdf': 'application/pdf',
            '.doc': 'application/msword',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.xls': 'application/vnd.ms-excel',
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.ppt': 'application/vnd.ms-powerpoint',
            '.pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation',
            '.ofd': 'application/octet-stream',
            '.txt': 'text/plain; charset=utf-8',
            '.md': 'text/markdown; charset=utf-8',
            '.csv': 'text/csv; charset=utf-8',
            '.dxf': 'application/dxf',
            '.dwg': 'application/acad',
        }
        mimetype = mimetypes_map.get(ext, 'application/octet-stream')
    return send_file(
        real_path,
        as_attachment=as_attachment,
        download_name=original_name,
        mimetype=mimetype,
        conditional=True
    )




@app.route('/collect/<link_id>/attachment/preview')
def collect_attachment_preview(link_id):
    """附件内联预览（图片/视频/音频直接展示）"""
    return _serve_link_attachment(link_id, as_attachment=False)


@app.route('/admin/links/<link_id>/toggle', methods=['POST'])
@login_required
def toggle_link(link_id):
    """启用/禁用链接"""
    if not _check_link_ownership(link_id):
        flash('无权操作该链接')
        return redirect(url_for('admin_links'))
    conn = get_db()
    link = conn.execute("SELECT status FROM links WHERE id = ?", (link_id,)).fetchone()
    if link:
        new_status = 'inactive' if link['status'] == 'active' else 'active'
        conn.execute("UPDATE links SET status = ? WHERE id = ?", (new_status, link_id))
        conn.commit()
    conn.close()
    flash('状态已更新')
    return redirect(url_for('admin_links'))

@app.route('/admin/links/<link_id>/delete', methods=['POST'])
@login_required
def delete_link(link_id):
    """删除链接及关联的所有上传文件和记录"""
    if not _check_link_ownership(link_id):
        flash('无权删除该链接')
        return redirect(url_for('admin_links'))
    conn = get_db()

    # 1. 查询所有关联的上传记录
    records = conn.execute(
        "SELECT id, stored_path FROM upload_records WHERE link_id = ?", (link_id,)
    ).fetchall()

    # 2. 删除磁盘上的文件（使用安全删除函数）
    deleted_count = 0
    for r in records:
        stored_path = r['stored_path']
        if not stored_path:
            continue
        try:
            if _safe_delete(stored_path):
                deleted_count += 1
        except Exception:
            pass

    # 3. 获取链接目录路径（在删除 DB 记录前获取）
    try:
        base_dir = create_upload_dir(link_id, '')
    except Exception:
        base_dir = None

    # 4. 删除数据库记录（download_logs 通过 record_id 关联，需先于 upload_records 删除）
    record_ids = [r['id'] for r in records]
    conn.execute("PRAGMA foreign_keys = OFF")
    if record_ids:
        placeholders = ','.join(['?' for _ in record_ids])
        conn.execute(f"DELETE FROM download_logs WHERE record_id IN ({placeholders})", record_ids)
    conn.execute("DELETE FROM upload_records WHERE link_id = ?", (link_id,))
    conn.execute("DELETE FROM upload_logs WHERE link_id = ?", (link_id,))
    conn.execute("DELETE FROM chunk_uploads WHERE link_id = ?", (link_id,))
    conn.execute("DELETE FROM links WHERE id = ?", (link_id,))
    conn.execute("PRAGMA foreign_keys = ON")
    conn.commit()
    conn.close()

    # 5. 删除整个上传目录（包含所有子文件夹和文件）
    folder_deleted = False
    if base_dir and base_dir != UPLOAD_BASE:
        try:
            folder_deleted = _safe_delete_dir(base_dir)
        except Exception:
            pass

    flash(f'链接已删除，清理了 {deleted_count} 个文件及 {len(records)} 条上传记录' + ('，同时删除了整个文件夹' if folder_deleted else ''))
    return redirect(url_for('admin_links'))

@app.route('/admin/links/batch_delete', methods=['POST'])
def batch_delete_links():
    """批量删除链接（AJAX）—— 自行处理登录/CSRF，返回JSON"""
    # 登录检查
    if not session.get('user_id'):
        return jsonify({'success': False, 'message': '未登录，请刷新页面后重试'}), 401
    # CSRF 检查
    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败，请刷新页面后重试'}), 403
    
    user_id = session.get('user_id')
    is_admin = session.get('is_admin', False)
    
    link_ids = request.form.get('link_ids', '')
    if not link_ids:
        return jsonify({'success': False, 'message': '未选择任何链接'})
    
    # 链接ID为8位hex字符串，过滤无效值
    ids = []
    for x in link_ids.split(','):
        x = x.strip()
        if x and re.match(r'^[a-f0-9]{8}$', x, re.I):
            ids.append(x)
    
    if not ids:
        return jsonify({'success': False, 'message': '未选择任何链接'})
    
    conn = get_db()
    deleted_count = 0
    file_count = 0
    record_count = 0
    
    for link_id in ids:
        # 权限检查
        if not is_admin:
            link = conn.execute("SELECT user_id FROM links WHERE id = ?", (link_id,)).fetchone()
            if not link or link['user_id'] != user_id:
                continue
        
        # 查询所有关联的上传记录
        records = conn.execute(
            "SELECT id, stored_path FROM upload_records WHERE link_id = ?", (link_id,)
        ).fetchall()
        
        # 删除磁盘上的文件
        for r in records:
            stored_path = r['stored_path']
            if not stored_path:
                continue
            try:
                if _safe_delete(stored_path):
                    file_count += 1
            except Exception:
                pass
        
        # 获取链接目录路径
        try:
            base_dir = create_upload_dir(link_id, '')
        except Exception:
            base_dir = None
        
        # 删除数据库记录
        record_ids = [r['id'] for r in records]
        conn.execute("PRAGMA foreign_keys = OFF")
        if record_ids:
            placeholders = ','.join(['?' for _ in record_ids])
            conn.execute(f"DELETE FROM download_logs WHERE record_id IN ({placeholders})", record_ids)
        conn.execute("DELETE FROM upload_records WHERE link_id = ?", (link_id,))
        conn.execute("DELETE FROM upload_logs WHERE link_id = ?", (link_id,))
        conn.execute("DELETE FROM chunk_uploads WHERE link_id = ?", (link_id,))
        conn.execute("DELETE FROM links WHERE id = ?", (link_id,))
        conn.execute("PRAGMA foreign_keys = ON")
        
        record_count += len(records)
        deleted_count += 1
        
        # 删除上传目录
        if base_dir and base_dir != UPLOAD_BASE:
            try:
                _safe_delete_dir(base_dir)
            except Exception:
                pass
    
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'message': f'已删除 {deleted_count} 个链接，清理了 {file_count} 个文件及 {record_count} 条上传记录'
    })

@app.route('/admin/records')
@login_required
def admin_records():
    """上传记录管理（高性能版：后台线程负责扫描和清理，此处仅查询）"""
    page = request.args.get('page', 1, type=int)
    if page < 1:
        page = 1
    user_id = session.get('user_id')
    is_admin = session.get('is_admin', False)
    per_page = int(get_user_setting(user_id, 'records_per_page', '50'))
    link_filter = request.args.get('link_id', '').strip()

    # 验证权限（非管理员只能看自己的链接）
    if link_filter and not is_admin:
        cv = get_db()
        owner = cv.execute("SELECT user_id FROM links WHERE id = ?", (link_filter,)).fetchone()
        cv.close()
        if not owner or owner['user_id'] != user_id:
            flash('无权访问该链接')
            return redirect(url_for('admin_records'))

    conn = get_db()

    # 构建高效的 COUNT 和数据查询
    # 管理员无过滤时无需 JOIN（upload_records 已有 link_id）
    # 使用特定字段而非 SELECT r.* 减少数据传输
    _rec_cols = ("r.id, r.original_name, r.file_size, r.file_size_display, r.download_count, "
                 "r.uploader_ip, r.uploaded_at, r.link_id, r.stored_path, r.stored_name, r.uploader_name")

    if is_admin:
        if link_filter:
            where = "WHERE r.link_id = ?"
            count_params = (link_filter,)
            data_params = (link_filter, per_page, (page - 1) * per_page)
        else:
            where = ""
            count_params = ()
            data_params = (per_page, (page - 1) * per_page)
    else:
        if link_filter:
            where = ("WHERE r.link_id IN (SELECT id FROM links l2 WHERE l2.id = ? AND l2.user_id = ?)")
            count_params = (link_filter, user_id)
            data_params = (link_filter, user_id, per_page, (page - 1) * per_page)
        else:
            where = ("WHERE r.link_id IN (SELECT id FROM links l2 WHERE l2.user_id = ?)")
            count_params = (user_id,)
            data_params = (user_id, per_page, (page - 1) * per_page)

    # COUNT 查询（轻量，不用 JOIN）
    count_sql = f"SELECT COUNT(*) FROM upload_records r {where}"
    count = conn.execute(count_sql, count_params).fetchone()[0]

    total_pages = max(1, (count + per_page - 1) // per_page)
    page = min(page, total_pages)

    # 数据查询：用子查询取 link_title、require_uploader 和 nickname（优先昵称）
    data_sql = f"""SELECT {_rec_cols},
                   (SELECT l.title FROM links l WHERE l.id = r.link_id) as link_title,
                   (SELECT l.require_uploader FROM links l WHERE l.id = r.link_id) as require_uploader,
                   COALESCE((SELECT u.nickname FROM links l JOIN users u ON l.user_id = u.id WHERE l.id = r.link_id),
                            (SELECT u.username FROM links l JOIN users u ON l.user_id = u.id WHERE l.id = r.link_id)) as creator_name
                   FROM upload_records r {where}
                   ORDER BY r.uploaded_at DESC
                   LIMIT ? OFFSET ?"""
    records = conn.execute(data_sql, data_params).fetchall()

    # 链接列表（用于筛选下拉）
    if is_admin:
        links = conn.execute("SELECT id, title FROM links ORDER BY created_at DESC").fetchall()
    else:
        links = conn.execute(
            "SELECT id, title FROM links WHERE user_id = ? ORDER BY created_at DESC",
            (user_id,)
        ).fetchall()
    conn.close()

    return render_template('admin_records.html',
        records=records,
        links=links,
        page=page,
        total_pages=total_pages,
        total_count=count,
        link_filter=link_filter,
        per_page=per_page)

@app.route('/admin/records/<int:record_id>/download')
@login_required
def admin_download_record(record_id):
    """下载上传记录中的文件"""
    if not _check_record_ownership(record_id):
        return '无权访问', 403
    
    conn = get_db()
    record = conn.execute(
        "SELECT stored_path, original_name FROM upload_records WHERE id = ?",
        (record_id,)
    ).fetchone()

    if not record:
        conn.close()
        return '文件不存在', 404

    conn.execute("UPDATE upload_records SET download_count = download_count + 1 WHERE id = ?", (record_id,))
    conn.commit()
    conn.close()

    _log_download(record_id, 'admin')
    return _safe_download(record['stored_path'], record['original_name'])

@app.route('/admin/records/<int:record_id>/preview-img')
@login_required
def admin_preview_image(record_id):
    """预览图片（HEIC 自动转为 JPEG 显示）"""
    if not _check_record_ownership(record_id):
        return '无权访问', 403
    conn = get_db()
    record = conn.execute(
        "SELECT stored_path, original_name FROM upload_records WHERE id = ?",
        (record_id,)
    ).fetchone()
    conn.close()
    if not record:
        return '文件不存在', 404
    
    ext = os.path.splitext(record['original_name'])[1].lower()
    
    if ext in ('.heic', '.heif'):
        if not HEIC_SUPPORT:
            return '服务端不支持 HEIC 预览，请安装 pillow-heif', 501
        buf = _serve_heic_as_jpeg(record['stored_path'])
        if not buf:
            return 'HEIC 转换失败', 500
        resp = make_response(buf.read())
        resp.headers['Content-Type'] = 'image/jpeg'
        resp.headers['Cache-Control'] = 'public, max-age=3600'
        return resp
    
    # 普通图片直接返回
    return _safe_download(record['stored_path'], record['original_name'])

# ============================================================

@app.route('/admin/records/<int:record_id>/download-logs')
@login_required
def admin_download_logs(record_id):
    """查看文件下载日志"""
    if not _check_record_ownership(record_id):
        return '无权访问', 403

    page = request.args.get('page', 1, type=int)
    if page < 1:
        page = 1
    per_page = 20

    conn = get_db()
    record = conn.execute(
        "SELECT original_name FROM upload_records WHERE id = ?", (record_id,)
    ).fetchone()
    if not record:
        conn.close()
        return '文件不存在', 404

    total = conn.execute(
        "SELECT COUNT(*) FROM download_logs WHERE record_id = ?", (record_id,)
    ).fetchone()[0]
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)
    offset = (page - 1) * per_page

    logs = conn.execute(
        """SELECT * FROM download_logs
           WHERE record_id = ?
           ORDER BY downloaded_at DESC
           LIMIT ? OFFSET ?""",
        (record_id, per_page, offset)
    ).fetchall()
    conn.close()

    return render_template('admin_download_logs.html',
                           record=record,
                           record_id=record_id,
                           logs=logs,
                           page=page,
                           total_pages=total_pages,
                           total=total,
                           prev_url=url_for('admin_download_logs', record_id=record_id, page=page-1) if page > 1 else None,
                           next_url=url_for('admin_download_logs', record_id=record_id, page=page+1) if page < total_pages else None,
                           now=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

@app.route('/admin/records/<int:record_id>/upload-logs')
@login_required
def admin_upload_logs(record_id):
    """查看文件上传日志"""
    if not _check_record_ownership(record_id):
        return '无权访问', 403

    page = request.args.get('page', 1, type=int)
    if page < 1:
        page = 1
    per_page = 20

    conn = get_db()
    record = conn.execute(
        """SELECT r.original_name, l.require_uploader
           FROM upload_records r
           LEFT JOIN links l ON r.link_id = l.id
           WHERE r.id = ?""", (record_id,)
    ).fetchone()
    if not record:
        conn.close()
        return '文件不存在', 404

    total = conn.execute(
        "SELECT COUNT(*) FROM upload_logs WHERE record_id = ?", (record_id,)
    ).fetchone()[0]
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)
    offset = (page - 1) * per_page

    logs = conn.execute(
        """SELECT * FROM upload_logs
           WHERE record_id = ?
           ORDER BY event_time DESC
           LIMIT ? OFFSET ?""",
        (record_id, per_page, offset)
    ).fetchall()
    conn.close()

    return render_template('admin_upload_logs.html',
                           record=record,
                           record_id=record_id,
                           logs=logs,
                           page=page,
                           total_pages=total_pages,
                           total=total,
                           prev_url=url_for('admin_upload_logs', record_id=record_id, page=page-1) if page > 1 else None,
                           next_url=url_for('admin_upload_logs', record_id=record_id, page=page+1) if page < total_pages else None,
                           require_uploader=bool(record['require_uploader']) if record['require_uploader'] else False,
                           now=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

@app.route('/admin/links/<link_id>/upload-logs')
@login_required
def admin_link_upload_logs(link_id):
    """查看收集链接的上传日志"""
    if not _check_link_ownership(link_id):
        return '无权访问', 403

    page = request.args.get('page', 1, type=int)
    if page < 1:
        page = 1
    per_page = 20

    conn = get_db()
    link = conn.execute("SELECT title, require_uploader FROM links WHERE id = ?", (link_id,)).fetchone()
    if not link:
        conn.close()
        return '链接不存在', 404

    total = conn.execute(
        "SELECT COUNT(*) FROM upload_logs WHERE link_id = ?", (link_id,)
    ).fetchone()[0]
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)
    offset = (page - 1) * per_page

    logs = conn.execute(
        """SELECT * FROM upload_logs
           WHERE link_id = ?
           ORDER BY event_time DESC
           LIMIT ? OFFSET ?""",
        (link_id, per_page, offset)
    ).fetchall()
    conn.close()

    return render_template('admin_upload_logs.html',
                           record={'original_name': link['title']},
                           link_title=link['title'],
                           require_uploader=bool(link['require_uploader']) if link['require_uploader'] else False,
                           back_url=url_for('admin_links'),
                           logs=logs,
                           page=page,
                           total_pages=total_pages,
                           total=total,
                           prev_url=url_for('admin_link_upload_logs', link_id=link_id, page=page-1) if page > 1 else None,
                           next_url=url_for('admin_link_upload_logs', link_id=link_id, page=page+1) if page < total_pages else None,
                           now=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

@app.route('/admin/records/<int:record_id>/preview_file')
@login_required
def admin_preview_file(record_id):
    """预览文件（用于 flyfish file-viewer 获取文件内容）"""
    if not _check_record_ownership(record_id):
        abort(403)
    
    conn = get_db()
    record = conn.execute(
        "SELECT stored_path, original_name FROM upload_records WHERE id = ?",
        (record_id,)
    ).fetchone()

    if not record:
        conn.close()
        return '文件不存在', 404

    conn.close()

    upload_base = get_upload_base()
    real_path = os.path.realpath(record['stored_path'])
    real_base = os.path.realpath(upload_base)
    
    # 安全检查
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        logger.warning(f"路径遍历攻击拦截: stored_path={record['stored_path']}, upload_base={real_base}")
        abort(403)
    
    if not os.path.isfile(real_path):
        abort(404)

    directory = os.path.dirname(real_path)
    filename = os.path.basename(real_path)
    
    # 不使用 as_attachment，允许SDK读取文件内容
    response = send_from_directory(
        directory, filename,
        download_name=record['original_name'],
        as_attachment=False
    )
    
    # 添加CORS响应头，允许JIT Viewer SDK跨域访问
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    
    return response

@app.route('/admin/records/<int:record_id>/txt_info')
@login_required
def admin_txt_info(record_id):
    """返回 TXT 文件章节目录和编码信息（管理后台）"""
    if not _check_record_ownership(record_id):
        abort(403)
    conn = get_db()
    record = conn.execute(
        "SELECT stored_path, original_name FROM upload_records WHERE id = ?",
        (record_id,)
    ).fetchone()
    if not record:
        conn.close()
        return jsonify({'error': '文件不存在'}), 404
    conn.close()
    upload_base = get_upload_base()
    real_path = os.path.realpath(record['stored_path'])
    real_base = os.path.realpath(upload_base)
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        abort(403)
    if not os.path.isfile(real_path):
        abort(404)
    chapters, encoding = _scan_txt_chapters(real_path)
    total_size = os.path.getsize(real_path)
    return jsonify({
        'chapters': chapters,
        'encoding': encoding,
        'total_size': total_size
    })


@app.route('/admin/records/<int:record_id>/txt_chunk')
@login_required
def admin_txt_chunk(record_id):
    """返回 TXT 文件的文本块（分页用，管理后台）"""
    if not _check_record_ownership(record_id):
        abort(403)
    conn = get_db()
    record = conn.execute(
        "SELECT stored_path, original_name FROM upload_records WHERE id = ?",
        (record_id,)
    ).fetchone()
    if not record:
        conn.close()
        return jsonify({'error': '文件不存在'}), 404
    conn.close()
    upload_base = get_upload_base()
    real_path = os.path.realpath(record['stored_path'])
    real_base = os.path.realpath(upload_base)
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        abort(403)
    if not os.path.isfile(real_path):
        abort(404)
    try:
        offset = int(request.args.get('offset', 0))
        size = min(int(request.args.get('size', 65536)), 524288)
    except (ValueError, TypeError):
        return jsonify({'error': '参数无效'}), 400
    text, file_size = _read_txt_chunk(real_path, offset, size)
    return jsonify({
        'text': text,
        'offset': offset,
        'size': len(text.encode('utf-8')),
        'total_size': file_size,
        'has_more': (offset + size) < file_size
    })


@app.route('/admin/records/<int:record_id>/delete', methods=['POST'])
@login_required
def delete_record(record_id):
    """删除单条上传记录（同时删除文件）"""
    if not validate_csrf():
        flash('安全验证失败，请刷新页面重试')
        return redirect(url_for('admin_records'))
    if not _check_record_ownership(record_id):
        flash('无权删除该记录')
        return redirect(url_for('admin_records'))
    
    conn = get_db()
    record = conn.execute(
        "SELECT * FROM upload_records WHERE id = ?", (record_id,)
    ).fetchone()

    if record:
        try:
            _safe_delete(record['stored_path'])
        except OSError:
            pass
        conn.execute("DELETE FROM upload_records WHERE id = ?", (record_id,))
        conn.commit()

    conn.close()
    flash('记录已删除')
    return redirect(url_for('admin_records'))

@app.route('/admin/preview_record/<int:record_id>')
@login_required
def admin_preview_record(record_id):
    """预览上传的文件"""
    if not _check_record_ownership(record_id):
        return '<p class="error">无权访问该文件</p>'
    
    conn = get_db()
    record = conn.execute(
        "SELECT * FROM upload_records WHERE id = ?", (record_id,)
    ).fetchone()
    conn.close()

    if not record:
        return '<p class="error">文件不存在</p>'

    stored_path = record['stored_path']
    original_name = record['original_name']
    
    # 路径遍历安全校验
    upload_base = get_upload_base()
    real_path = os.path.realpath(stored_path)
    real_base = os.path.realpath(upload_base)
    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        logger.warning(f"路径遍历攻击拦截: stored_path={stored_path}, upload_base={real_base}")
        return '<p class="error">无权访问该文件</p>'
    
    # 获取文件扩展名
    _, ext = os.path.splitext(original_name.lower())
    
    # 图片类型
    image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.svg']
    # 视频类型
    video_extensions = ['.mp4', '.webm', '.ogg', '.mov']
    # 文本类型
    text_extensions = ['.txt', '.md', '.log', '.json', '.xml', '.csv']
    
    import html as _html
    safe_original_name = _html.escape(original_name, quote=True)
    
    if ext in image_extensions:
        # 图片预览
        return f'<img src="/admin/records/{record_id}/download" alt="{safe_original_name}" style="max-width:100%;max-height:70vh;">'
    elif ext in video_extensions:
        # 视频预览
        return f'<video controls style="max-width:100%;max-height:70vh;"><source src="/admin/records/{record_id}/download" type="video/mp4">您的浏览器不支持视频播放</video>'
    elif ext in text_extensions:
        # 文本文件预览（使用已校验的 real_path，避免 TOCTOU）
        try:
            with open(real_path, 'r', encoding='utf-8') as f:
                content = f.read()
            if ext == '.md':
                # Markdown文件返回内容，前端用marked.js渲染
                return content
            else:
                import html as _html2
                safe_content = _html2.escape(content, quote=False)
                return f'<pre style="white-space:pre-wrap;word-wrap:break-word;">{safe_content}</pre>'
        except Exception as e:
            return f'<p class="error">无法读取文件内容: {_html.escape(str(e))}</p>'
    else:
        # 不支持的类型（Office文件在前端用JavaScript库处理）
        return f'<p class="info">此文件类型需要前端JavaScript库预览<br>文件名: {safe_original_name}</p>'

@app.route('/admin/records/batch-delete', methods=['POST'])
@login_required
def admin_batch_delete_records():
    """批量删除记录"""
    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败，请刷新页面重试'})
    ids = request.form.getlist('ids[]')
    if not ids:
        return jsonify({'success': False, 'message': '未选择记录'})

    is_admin = session.get('is_admin', False)
    user_id = session.get('user_id')
    conn = get_db()
    deleted = 0
    for rid in ids:
        # 非管理员校验所有权
        if not is_admin:
            row = conn.execute(
                """SELECT r.stored_path FROM upload_records r
                   INNER JOIN links l ON r.link_id = l.id
                   WHERE r.id = ? AND l.user_id = ?""",
                (rid, user_id)
            ).fetchone()
        else:
            row = conn.execute(
                "SELECT stored_path FROM upload_records WHERE id = ?", (rid,)
            ).fetchone()
        if row:
            try:
                _safe_delete(row['stored_path'])
            except OSError:
                pass
            conn.execute("DELETE FROM upload_records WHERE id = ?", (rid,))
            deleted += 1
    conn.commit()
    conn.close()

    flash(f'已删除 {deleted} 条记录')
    return redirect(url_for('admin_records'))

@app.route('/admin/records/batch-download', methods=['POST'])
@login_required
def batch_download_records():
    """批量打包下载选中记录"""
    if not validate_csrf():
        return '安全验证失败，请刷新页面重试', 403
    ids = request.form.getlist('ids[]')
    if not ids:
        return '未选择任何记录', 400

    is_admin = session.get('is_admin', False)
    user_id = session.get('user_id')
    conn = get_db()

    records = []
    for rid in ids:
        try:
            if not is_admin:
                row = conn.execute(
                    """SELECT r.stored_path, r.original_name, r.link_id, r.uploader_name, r.file_size
                       FROM upload_records r
                       INNER JOIN links l ON r.link_id = l.id
                       WHERE r.id = ? AND l.user_id = ?""",
                    (rid, user_id)
                ).fetchone()
            else:
                row = conn.execute(
                    "SELECT stored_path, original_name, link_id, uploader_name, file_size FROM upload_records WHERE id = ?",
                    (rid,)
                ).fetchone()
            if row:
                records.append({
                    'stored_path': row['stored_path'],
                    'original_name': row['original_name'],
                    'link_id': row['link_id'],
                    'uploader_name': row['uploader_name'],
                    'file_size': row['file_size'],
                })
        except Exception:
            pass
    conn.close()

    if not records:
        return '没有可下载的记录', 400

    upload_base = os.path.realpath(get_upload_base())

    # 创建临时 ZIP 文件
    tmp = tempfile.NamedTemporaryFile(suffix='.zip', delete=False)
    tmp_path = tmp.name

    try:
        # 跟踪 ZIP 内已用路径，防止同名覆盖
        used_names = {}
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zf:
            for r in records:
                stored_path = r['stored_path']
                original_name = r['original_name']
                uploader_name = r['uploader_name'] or ''
                link_id = r['link_id']

                # 路径遍历安全检查
                real_path = os.path.realpath(stored_path)
                if not (real_path.startswith(upload_base + os.sep) or real_path == upload_base):
                    logger.warning(f"批量下载路径遍历拦截: {stored_path}")
                    continue
                if not os.path.isfile(real_path):
                    continue

                # 获取链接标题作为顶层文件夹
                link = get_link_by_id(link_id)
                link_title = link.get('title', '') or link_id if link else link_id
                safe_title = re.sub(r'[<>:"/\\|?*]', '_', link_title.strip())

                # 构建 ZIP 内路径：链接标题 / [上传者 /] 文件名
                if uploader_name:
                    safe_uploader = re.sub(r'[<>:"/\\|?*]', '_', uploader_name.strip())
                    arcname = f"{safe_title}/{safe_uploader}/{original_name}"
                else:
                    arcname = f"{safe_title}/{original_name}"

                # 处理重名
                if arcname in used_names:
                    used_names[arcname] += 1
                    base, ext = os.path.splitext(arcname)
                    arcname = f"{base}({used_names[arcname]}){ext}"
                else:
                    used_names[arcname] = 0

                # 智能压缩：已压缩格式只存储不重新压缩
                compress_type = _get_zip_compress_type(original_name)
                zf.write(real_path, arcname=arcname, compress_type=compress_type)

        tmp.close()

        @after_this_request
        def cleanup_zip(response):
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
            return response

        return send_file(
            tmp_path,
            download_name='文件收集_批量下载.zip',
            as_attachment=True,
            mimetype='application/zip'
        )

    except Exception as e:
        logger.error(f"批量下载失败: {e}\n{traceback.format_exc()}")
        try:
            tmp.close()
            os.unlink(tmp_path)
        except OSError:
            pass
        return '打包下载失败，请重试', 500

@app.route('/admin/ajax/set-per-page', methods=['POST'])
@login_required
def ajax_set_per_page():
    """AJAX 保存分页设置"""
    if not validate_csrf():
        return jsonify({'ok': False, 'error': '安全验证失败'}), 403
    key = request.form.get('key', '').strip()
    raw = request.form.get('value', '').strip()
    user_id = session.get('user_id')

    if key not in ('links_per_page', 'records_per_page'):
        return jsonify({'ok': False, 'error': '无效的设置项'}), 400

    try:
        _v = float(raw)
        if _v != int(_v):
            raise ValueError
        val = int(_v)
        max_val = 100 if key == 'links_per_page' else 200
        if val < 5 or val > max_val:
            raise ValueError
    except ValueError:
        return jsonify({'ok': False, 'error': f'请输入 5-{200 if key == "records_per_page" else 100} 的整数'}), 400

    set_user_setting(user_id, key, str(val))
    return jsonify({'ok': True})

@app.route('/admin/settings', methods=['GET', 'POST'])
@admin_required
def admin_settings():
    """系统设置"""
    if request.method == 'POST':
        if not validate_csrf():
            flash('安全验证失败，请刷新页面重试')
            return redirect(url_for('admin_settings'))
        action = request.form.get('action', '')

        if action == 'account':
            new_username = request.form.get('new_username', '').strip()
            new_nickname = request.form.get('new_nickname', '').strip()
            new_email = request.form.get('new_email', '').strip().lower()
            old_pass = request.form.get('old_password', '')
            new_pass = request.form.get('new_password', '')
            confirm_pass = request.form.get('confirm_password', '')

            admin_hash = get_setting('admin_password_hash', '')
            if not check_password_hash(admin_hash, old_pass):
                flash('原密码错误，无法修改账号信息')
            elif not new_username:
                flash('管理员账号不能为空')
            else:
                old_admin_user = get_setting('admin_username', 'admin')
                set_setting('admin_username', new_username)
                # 同步更新 users 表中的管理员记录
                conn2 = get_db()
                updates_sql = "UPDATE users SET username = ?"
                updates_params = [new_username]
                if new_nickname:
                    updates_sql += ", nickname = ?"
                    updates_params.append(new_nickname)
                if new_email:
                    if '@' not in new_email:
                        flash('请输入有效的邮箱地址')
                        conn2.close()
                        return redirect(url_for('admin_settings'))
                    conflict = conn2.execute("SELECT id FROM users WHERE email = ? AND id != (SELECT id FROM users WHERE username = ? AND is_admin = 1 LIMIT 1)",
                                           (new_email, new_username)).fetchone()
                    if conflict:
                        conn2.close()
                        flash('该邮箱已被其他用户使用')
                        return redirect(url_for('admin_settings'))
                    updates_sql += ", email = ?"
                    updates_params.append(new_email)
                updates_sql += " WHERE username = ? AND is_admin = 1"
                updates_params.append(old_admin_user)
                conn2.execute(updates_sql, updates_params)
                conn2.commit()
                conn2.close()
                if new_nickname and session.get('user_id'):
                    session['nickname'] = new_nickname
                if new_pass:
                    if len(new_pass) < 8:
                        flash('新密码至少8位，且需包含字母和数字')
                    elif not re.search(r'[a-zA-Z]', new_pass) or not re.search(r'[0-9]', new_pass):
                        flash('新密码必须同时包含字母和数字')
                    elif new_pass != confirm_pass:
                        flash('两次密码不一致')
                    else:
                        set_setting('admin_password_hash', generate_password_hash(new_pass))
                        # 同步更新 users 表中的管理员密码
                        conn3 = get_db()
                        conn3.execute("UPDATE users SET password_hash = ? WHERE username = ? AND is_admin = 1",
                                      (generate_password_hash(new_pass), new_username))
                        conn3.commit()
                        conn3.close()
                        flash('账号信息修改成功')
                else:
                    flash('账号信息修改成功（密码未变）')

        elif action == 'login_tip':
            tip = request.form.get('login_tip_text', '').strip()
            set_setting('login_tip', tip)
            flash('登录页提示已更新')

        elif action == 'defaults':
            max_files = request.form.get('default_max_files', str(DEFAULT_MAX_FILES))
            max_size = request.form.get('default_max_size', str(DEFAULT_MAX_FILE_SIZE_GB))
            upload_batch = request.form.get('upload_batch_limit', '30')

            try:
                _mf = float(max_files)
                if _mf != int(_mf):
                    raise ValueError('默认最大文件数必须为整数')
                max_files = int(_mf)
                max_size = round(float(max_size), 6)
                upload_batch = int(upload_batch)
                if upload_batch < 5 or upload_batch > 100:
                    raise ValueError('单次上传个数必须在 5-100 之间')
                if max_files < 0:
                    raise ValueError('默认最大文件数不能为负数')
                if max_size < 0.01 or max_size > 64:
                    raise ValueError('单文件上限必须在 0.01-64 GB 之间')
            except ValueError as e:
                flash(str(e) if '必须' in str(e) or '必须在' in str(e) else '默认值格式错误')
                return redirect(url_for('admin_settings'))

            set_setting('max_files', str(max_files))
            set_setting('max_file_size_gb', str(max_size))
            set_setting('upload_batch_limit', str(upload_batch))

            # 通行证有效期
            passcode_ttl = request.form.get('passcode_ttl_minutes', '120')
            try:
                _pt = float(passcode_ttl)
                if _pt != int(_pt):
                    raise ValueError('通行证有效期必须为整数')
                p_val = int(_pt)
                if p_val < 1 or p_val > 43200:
                    raise ValueError('通行证有效期必须在 1-43200 分钟之间')
            except ValueError as e:
                flash(str(e) if '必须' in str(e) else '通行证有效期格式错误')
                return redirect(url_for('admin_settings'))
            set_setting('passcode_ttl_minutes', str(p_val))

            # 链接有效期
            link_expire = request.form.get('default_link_expire_days', '30')
            try:
                _le = float(link_expire)
                if _le != int(_le):
                    raise ValueError('链接有效期天数必须为整数')
                l_val = int(_le)
                if l_val < 1 or l_val > 3650:
                    raise ValueError('链接有效期天数必须在 1-3650 天之间')
            except ValueError as e:
                flash(str(e) if '必须' in str(e) else '链接有效期格式错误')
                return redirect(url_for('admin_settings'))
            set_setting('default_link_expire_days', str(l_val))

            # 禁止文件类型
            blocked = request.form.get('blocked_extensions_input', '').strip()
            if blocked:
                import re as _re2
                cleaned = _re2.sub(r'\s+', ' ', blocked).strip()
                set_setting('blocked_extensions', cleaned)
            else:
                set_setting('blocked_extensions', '')

            # 收集附件上限
            att_max = request.form.get('attachment_max_mb', str(DEFAULT_ATTACHMENT_MAX_MB))
            try:
                _am = float(att_max)
                if _am < 0.1:
                    raise ValueError('收集附件上限最小为 0.1 MB')
                set_setting('attachment_max_mb', str(round(_am, 1)))
            except ValueError as e:
                flash(str(e) if '最小' in str(e) else '附件上限格式错误')
                return redirect(url_for('admin_settings'))

            # 收集链接分页 + 上传记录分页
            raw_links = request.form.get('links_per_page_val', '10')
            try:
                _v = float(raw_links)
                if _v != int(_v):
                    raise ValueError('收集链接每页数量必须为整数')
                val_l = int(_v)
                if val_l < 5 or val_l > 100:
                    raise ValueError('收集链接每页数量必须在 5-100 之间')
            except ValueError as e:
                flash(str(e) if '必须' in str(e) else '每页数量格式错误')
                return redirect(url_for('admin_settings'))
            set_setting('links_per_page', str(val_l))
            raw_records = request.form.get('records_per_page_val', '50')
            try:
                _v2 = float(raw_records)
                if _v2 != int(_v2):
                    raise ValueError('上传记录每页数量必须为整数')
                val_r = int(_v2)
                if val_r < 5 or val_r > 200:
                    raise ValueError('上传记录每页数量必须在 5-200 之间')
            except ValueError as e:
                flash(str(e) if '必须' in str(e) else '每页数量格式错误')
                return redirect(url_for('admin_settings'))
            set_user_setting(session.get('user_id'), 'records_per_page', str(val_r))

            flash('设置已保存')

        elif action == 'site_title':
            title = request.form.get('site_title', '').strip()
            if not title:
                flash('站点标题不能为空')
                return redirect(url_for('admin_settings'))
            set_setting('site_title', title)
            flash('站点标题已保存')

        elif action == 'share_page':
            share_title = request.form.get('share_page_title', '').strip()
            share_footer = request.form.get('share_footer_text', '').strip()
            set_setting('share_page_title', share_title)
            set_setting('share_footer_text', share_footer)
            flash('分享页设置已保存')

        elif action == 'collect_page':
            footer_text = request.form.get('collect_footer_text', '').strip()
            public_url = request.form.get('public_url', '').strip()
            set_setting('collect_footer_text', footer_text)
            set_setting('public_url', public_url)
            flash('收集页设置已保存')

        elif action == 'landing_page':
            enabled = request.form.get('landing_page_enabled', '0')
            set_setting('landing_page_enabled', enabled)
            flash('首页设置已保存')

        elif action == 'registration':
            allow_reg = request.form.get('allow_registration', '0')
            expire_days = request.form.get('default_invite_expire_days', '7')
            try:
                _d = float(expire_days)
                if _d != int(_d):
                    raise ValueError('有效期必须为整数')
                days = int(_d)
                if days < 1 or days > 365:
                    raise ValueError('有效期必须在 1-365 天之间')
            except ValueError as e:
                flash(str(e) if '必须' in str(e) else '有效期格式错误')
                return redirect(url_for('admin_settings'))
            set_setting('allow_registration', allow_reg)
            set_setting('default_invite_expire_days', str(days))
            flash('用户注册设置已保存')

        elif action == 'passcode_ttl':
            minutes = request.form.get('passcode_ttl_minutes', '120')
            try:
                _m = float(minutes)
                if _m != int(_m):
                    raise ValueError('通行证有效期必须为整数')
                val = int(_m)
                if val < 1 or val > 43200:
                    raise ValueError('通行证有效期必须在 1-43200 分钟之间')
            except ValueError as e:
                flash(str(e) if '必须' in str(e) else '通行证有效期格式错误')
                return redirect(url_for('admin_settings'))
            set_setting('passcode_ttl_minutes', str(val))
            flash('通行证有效期已保存')

        elif action == 'upload_path':
            custom_path = request.form.get('custom_upload_path', '').strip()
            if custom_path:
                # 清空可能存在的子路径
                custom_path = os.path.normpath(custom_path)
                if not os.path.isabs(custom_path):
                    flash('上传路径必须是绝对路径，例如 /vol2/1000/文件收集')
                elif not os.path.exists(custom_path):
                    flash(f'路径不存在: {custom_path}')
                else:
                    old_base = get_upload_base()
                    set_setting('custom_upload_path', custom_path)
                    refresh_upload_base()
                    new_base = UPLOAD_BASE
                    # 确保目录可写
                    try:
                        os.makedirs(UPLOAD_BASE, mode=0o755, exist_ok=True)
                    except PermissionError:
                        pass
                    flash('上传路径已更新。请确认飞牛应用设置中已授权该文件夹的读写权限。')
            else:
                old_base = get_upload_base()
                set_setting('custom_upload_path', '')
                refresh_upload_base()
                new_base = UPLOAD_BASE
                flash('已恢复默认上传路径')
            # 无论新旧路径是否相同，都检查并修复数据库中不匹配的 stored_path
            _migrate_stored_paths(old_base, new_base)

        elif action == 'default_link_expiry':
            raw = request.form.get('default_link_expire_days', '30')
            try:
                _v = float(raw)
                if _v != int(_v):
                    raise ValueError('链接有效期天数必须为整数')
                days = int(_v)
                if days < 1 or days > 3650:
                    raise ValueError('链接有效期天数必须在 1-3650 天之间')
            except ValueError as e:
                flash(str(e) if '必须' in str(e) else '链接有效期格式错误')
                return redirect(url_for('admin_settings'))
            set_setting('default_link_expire_days', str(days))
            flash('默认链接有效期已保存')

        elif action == 'links_per_page':
            raw = request.form.get('links_per_page_val', '50')
            try:
                _v = float(raw)
                if _v != int(_v):
                    raise ValueError('每页显示数量必须为整数')
                val = int(_v)
                if val < 5 or val > 100:
                    raise ValueError('每页显示数量必须在 5-100 之间')
            except ValueError as e:
                flash(str(e) if '必须' in str(e) else '每页数量格式错误')
                return redirect(url_for('admin_settings'))
            set_setting('links_per_page', str(val))
            flash('收集链接显示数量已保存')

        elif action == 'records_per_page':
            raw = request.form.get('records_per_page_val', '50')
            try:
                _v = float(raw)
                if _v != int(_v):
                    raise ValueError('每页显示数量必须为整数')
                val = int(_v)
                if val < 5 or val > 200:
                    raise ValueError('每页显示数量必须在 5-200 之间')
            except ValueError as e:
                flash(str(e) if '必须' in str(e) else '每页数量格式错误')
                return redirect(url_for('admin_settings'))
            set_user_setting(session.get('user_id'), 'records_per_page', str(val))
            flash('上传记录单页数量已保存（仅对自己生效）')

        # ---- 合并卡片 actions ----

        elif action == 'page_branding':
            # 站点标题 + 收集页 + 分享页 + 登录页提示
            title = request.form.get('site_title', '').strip()
            if not title:
                flash('站点标题不能为空')
                return redirect(url_for('admin_settings'))
            set_setting('site_title', title)
            footer_text = request.form.get('collect_footer_text', '').strip()
            set_setting('collect_footer_text', footer_text)
            # 反代运行时 public_url 为只读，不覆盖用户手动设置的值
            if not RPROXY_PM.get_public_url():
                public_url = request.form.get('public_url', '').strip()
                set_setting('public_url', public_url)
            share_title = request.form.get('share_page_title', '').strip()
            share_footer = request.form.get('share_footer_text', '').strip()
            set_setting('share_page_title', share_title)
            set_setting('share_footer_text', share_footer)
            tip = request.form.get('login_tip_text', '').strip()
            set_setting('login_tip', tip)
            flash('页面与品牌设置已保存')

        elif action == 'expiry_settings':
            # 通行证有效期 + 默认链接有效期
            minutes = request.form.get('passcode_ttl_minutes', '120')
            try:
                _m = float(minutes)
                if _m != int(_m):
                    raise ValueError('通行证有效期必须为整数')
                val = int(_m)
                if val < 1 or val > 43200:
                    raise ValueError('通行证有效期必须在 1-43200 分钟之间')
            except ValueError as e:
                flash(str(e) if '必须' in str(e) else '通行证有效期格式错误')
                return redirect(url_for('admin_settings'))
            set_setting('passcode_ttl_minutes', str(val))
            raw = request.form.get('default_link_expire_days', '30')
            try:
                _v = float(raw)
                if _v != int(_v):
                    raise ValueError('链接有效期天数必须为整数')
                days = int(_v)
                if days < 1 or days > 3650:
                    raise ValueError('链接有效期天数必须在 1-3650 天之间')
            except ValueError as e:
                flash(str(e) if '必须' in str(e) else '链接有效期格式错误')
                return redirect(url_for('admin_settings'))
            set_setting('default_link_expire_days', str(days))
            flash('有效期设置已保存')

        elif action == 'feature_toggles':
            # 用户注册 + 首页设置
            allow_reg = request.form.get('allow_registration', '0')
            set_setting('allow_registration', allow_reg)
            enabled = request.form.get('landing_page_enabled', '0')
            set_setting('landing_page_enabled', enabled)
            flash('功能开关已保存')

        elif action == 'smtp_config':
            smtp_fields = ['smtp_host', 'smtp_port', 'smtp_username', 'smtp_password',
                           'smtp_use_tls', 'smtp_from_email', 'smtp_from_name']
            for f in smtp_fields:
                val = request.form.get(f, '').strip()
                set_setting(f, val)
            flash('邮箱配置已保存')
            # 自动同步管理员邮箱到 users 表，确保「忘记密码」功能可用
            admin_email = request.form.get('smtp_from_email', '').strip()
            if admin_email and '@' in admin_email:
                try:
                    conn3 = get_db()
                    # 为管理员账号设置邮箱
                    admin_username = get_setting('admin_username', 'admin')
                    conn3.execute(
                        "UPDATE users SET email = ? WHERE username = ? AND is_admin = 1 AND (email = '' OR email IS NULL)",
                        (admin_email, admin_username)
                    )
                    updated = conn3.execute(
                        "SELECT changes()"
                    ).fetchone()[0]
                    if updated:
                        conn3.commit()
                        logger.info(f"管理员邮箱已自动同步: {admin_email}")
                    conn3.close()
                except Exception as e:
                    logger.error(f"同步管理员邮箱失败: {e}")

        elif action == 'attachment_max':
            raw = request.form.get('attachment_max_mb', str(DEFAULT_ATTACHMENT_MAX_MB))
            try:
                _v = float(raw)
                if _v < 0.1:
                    raise ValueError('附件上限最小为 0.1 MB')
                mb = round(_v, 1)
            except ValueError as e:
                flash(str(e) if '最小' in str(e) else '附件上限格式错误')
                return redirect(url_for('admin_settings'))
            set_setting('attachment_max_mb', str(mb))
            flash('附件大小上限已保存')

        elif action == 'blocked_extensions':
            raw = request.form.get('blocked_extensions_input', '').strip()
            # 留空则使用默认禁止列表
            if raw:
                import re as _re
                cleaned = _re.sub(r'\s+', ' ', raw).strip()
                set_setting('blocked_extensions', cleaned)
                flash('禁止上传的文件类型已更新')
            else:
                set_setting('blocked_extensions', '')
                flash('已恢复默认禁止列表（.exe .php .sh .bat 等危险类型）')

        return redirect(url_for('admin_settings'))

    admin_user = get_setting('admin_username', DEFAULT_ADMIN_USER)
    # 获取管理员邮箱和昵称（从 users 表）
    conn = get_db()
    admin_row = conn.execute("SELECT email, nickname FROM users WHERE username = ? AND is_admin = 1",
                             (admin_user,)).fetchone()
    conn.close()
    admin_email = admin_row['email'] if admin_row else ''
    admin_nickname = admin_row['nickname'] if admin_row else ''
    custom_upload_path = get_setting('custom_upload_path', '')
    defaults = {
        'max_files': get_setting('max_files', str(DEFAULT_MAX_FILES)),
        'max_file_size_gb': get_setting('max_file_size_gb', str(DEFAULT_MAX_FILE_SIZE_GB)),
        'upload_batch_limit': get_setting('upload_batch_limit', '30'),
        'attachment_max_mb': get_setting('attachment_max_mb', str(DEFAULT_ATTACHMENT_MAX_MB)),
        'site_title': get_setting('site_title', '文件收集器'),
        'login_tip': get_setting('login_tip', '默认账户 admin / admin123，请及时修改'),
        'collect_footer_text': get_setting('collect_footer_text', ''),
        'share_page_title': get_setting('share_page_title', ''),
        'share_footer_text': get_setting('share_footer_text', ''),
        'public_url': _resolve_public_url(),
        'landing_page_enabled': get_setting('landing_page_enabled', '1'),
        'passcode_ttl_minutes': get_setting('passcode_ttl_minutes', '120'),
        'blocked_extensions': get_setting('blocked_extensions', ''),
        'allow_registration': get_setting('allow_registration', '0'),
        'default_invite_expire_days': get_setting('default_invite_expire_days', '7'),
        'default_link_expire_days': get_setting('default_link_expire_days', '30'),
        'links_per_page': get_setting('links_per_page', '50'),
        'records_per_page': get_user_setting(session.get('user_id'), 'records_per_page', '50'),
        'smtp_host': get_setting('smtp_host', ''),
        'smtp_port': get_setting('smtp_port', '587'),
        'smtp_username': get_setting('smtp_username', ''),
        'smtp_password': get_setting('smtp_password', ''),
        'smtp_use_tls': get_setting('smtp_use_tls', '1'),
        'smtp_from_email': get_setting('smtp_from_email', ''),
        'smtp_from_name': get_setting('smtp_from_name', '文件收集器'),
    }
    sys_info = {
        'db_path': os.path.dirname(DB_PATH),  # 仅显示目录，不暴露完整文件名
        'upload_base': get_upload_base(),
        'data_dir': os.path.basename(DATA_DIR) or DATA_DIR,
        'port': str(PORT),
    }
    return render_template('admin_settings.html',
        defaults=defaults,
        admin_username=admin_user,
        admin_email=admin_email,
        admin_nickname=admin_nickname,
        custom_upload_path=custom_upload_path,
        sys_info=sys_info,
        rp_status=RPROXY_PM.status(),
        version=VERSION)


# ============================================================
# 内置反向代理路由
# ============================================================

def _rp_get_pm():
    return RPROXY_PM

@app.route('/admin/settings/reverse-proxy')
@admin_required
def admin_reverse_proxy():
    """反向代理配置页面"""
    pm = _rp_get_pm()
    certs = CertManager.get_certs_for_display()
    config = pm.get_config()
    status = pm.status()
    logs = pm.get_logs(50)
    return render_template('admin_reverse_proxy.html',
        certs=certs,
        config=config,
        status=status,
        logs=logs,
        all_certs_json=json.dumps(certs, ensure_ascii=False),
        csrf_token=session.get('csrf_token', ''))


@app.route('/api/reverse-proxy/status')
@admin_required
def api_reverse_proxy_status():
    return jsonify(_rp_get_pm().status())


@app.route('/api/reverse-proxy/certs')
@admin_required
def api_reverse_proxy_certs():
    return jsonify({'certs': CertManager.get_certs_for_display()})


@app.route('/api/reverse-proxy/start', methods=['POST'])
@admin_required
def api_reverse_proxy_start():
    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败'})
    domain = request.form.get('domain', '').strip()
    port = request.form.get('port', '7786').strip()
    gzip_enabled = request.form.get('gzip', '1') == '1'
    hsts_enabled = request.form.get('hsts', '1') == '1'
    timeout = request.form.get('timeout', '600').strip()
    if not domain:
        return jsonify({'success': False, 'message': '请选择域名'})
    try:
        port = int(port)
        timeout = int(timeout)
    except ValueError:
        return jsonify({'success': False, 'message': '端口或超时格式错误'})
    if port < 1 or port > 65535:
        return jsonify({'success': False, 'message': '端口范围 1-65535'})
    if port in (80, 443, 8080):
        return jsonify({'success': False, 'message': f'端口 {port} 已被飞牛系统占用，请使用其他端口（如 7786）'})
    certs = CertManager.load_certs()
    cert_path = None
    key_path = None
    for cert in certs:
        sans = cert.get('san', [])
        if domain in sans or domain == cert.get('domain'):
            cert_path = cert.get('fullchain') or cert.get('certificate')
            key_path = cert.get('privateKey')
            break
    if not cert_path:
        return jsonify({'success': False, 'message': f'未找到域名 {domain} 的证书'})

    success, msg = RPROXY_PM.start(
        domain, port, cert_path, key_path,
        f'http://127.0.0.1:{PORT}',
        gzip_enabled, hsts_enabled, timeout
    )
    return jsonify({'success': success, 'message': msg})


@app.route('/api/reverse-proxy/stop', methods=['POST'])
@admin_required
def api_reverse_proxy_stop():
    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败'})
    success, msg = _rp_get_pm().stop()
    return jsonify({'success': success, 'message': msg})


@app.route('/api/reverse-proxy/reload-cert', methods=['POST'])
@admin_required
def api_reverse_proxy_reload_cert():
    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败'})
    success, msg = _rp_get_pm().reload_cert()
    return jsonify({'success': success, 'message': msg})


@app.route('/api/reverse-proxy/logs')
@admin_required
def api_reverse_proxy_logs():
    return jsonify({'logs': _rp_get_pm().get_logs(200)})


@app.route('/api/reverse-proxy/logs/clear', methods=['POST'])
@admin_required
def api_reverse_proxy_clear_logs():
    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败'})
    _rp_get_pm().clear_logs()
    return jsonify({'success': True})


@app.route('/api/reverse-proxy/logs/export')
@admin_required
def api_reverse_proxy_export_logs():
    from flask import Response
    logs = _rp_get_pm().get_logs(9999)
    content = '\n'.join(
        f"[{l['time']}] {l['level']} {l['msg']}" for l in logs
    )
    return Response(
        content,
        mimetype='text/plain',
        headers={'Content-Disposition': 'attachment; filename=reverse_proxy_logs.txt'}
    )


@app.route('/admin/settings/test-smtp', methods=['POST'])
@admin_required
def test_smtp():
    """测试 SMTP 邮件发送"""
    if not validate_csrf():
        return jsonify({'success': False, 'message': '安全验证失败'})
    to_email = request.form.get('test_email', '').strip()
    if not to_email or '@' not in to_email:
        return jsonify({'success': False, 'message': '请输入有效的测试邮箱地址'})

    config = get_smtp_config()
    if not config['host'] or not config['from_email']:
        return jsonify({'success': False, 'message': '请先配置 SMTP 服务器和发件人邮箱'})

    subject = f'[{config["from_name"]}] 测试邮件'
    body = f'''<!DOCTYPE html>
<html>
<head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f5f7fa;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;">
<table width="100%%" cellpadding="0" cellspacing="0" style="background:#f5f7fa;padding:30px 0;">
<tr><td align="center">
<table width="480" cellpadding="0" cellspacing="0" style="background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.06);">
  <tr><td style="background:#059669;padding:24px 32px;text-align:center;">
    <h2 style="color:#fff;margin:0;font-size:18px;">{config["from_name"]}</h2>
  </td></tr>
  <tr><td style="padding:32px;text-align:center;">
    <div style="width:56px;height:56px;border-radius:50%%;background:#ecfdf5;margin:0 auto 16px;display:flex;align-items:center;justify-content:center;">
      <svg width="28" height="28" viewBox="0 0 24 24" fill="none" stroke="#059669" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><path d="M20 6L9 17l-5-5"/></svg>
    </div>
    <h3 style="margin:0 0 8px;color:#111827;font-size:16px;">邮件发送测试成功！</h3>
    <p style="color:#6b7280;font-size:14px;line-height:1.6;margin:0;">如果您收到此邮件，说明 SMTP 配置正确，邮件功能可以正常使用。</p>
  </td></tr>
  <tr><td style="border-top:1px solid #e5e7eb;padding:16px 32px;text-align:center;">
    <p style="color:#9ca3af;font-size:12px;margin:0;">此邮件由 {config["from_name"]} 系统自动发送</p>
  </td></tr>
</table>
</td></tr>
</table>
</body>
</html>'''
    ok, msg = send_email(to_email, subject, body)
    return jsonify({'success': ok, 'message': msg})

@app.route('/admin/settings/backup-db')
@admin_required
def backup_database():
    """下载数据库备份"""
    if not os.path.exists(DB_PATH):
        flash('数据库文件不存在')
        return redirect(url_for('admin_settings'))
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    download_name = f'file-collector-backup-{timestamp}.db'
    return send_from_directory(
        os.path.dirname(DB_PATH),
        os.path.basename(DB_PATH),
        download_name=download_name,
        as_attachment=True
    )

@app.route('/admin/settings/restore-db', methods=['POST'])
@admin_required
def restore_database():
    """导入数据库备份"""
    if not validate_csrf():
        flash('安全验证失败，请刷新页面重试')
        return redirect(url_for('admin_settings'))
    if 'db_file' not in request.files:
        flash('请选择数据库文件')
        return redirect(url_for('admin_settings'))

    file = request.files['db_file']
    if not file or not file.filename:
        flash('请选择数据库文件')
        return redirect(url_for('admin_settings'))

    if not file.filename.endswith('.db'):
        flash('仅支持 .db 格式的数据库文件')
        return redirect(url_for('admin_settings'))

    # 保存到临时位置并验证
    tmp = tempfile.NamedTemporaryFile(suffix='.db', delete=False)
    tmp_path = tmp.name
    tmp.close()
    try:
        file.save(tmp_path)

        # 验证文件头：SQLite 数据库前 16 字节必须是 "SQLite format 3\0"
        with open(tmp_path, 'rb') as f:
            header = f.read(16)
        if header != b'SQLite format 3\x00':
            flash('无效的数据库文件：文件格式不正确')
            os.unlink(tmp_path)
            return redirect(url_for('admin_settings'))

        # 验证是否为有效的 SQLite 数据库且包含必要表结构
        try:
            test_conn = sqlite3.connect(tmp_path)
            test_conn.row_factory = sqlite3.Row
            tables = test_conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name IN ('settings','links','upload_records')"
            ).fetchall()
            test_conn.close()
            if len(tables) < 3:
                flash('无效的数据库文件：缺少必要的表结构')
                os.unlink(tmp_path)
                return redirect(url_for('admin_settings'))
        except sqlite3.Error:
            flash('无效的数据库文件：无法读取数据库')
            os.unlink(tmp_path)
            return redirect(url_for('admin_settings'))

        # 备份当前数据库
        backup_path = DB_PATH + f'.bak_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        shutil.copy2(DB_PATH, backup_path)

        # 替换数据库
        shutil.copy2(tmp_path, DB_PATH)
        os.unlink(tmp_path)

        # 重新初始化（执行迁移）
        try:
            init_db()
        except Exception as e:
            # 迁移失败，回滚到旧数据库
            logger.error(f"数据库迁移失败，回滚: {e}")
            shutil.copy2(backup_path, DB_PATH)
            flash(f'数据库迁移失败，已自动回滚。原因：{e}')
            return redirect(url_for('admin_settings'))

        refresh_upload_base()  # 恢复后刷新自定义上传路径

        # 轮换 secret_key（仅保存到数据库，下次重启生效）
        # 注意：不能在运行时改 app.secret_key，因为多 gunicorn worker 下只有当前 worker 会更新，
        #       其他 worker 仍是旧 key，导致 cookie 跨 worker 解密失败 → "安全验证失败"
        new_secret = secrets.token_hex(32)
        set_setting('secret_key', new_secret)

        # 检查还原后的数据库中的路径是否安全（在上传目录范围内）
        conn = get_db()
        rows = conn.execute("SELECT id, stored_path FROM upload_records").fetchall()
        bad_paths = 0
        upload_base = get_upload_base()
        real_base = os.path.realpath(upload_base)
        for row in rows:
            try:
                real_path = os.path.realpath(row['stored_path'])
                if not (real_path.startswith(real_base + os.sep) or real_path == real_base):
                    logger.warning(f"还原数据库发现不安全路径: {row['stored_path']}")
                    bad_paths += 1
            except Exception:
                bad_paths += 1
        conn.close()

        flash('数据库已成功导入！Secret Key 已轮换（重启后生效）。')
        if bad_paths > 0:
            flash(f'警告：发现 {bad_paths} 条记录的文件路径不在上传目录中，已跳过删除保护。')

        if bad_paths > 0:
            flash(f'警告：发现 {bad_paths} 条记录的文件路径不在上传目录中，已跳过删除保护。')
    except Exception as e:
        logger.error(f"数据库还原失败: {e}")
        flash(f'导入失败，请检查文件是否正确')
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

    return redirect(url_for('admin_settings'))

@app.route('/logout')
def logout():
    """退出登录"""
    session.clear()
    return redirect(url_for('admin_login'))

# ============================================================
# API 路由
# ============================================================
@app.route('/api/status')
def api_status():
    """应用状态检查"""
    conn = get_db()
    try:
        conn.execute("SELECT 1 FROM settings LIMIT 1")
        db_ok = True
    except Exception:
        db_ok = False
    conn.close()

    return jsonify({
        'status': 'running',
        'db_ok': db_ok,
        'upload_dir_exists': os.path.isdir(UPLOAD_BASE),
    })

# ============================================================
# 后台文件监控线程
# ============================================================
_bg_scanner_lock = threading.Lock()
_bg_scanner_state = {
    'running': False,
    'last_scan': None,
    'last_scan_duration': 0,
    'total_new_files': 0,
    'total_orphans_cleaned': 0,
    'error': None,
}

def _bg_file_scanner(interval=30):
    """后台文件扫描线程：定期扫描所有链接文件夹的新文件并清理孤儿记录
    使用 settings 表中的时间戳锁防止多 worker 重复扫描。"""
    # 启动时随机延迟，避免多个 worker 同时竞争
    import random
    time.sleep(random.uniform(0, 3))
    
    _bg_scanner_state['running'] = True
    
    while _bg_scanner_state['running']:
        try:
            # 检查是否需要扫描
            conn = get_db()
            last_val = conn.execute(
                "SELECT value FROM settings WHERE key = 'last_bg_scan'"
            ).fetchone()
            
            now = time.time()
            if last_val:
                try:
                    last_ts = float(last_val['value'])
                    if now - last_ts < interval * 0.8:
                        conn.close()
                        time.sleep(interval * 0.3)
                        continue
                except ValueError:
                    pass
            
            # 乐观锁：更新扫描时间戳来获取锁
            try:
                conn.execute(
                    "INSERT OR REPLACE INTO settings (key, value) VALUES ('last_bg_scan', ?)",
                    (str(now),)
                )
                conn.commit()
            except Exception:
                conn.close()
                time.sleep(interval * 0.3)
                continue
            
            # 执行扫描
            start_time = time.time()
            
            links = conn.execute("SELECT id FROM links").fetchall()
            total_new = 0
            for link in links:
                new_files = scan_link_folder(link['id'], conn)
                total_new += len(new_files)
            
            total_orphan = cleanup_orphan_records(conn)
            
            conn.close()
            
            elapsed = round(time.time() - start_time, 2)
            
            with _bg_scanner_lock:
                _bg_scanner_state['last_scan'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                _bg_scanner_state['last_scan_duration'] = elapsed
                _bg_scanner_state['total_new_files'] += total_new
                _bg_scanner_state['total_orphans_cleaned'] += total_orphan
                _bg_scanner_state['error'] = None
            
            if total_new > 0 or total_orphan > 0:
                logger.info(f"后台扫描: +{total_new} 新文件, -{total_orphan} 孤儿, 耗时 {elapsed}s")
                
        except Exception as e:
            logger.error(f"后台扫描异常: {e}")
            with _bg_scanner_lock:
                _bg_scanner_state['error'] = str(e)[:200]
        
        time.sleep(interval)

def start_bg_scanner(interval=30):
    """启动后台文件扫描线程"""
    thread = threading.Thread(target=_bg_file_scanner, args=(interval,), daemon=True)
    thread.start()
    return thread

@app.route('/api/scan-status')
@login_required
def api_scan_status():
    """获取后台扫描状态（前端轮询用）"""
    with _bg_scanner_lock:
        state = dict(_bg_scanner_state)
    return jsonify(state)

@app.route('/favicon.ico')
def favicon_root():
    """根路径 /favicon.ico 重定向到静态目录"""
    return redirect(url_for('static', filename='favicon.ico'))


# 检查更新缓存（5分钟内不重复请求）
_check_update_cache = {'last_time': 0, 'result': None}
_CHECK_UPDATE_CACHE_SECONDS = 300


@app.route('/api/check-update')
@admin_required
def check_update():
    """检查 GitHub Releases 是否有新版本（5分钟内返回缓存结果）"""
    now = time.time()
    if now - _check_update_cache['last_time'] < _CHECK_UPDATE_CACHE_SECONDS \
            and _check_update_cache['result'] is not None:
        return jsonify(_check_update_cache['result'])

    try:
        req = urllib.request.Request(
            'https://api.github.com/repos/Contribuv/file-collector/releases',
            headers={
                'User-Agent': 'file-collector/' + VERSION,
                'Accept': 'application/vnd.github.v3+json',
            }
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            import json
            releases = json.loads(resp.read().decode('utf-8'))

        if not releases:
            result = {'has_update': False}
            _check_update_cache['last_time'] = now
            _check_update_cache['result'] = result
            return jsonify(result)

        # 从最新 release 提取版本号
        latest = releases[0]
        tag = latest.get('tag_name', '').lstrip('v')
        html_url = latest.get('html_url', 'https://github.com/Contribuv/file-collector/releases')

        def parse_version(v):
            parts = v.split('.')
            return tuple(int(p) for p in parts if p.isdigit())

        try:
            if parse_version(tag) > parse_version(VERSION):
                result = {'has_update': True, 'latest_version': tag, 'url': html_url}
            else:
                result = {'has_update': False}
        except (ValueError, IndexError):
            result = {'has_update': False}

        _check_update_cache['last_time'] = now
        _check_update_cache['result'] = result
        return jsonify(result)
    except Exception as e:
        logger.warning(f"检查更新失败: {e}")
        return jsonify({'has_update': False, 'error': '检查更新失败'})

# ============================================================
# 错误处理
# ============================================================
@app.errorhandler(500)
def internal_error(e):
    """500 错误处理 - API 路由返回 JSON"""
    logger.error(f"500 错误: {request.path}\n{traceback.format_exc()}")
    # 对 collect/api 路由返回 JSON
    if '/collect/' in request.path or '/api/' in request.path:
        return jsonify({'success': False, 'message': '服务器内部错误，请稍后重试'}), 500
    return render_template('error.html',
        error_code=500,
        error_title='服务器内部错误',
        error_message='服务器遇到内部错误，无法完成您的请求。'), 500

@app.errorhandler(404)
def page_not_found(e):
    return render_template('error.html',
        error_code=404,
        error_title='页面未找到',
        error_message='您访问的页面不存在或已被移除。'), 404

@app.errorhandler(410)
def page_gone(e):
    return render_template('error.html',
        error_code=410,
        error_title='资源已失效',
        error_message='该资源已过期或不再可用。'), 410


# ============================================================
# WSGI 入口（供 Gunicorn 调用）
# ============================================================
# app 对象已在文件顶部创建，Gunicorn 通过 `app:app` 加载

if __name__ == '__main__':
    import multiprocessing
    import gunicorn.app.base

    cpu_count = multiprocessing.cpu_count()
    workers = min(cpu_count + 1, 4)

    class GunicornApp(gunicorn.app.base.BaseApplication):
        def __init__(self, app, options=None):
            self.application = app
            self.options = options or {}
            super().__init__()

        def load_config(self):
            for k, v in self.options.items():
                if k in self.cfg.settings and v is not None:
                    self.cfg.set(k.lower(), v)

        def load(self):
            return self.application

    # Gunicorn worker 启动后，启动后台文件扫描线程
    def post_worker_init(worker):
        # 启动时自动修复可能不匹配的 stored_path（如管理员修改了上传路径后重启）
        current_base = get_upload_base()
        _migrate_stored_paths(current_base, current_base)
        start_bg_scanner(interval=30)

    options = {
        'bind': f'0.0.0.0:{PORT}',
        'workers': workers,
        'timeout': 0,  # 不限制超时，MAX_CONTENT_LENGTH 和频率限制已提供保护
        'accesslog': '-',
        'errorlog': '-',
        'loglevel': 'info',
        'post_worker_init': post_worker_init,
    }

    logger.info(f"文件收集器 v{VERSION} 启动中 (Gunicorn)...")
    logger.info(f"数据目录: {DATA_DIR}")
    logger.info(f"上传目录: {UPLOAD_BASE}")
    logger.info(f"监听端口: {PORT}")
    logger.info(f"Worker 进程: {workers}")
    logger.info(f"管理后台: http://localhost:{PORT}/admin")

    GunicornApp(app, options).run()